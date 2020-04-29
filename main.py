﻿#!/usr/bin/env python
# -*- coding: utf-8 -*-
import collections

from utils.min_heap import MinHeap
from utils.util import format_time, _from_rgb



## 코드를 무단으로 복제하여 개조 및 배포하지 말 것##

import itertools
import os
import threading
import time
import tkinter.font
import tkinter.messagebox
import tkinter.ttk
import webbrowser
from collections import Counter
from math import floor
from tkinter import *

import numpy as np
from openpyxl import load_workbook
now_version = "v3.2.8.1"
ver_time = '2020-04-26'



# copy from https://gist.github.com/bakineugene/76c8f9bcec5b390e45df
# http://tkinter.unpythonic.net/wiki/VerticalScrolledFrame

class VerticalScrolledFrame(Frame):
    """A pure Tkinter scrollable frame that actually works!
    * Use the 'interior' attribute to place widgets inside the scrollable frame
    * Construct and pack/place/grid normally
    * This frame only allows vertical scrolling
    """

    def __init__(self, parent, *args, **kw):
        Frame.__init__(self, parent, *args, **kw)

        # create a canvas object and a vertical scrollbar for scrolling it
        vscrollbar = Scrollbar(self, orient=VERTICAL)
        vscrollbar.pack(fill=Y, side=RIGHT, expand=FALSE)
        canvas = Canvas(self, bd=0, highlightthickness=0,
                        yscrollcommand=vscrollbar.set)
        canvas.pack(side=LEFT, fill=BOTH, expand=TRUE)
        vscrollbar.config(command=canvas.yview)

        # reset the view
        canvas.xview_moveto(0)
        canvas.yview_moveto(0)

        # create a frame inside the canvas which will be scrolled with it
        self.interior = interior = Frame(canvas)
        interior_id = canvas.create_window(0, 0, window=interior,
                                           anchor=NW)

        # track changes to the canvas and frame width and sync them,
        # also updating the scrollbar
        def _configure_interior(event):
            # update the scrollbars to match the size of the inner frame
            size = (interior.winfo_reqwidth(), interior.winfo_reqheight())
            canvas.config(scrollregion="0 0 %s %s" % size)
            if interior.winfo_reqwidth() != canvas.winfo_width():
                # update the canvas's width to fit the inner frame
                canvas.config(width=interior.winfo_reqwidth())

        interior.bind('<Configure>', _configure_interior)

        def _configure_canvas(event):
            if interior.winfo_reqwidth() != canvas.winfo_width():
                # update the inner frame's width to fill the canvas
                canvas.itemconfigure(interior_id, width=canvas.winfo_width())

        canvas.bind('<Configure>', _configure_canvas)


# Combopicker code from @SilverHalo https://stackoverflow.com/questions/34549752/how-do-i-enable-multiple-selection-of-values-from-a-combobox
class Picker(tkinter.ttk.Frame):
    def __init__(self, master=None, activebackground='#b1dcfb', dict_intvar_item={}, values=[], entry_wid=None, activeforeground='black', selectbackground='#003eff', selectforeground='white', command=None, borderwidth=1, relief="solid"):
        self._selected_item = None

        self._values = values

        self._entry_wid = entry_wid

        self._sel_bg = selectbackground
        self._sel_fg = selectforeground

        self._act_bg = activebackground
        self._act_fg = activeforeground

        self._command = command
        tkinter.ttk.Frame.__init__(self, master, borderwidth=borderwidth, relief=relief)

        self.bind("<FocusIn>", lambda event: self.event_generate('<<PickerFocusIn>>'))
        self.bind("<FocusOut>", lambda event: self.event_generate('<<PickerFocusOut>>'))

        self._font = tkinter.font.Font()

        self.frame = VerticalScrolledFrame(self)
        self.frame.pack()

        self.dict_checkbutton = {}
        self.dict_checkbutton_var = {}
        self.dict_intvar_item = dict_intvar_item

        for index, item in enumerate(self._values):
            self.dict_intvar_item[item] = tkinter.IntVar()
            self.dict_checkbutton[item] = tkinter.ttk.Checkbutton(self.frame.interior, text=item, variable=self.dict_intvar_item[item], command=lambda ITEM=item: self._command(ITEM))
            self.dict_checkbutton[item].grid(row=index, column=0, sticky=tkinter.NSEW)
            self.dict_intvar_item[item].set(0)


class Combopicker(tkinter.ttk.Entry, Picker):
    def __init__(self, master, values=[], entryvar=None, entrywidth=None, entrystyle=None, onselect=None, activebackground='#b1dcfb', activeforeground='black', selectbackground='#003eff', selectforeground='white', borderwidth=1,
                 relief="solid"):

        if entryvar is not None:
            self.entry_var = entryvar
        else:
            self.entry_var = tkinter.StringVar()

        self.dict_intvar_item = {}

        entry_config = {}
        if entrywidth is not None:
            entry_config["width"] = entrywidth

        if entrystyle is not None:
            entry_config["style"] = entrystyle

        tkinter.ttk.Entry.__init__(self, master, textvariable=self.entry_var, **entry_config, state="readonly")

        self._is_menuoptions_visible = False

        self.picker_frame = Picker(self.winfo_toplevel(), dict_intvar_item=self.dict_intvar_item, values=values, entry_wid=self.entry_var, activebackground=activebackground, activeforeground=activeforeground,
                                   selectbackground=selectbackground, selectforeground=selectforeground,
                                   command=self._on_selected_check)

        self.bind_all("<1>", self._on_click, "+")

        self.bind("<Escape>", lambda event: self.hide_picker())

    @property
    def current_value(self):
        try:
            value = self.entry_var.get()
            return value
        except ValueError:
            return None

    def get_selected_entrys(self):
        return self.entry_var.get().split(",")

    @current_value.setter
    def current_value(self, INDEX):
        self.entry_var.set(self.values.index(INDEX))

    def set(self, checked_values):
        # 清空所有选项
        for item, intvar in self.dict_intvar_item.items():
            intvar.set(0)
        # 清空内容栏
        self.entry_var.set("")

        if checked_values is None:
            return

        # 添加选项与内容
        temp_value = ""
        for item in checked_values:
            try:
                self.dict_intvar_item[item].set(1)
                if len(temp_value) != 0:
                    temp_value += ","
                temp_value += str(item)
            except:
                pass

        self.entry_var.set(temp_value)

    def _on_selected_check(self, SELECTED):

        value = []
        if self.entry_var.get() != "" and self.entry_var.get() != None:
            temp_value = self.entry_var.get()
            value = temp_value.split(",")

        if str(SELECTED) in value:
            value.remove(str(SELECTED))

        else:
            value.append(str(SELECTED))

        value.sort()

        temp_value = ""
        for index, item in enumerate(value):
            if item != "":
                if index != 0:
                    temp_value += ","
                temp_value += str(item)

        self.entry_var.set(temp_value)

    def _on_click(self, event):
        str_widget = str(event.widget)

        if str_widget == str(self):
            if not self._is_menuoptions_visible:
                self.show_picker()
        else:
            if not str_widget.startswith(str(self.picker_frame)) and self._is_menuoptions_visible:
                self.hide_picker()

    def show_picker(self):
        if not self._is_menuoptions_visible:
            self.picker_frame.place(in_=self, relx=0, rely=1, relwidth=1)
            self.picker_frame.lift()

        self._is_menuoptions_visible = True

    def hide_picker(self):
        if self._is_menuoptions_visible:
            self.picker_frame.place_forget()

        self._is_menuoptions_visible = False


###########################################################
#                         逻辑相关常量                     #
###########################################################

# 可升级得到的工作服列表
work_uniforms = [
    "11150", "12150", "13150", "14150", "15150",  # 工作服防具：大自然
    "21190", "22190", "23190",  # 工作服首饰：权能
    "31230", "32230", "33230",  # 工作服特殊装备：能量
]
# 智慧产物列表
the_product_of_wisdoms = [
    "13390150", "22390240", "23390450", "33390750", "21400340", "31400540", "32410650",
]

# 自定义存档的列定义
g_col_custom_save_key = 14
g_col_custom_save_value_begin = 15

# 自定义存档的行定义
g_row_custom_save_save_name = 1  # 存档名
g_row_custom_save_weapon = 2  # 武器
g_row_custom_save_job = 3  # 职业选择
g_row_custom_save_fight_time = 4  # 输出时间
g_row_custom_save_title = 5  # 称号选择
g_row_custom_save_pet = 6  # 宠物选择
g_row_custom_save_cd = 7  # 冷却补正
g_row_custom_save_speed = 8  # 选择速度
g_row_custom_save_has_baibianguai = 9  # 是否拥有百变怪
g_row_custom_save_can_upgrade_work_uniforms_nums = 10  # 当前拥有的材料够升级多少件工作服
g_row_custom_save_transfer_from = 11  # 跨界来源账户列表
g_row_custom_save_max_transfer_count = 12  # 最大跨界数目

# 国服特色
styles = [
    '使徒降临', '伟大的意志',  # 2020春节普通称号和至尊称号
    '超越极限者',  # 2019国庆称号
    '秘境迷踪', '神选之英杰',  # 2019春节普通称号和至尊称号
    '神之试炼的奖赏',  # 2018国庆称号
    '兽人守护神', '天选之人',  # 2018春节普通称号和至尊称号
    '海洋霸主',  # 2017国庆称号
    '龙之挑战', '龙之威仪',  # 2017春节普通称号和至尊称号
    '最强战神',  # 心悦称号
    '与贝奇邂逅',  # 我的奶妈在用的称号，方便自己用
    '其他（直接比较）'
]
creatures = [
    '弓手维多利亚', '神官格洛丽亚', '雷光之箭维多利亚', '暴风圣女格洛丽亚',  # 2020春节普通宠物和至尊宠物
    '骑士莱恩', '吟游诗人薇泽达', '古国英豪莱恩', '太初之音薇泽达',  # 2019春节普通宠物和至尊宠物
    '雪兔蒂娅', '火狐艾芙', '冰雪魔法师蒂娅', '炽焰咒术师艾芙',  # 2018春节普通宠物和至尊宠物
    '艾莉丝', '克里斯',  # 2017春节普通宠物
    '其他（直接比较）'
]

# 各个词条加成在base_array中的下标        # data.xlsx的one sheet字段说明
index_strength_and_intelligence = 0  # 0-C-stat-力智
index_physical_magical_independent_attack_power = 1  # 1-D-att-物理/魔法/独立攻击力
index_extra_percent_attack_damage = 2  # 2-E-damper-攻击时额外增加X%的伤害增加量
index_extra_percent_crit_damage = 3  # 3-F-criper-暴击时，额外增加X%的伤害增加量
index_extra_percent_addtional_damage = 4  # 4-G-bonper-攻击时，附加X%的伤害，也就是白字
index_extra_percent_elemental_damage = 5  # 5-H-elebon-攻击时，附加X%的属性伤害
index_extra_percent_final_damage = 6  # 6-I-allper-最终伤害+X%
index_extra_percent_physical_magical_independent_attack_power = 7  # 7-J-attper-物理/魔法/独立攻击力 +X%
index_extra_percent_strength_and_intelligence = 8  # 8-K-staper-力智+X%
index_extra_all_element_strength = 9  # 9-L-ele-所有属性强化+X
index_extra_percent_continued_damage = 10  # 10-M-sloper-发生持续伤害5秒，伤害量为对敌人造成伤害的X%
index_extra_percent_skill_attack_power = 11  # 11-N-skiper-技能攻击力 +X%
index_extra_percent_special_effect = 12  # 12-O-special-特殊词条补正，如歧路和不息的装备，详见自定义中这俩装备相关配置
index_extra_percent_attack_speed = 13  # 13-P-speed-攻击速度 +X%
index_extra_percent_magic_physical_crit_rate = 14  # 14-Q-critical-魔法/物理暴击率 +X%
index_extra_active_skill_effect = 15  # 15-R-active-主动技能增加等级所带来的的影响（目前C的伤害计算没有计入该值，仅奶系职业用到）
index_extra_passive_transfer_skill = 16  # 16-S-pas1-增加转职被动的等级
index_extra_passive_first_awaken_skill = 17  # 17-T-pas2-增加一绝被动的等级
index_extra_passive_second_awaken_skill = 18  # 18-U-pas3-增加二觉被动的等级
index_extra_passive_third_awaken_skill = 19  # 19-V-pas4-增加三觉被动的等级
index_cool_correction = 20  # 20-Y-cool_skill-冷却矫正系数，每冷却1%，记0.35这个值
index_extra_active_second_awaken_skill = 21  # 21-AK-active2-二觉主动技能
index_extra_active_skill_lv_1_45 = 22  # 22-AO-pas0-1_45主动技能
index_extra_active_skill_lv_50 = 23  # 23-AP-pas1-50主动技能
index_extra_active_skill_lv_60_80 = 24  # 24-AQ-pas2-60_80主动技能
index_extra_active_skill_lv_85 = 25  # 25-AR-pas3-85主动技能
index_extra_active_skill_lv_95 = 26  # 26-AS-pas4-95主动技能
index_extra_active_skill_lv_100 = 27  # 27-AT-pas5-100主动技能


###########################################################
#                         逻辑相关函数                     #
###########################################################

# 获取国服特殊加成属性
def get_shuchu_bonus_attributes():
    bonus_array = np.array([0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
    # 获取称号的加成
    style = style_select.get()
    if style == '使徒降临':
        bonus_array[index_physical_magical_independent_attack_power] += 60  # 三攻+60
        bonus_array[index_strength_and_intelligence] += 80  # 四维+80
        bonus_array[index_extra_percent_attack_speed] += 3  # 攻击速度+3%
        bonus_array[index_extra_all_element_strength] += 15  # 所有属性强化+15
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 12  # 物理、魔法暴击率+12%
        bonus_array[index_extra_percent_addtional_damage] += 12  # 攻击时，附加12%的伤害
        bonus_array[index_extra_percent_strength_and_intelligence] += 3  # 增加3%的力量、智力
        bonus_array[index_strength_and_intelligence] += 0.03 * 35 * 20 / 30  # 攻击时，3%概率增加35点力量，持续20s，冷却30s
        bonus_array[
            index_extra_percent_magic_physical_crit_rate] += 0.05 * 5 * 20 / 30  # 释放技能时，5%概率增加5%物理、魔法暴击率，持续20s，冷却30s
    elif style == '伟大的意志':
        bonus_array[index_physical_magical_independent_attack_power] += 65  # 三攻+65
        bonus_array[index_strength_and_intelligence] += 90  # 四维+90
        bonus_array[index_extra_percent_attack_speed] += 4  # 攻击速度+4%
        bonus_array[index_extra_all_element_strength] += 20  # 所有属性强化+20
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 15  # 物理、魔法暴击率+15%
        bonus_array[index_extra_percent_strength_and_intelligence] += 4  # 增加4%的力量、智力
        bonus_array[index_extra_percent_crit_damage] += 18  # 暴击时，额外增加18%的伤害增加量
        bonus_array[index_strength_and_intelligence] += 0.03 * 35 * 20 / 30  # 攻击时，3%概率增加35点力量，持续20s，冷却30s
        bonus_array[
            index_extra_percent_magic_physical_crit_rate] += 0.05 * 5 * 20 / 30  # 释放技能时，5%概率增加5%物理、魔法暴击率，持续20s，冷却30s
    elif style == '超越极限者':
        bonus_array[index_strength_and_intelligence] += 60  # 四维+60
        bonus_array[index_extra_percent_attack_speed] += 3  # 攻击速度+3%
        bonus_array[index_extra_all_element_strength] += 15  # 所有属性强化+15
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 12  # 物理、魔法暴击率+12%
        bonus_array[index_extra_percent_crit_damage] += 15  # 暴击时，额外增加15%的伤害增加量
        bonus_array[index_extra_all_element_strength] += 0.03 * 10 * 30 / 40  # 攻击时，3%概率增加10点属强，持续30s，冷却40s
        bonus_array[index_extra_percent_attack_speed] += 0.03 * 3 * 30 / 40  # 攻击时，3%概率增加3%三速，持续30s，冷却40s
    elif style == '秘境迷踪':
        bonus_array[index_physical_magical_independent_attack_power] += 40  # 三攻+40
        bonus_array[index_strength_and_intelligence] += 70  # 四维+70
        bonus_array[index_extra_percent_attack_speed] += 3  # 攻击速度+3%
        bonus_array[index_extra_all_element_strength] += 15  # 所有属性强化+15
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 12  # 物理、魔法暴击率+12%
        bonus_array[index_extra_percent_addtional_damage] += 10  # 攻击时，附加10%的伤害
        bonus_array[index_strength_and_intelligence] += 0.03 * 35 * 20 / 30  # 攻击时，3%概率增加35点力量，持续20s，冷却30s
        bonus_array[
            index_extra_percent_magic_physical_crit_rate] += 0.05 * 5 * 20 / 30  # 释放技能时，5%概率增加5%物理、魔法暴击率，持续20s，冷却30s
    elif style == '神选之英杰':
        bonus_array[index_physical_magical_independent_attack_power] += 45  # 三攻+45
        bonus_array[index_strength_and_intelligence] += 75  # 四维+75
        bonus_array[index_extra_percent_attack_speed] += 4  # 攻击速度+4%
        bonus_array[index_extra_all_element_strength] += 20  # 所有属性强化+20
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 15  # 物理、魔法暴击率+15%
        bonus_array[index_extra_percent_crit_damage] += 18  # 暴击时，额外增加18%的伤害增加量
        bonus_array[index_strength_and_intelligence] += 0.03 * 35 * 20 / 30  # 攻击时，3%概率增加35点力量，持续20s，冷却30s
        bonus_array[
            index_extra_percent_magic_physical_crit_rate] += 0.05 * 6 * 20 / 30  # 释放技能时，5%概率增加6%物理、魔法暴击率，持续20s，冷却30s
    elif style == '神之试炼的奖赏':
        bonus_array[index_strength_and_intelligence] += 55  # 四维+55
        bonus_array[index_extra_percent_attack_speed] += 3  # 攻击速度+3%
        bonus_array[index_extra_all_element_strength] += 15  # 所有属性强化+15
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 10  # 物理、魔法暴击率+10%
        bonus_array[index_extra_percent_crit_damage] += 15  # 暴击时，额外增加15%的伤害增加量
        bonus_array[index_extra_all_element_strength] += 0.03 * 10 * 30 / 40  # 攻击时，3%概率增加10点属强，持续30s，冷却40s
        bonus_array[index_extra_percent_attack_speed] += 0.03 * 3 * 30 / 40  # 攻击时，3%概率增加3%三速，持续30s，冷却40s
    elif style == '兽人守护神':
        bonus_array[index_physical_magical_independent_attack_power] += 30  # 三攻+30
        bonus_array[index_strength_and_intelligence] += 70  # 四维+70
        bonus_array[index_extra_percent_attack_speed] += 3  # 攻击速度+3%
        bonus_array[index_extra_all_element_strength] += 15  # 所有属性强化+15
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 12  # 物理、魔法暴击率+12%
        bonus_array[index_extra_percent_addtional_damage] += 10  # 攻击时，附加10%的伤害
        bonus_array[index_strength_and_intelligence] += 0.03 * 35 * 20 / 30  # 攻击时，3%概率增加35点力量，持续20s，冷却30s
        bonus_array[
            index_extra_percent_magic_physical_crit_rate] += 0.05 * 5 * 20 / 30  # 释放技能时，5%概率增加5%物理、魔法暴击率，持续20s，冷却30s
    elif style == '天选之人':
        bonus_array[index_physical_magical_independent_attack_power] += 35  # 三攻+35
        bonus_array[index_strength_and_intelligence] += 75  # 四维+75
        bonus_array[index_extra_percent_attack_speed] += 4  # 攻击速度+4%
        bonus_array[index_extra_all_element_strength] += 20  # 所有属性强化+20
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 15  # 物理、魔法暴击率+15%
        bonus_array[index_extra_percent_final_damage] += 12  # 最终伤害增加12%
        bonus_array[index_strength_and_intelligence] += 0.03 * 35 * 20 / 30  # 攻击时，3%概率增加35点力量，持续20s，冷却30s
        bonus_array[
            index_extra_percent_magic_physical_crit_rate] += 0.05 * 6 * 20 / 30  # 释放技能时，5%概率增加6%物理、魔法暴击率，持续20s，冷却30s
    elif style == '海洋霸主':
        bonus_array[index_strength_and_intelligence] += 55  # 四维+55
        bonus_array[index_extra_percent_attack_speed] += 3  # 攻击速度+3%
        bonus_array[index_extra_all_element_strength] += 15  # 所有属性强化+15
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 10  # 物理、魔法暴击率+10%
        bonus_array[index_extra_percent_crit_damage] += 10  # 暴击时，额外增加10%的伤害增加量
        bonus_array[index_extra_all_element_strength] += 0.03 * 10 * 30 / 40  # 攻击时，3%概率增加10点属强，持续30s，冷却40s
        bonus_array[index_extra_percent_attack_speed] += 0.03 * 3 * 30 / 40  # 攻击时，3%概率增加3%三速，持续30s，冷却40s
    elif style == '龙之挑战':
        bonus_array[index_physical_magical_independent_attack_power] += 30  # 三攻+30
        bonus_array[index_strength_and_intelligence] += 60  # 四维+60
        bonus_array[index_extra_percent_attack_speed] += 3  # 攻击速度+3%
        bonus_array[index_extra_all_element_strength] += 15  # 所有属性强化+15
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 10  # 物理、魔法暴击率+10%
        bonus_array[index_extra_percent_addtional_damage] += 10  # 攻击时，附加10%的伤害
        bonus_array[index_strength_and_intelligence] += 0.03 * 35 * 20 / 30  # 攻击时，3%概率增加35点力量，持续20s，冷却30s
        bonus_array[
            index_extra_percent_magic_physical_crit_rate] += 0.05 * 5 * 20 / 30  # 释放技能时，5%概率增加5%物理、魔法暴击率，持续20s，冷却30s
    elif style == '龙之威仪':
        bonus_array[index_physical_magical_independent_attack_power] += 35  # 三攻+35
        bonus_array[index_strength_and_intelligence] += 65  # 四维+65
        bonus_array[index_extra_percent_attack_speed] += 3  # 攻击速度+3%
        bonus_array[index_extra_all_element_strength] += 15  # 所有属性强化+15
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 10  # 物理、魔法暴击率+10%
        bonus_array[index_extra_percent_addtional_damage] += 12  # 攻击时，附加12%的伤害
        bonus_array[index_strength_and_intelligence] += 0.03 * 35 * 20 / 30  # 攻击时，3%概率增加35点力量，持续20s，冷却30s
        bonus_array[
            index_extra_percent_magic_physical_crit_rate] += 0.05 * 6 * 20 / 30  # 释放技能时，5%概率增加6%物理、魔法暴击率，持续20s，冷却30s
    elif style == '最强战神':
        bonus_array[index_physical_magical_independent_attack_power] += 35  # 三攻+35
        bonus_array[index_strength_and_intelligence] += 75  # 四维+75
        bonus_array[index_extra_percent_attack_speed] += 4  # 攻击速度+4%
        bonus_array[index_extra_all_element_strength] += 20  # 所有属性强化+20
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 15  # 物理、魔法暴击率+15%
        bonus_array[index_extra_percent_final_damage] += 12  # 最终伤害增加12%
        bonus_array[index_strength_and_intelligence] += 0.03 * 35 * 20 / 30  # 攻击时，3%概率增加35点力量，持续20s，冷却30s
        bonus_array[
            index_extra_percent_magic_physical_crit_rate] += 0.05 * 6 * 20 / 30  # 释放技能时，5%概率增加6%物理、魔法暴击率，持续20s，冷却30s
    elif style == '与贝奇邂逅':
        bonus_array[index_strength_and_intelligence] += 90  # 四维+90
        bonus_array[index_extra_percent_attack_speed] += 3  # 攻击速度+3%
        bonus_array[index_extra_all_element_strength] += 12  # 所有属性强化+12
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 8  # 物理、魔法暴击率+8%
        bonus_array[index_extra_percent_addtional_damage] += 10  # 攻击时，附加10%的伤害
        bonus_array[index_strength_and_intelligence] += 0.05 * 25 * 30 / 30  # 攻击时，5%概率增加25点力量，持续30s，冷却30s
        bonus_array[index_extra_percent_attack_speed] += 0.05 * 2 * 30 / 30  # 攻击时，5%概率增加2%三速，持续30s，冷却30s

    # 获取宠物的加成
    creature = creature_select.get()
    if creature in ['弓手维多利亚', '神官格洛丽亚']:
        bonus_array[index_extra_percent_addtional_damage] += 10 * 10 / 30  # 宠物技能+10%攻击力，持续10s，冷却30s
        bonus_array[index_extra_percent_attack_speed] += 5 * 10 / 30  # 宠物技能+5%三速，只考虑普通技能，持续10s，冷却30s
        bonus_array[index_strength_and_intelligence] += 140  # 四维+140
        bonus_array[index_extra_percent_attack_speed] += 5  # 三速+5%
        bonus_array[index_extra_all_element_strength] += 24  # 所有属强+24
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 10  # 暴击率+10%
        bonus_array[index_extra_passive_transfer_skill] += 1  # 转职被动+1， 1-50lv+1
        bonus_array[index_extra_passive_first_awaken_skill] += 1  # 一觉被动+1， 1-50lv+1
        bonus_array[index_extra_active_skill_lv_1_45] += 1  # 1-45级主动+1， 1-50lv+1
        bonus_array[index_extra_active_skill_lv_50] += 1  # 50级主动+1， 1-50lv+1
        bonus_array[index_extra_percent_addtional_damage] += 12  # 攻击时，附加12%的伤害
        bonus_array[index_extra_percent_strength_and_intelligence] += 10  # 力智+10%
        bonus_array[index_cool_correction] += 1.75  # 技能冷却每减1%，冷却矫正系数增加0.35
    elif creature in ['雷光之箭维多利亚', '暴风圣女格洛丽亚']:
        bonus_array[index_extra_percent_addtional_damage] += 10 * 10 / 30  # 宠物技能+10%攻击力，持续10s，冷却30s
        bonus_array[index_extra_percent_attack_speed] += 5 * 10 / 30  # 宠物技能+5%三速，只考虑普通技能，持续10s，冷却30s
        bonus_array[index_strength_and_intelligence] += 150  # 四维+150
        bonus_array[index_extra_percent_attack_speed] += 5  # 三速+5%
        bonus_array[index_extra_all_element_strength] += 24  # 所有属强+24
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 10  # 暴击率+10%
        bonus_array[index_extra_passive_transfer_skill] += 1  # 转职被动+1， 1-50lv+1
        bonus_array[index_extra_passive_first_awaken_skill] += 1  # 一觉被动+1， 1-50lv+1
        bonus_array[index_extra_active_skill_lv_1_45] += 1  # 1-45级主动+1， 1-50lv+1
        bonus_array[index_extra_active_skill_lv_50] += 1  # 50级主动+1， 1-50lv+1
        bonus_array[index_extra_percent_addtional_damage] += 15  # 攻击时，附加15%的伤害
        bonus_array[index_extra_percent_final_damage] += 5  # 最终伤害+5%
        bonus_array[index_extra_percent_strength_and_intelligence] += 12  # 力智+12%
        bonus_array[index_cool_correction] += 1.75  # 技能冷却每减1%，冷却矫正系数增加0.35
    elif creature in ['骑士莱恩', '吟游诗人薇泽达']:
        bonus_array[index_extra_percent_addtional_damage] += 10 * 10 / 30  # 宠物技能+10%攻击力，持续10s，冷却30s
        bonus_array[index_extra_percent_attack_speed] += 5 * 10 / 30  # 宠物技能+5%三速，只考虑普通技能，持续10s，冷却30s
        bonus_array[index_strength_and_intelligence] += 120  # 四维+120
        bonus_array[index_extra_percent_attack_speed] += 5  # 三速+5%
        bonus_array[index_extra_all_element_strength] += 24  # 所有属强+24
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 10  # 暴击率+10%
        bonus_array[index_extra_passive_transfer_skill] += 1  # 转职被动+1， 1-50lv+1
        bonus_array[index_extra_passive_first_awaken_skill] += 1  # 一觉被动+1， 1-50lv+1
        bonus_array[index_extra_active_skill_lv_1_45] += 1  # 1-45级主动+1， 1-50lv+1
        bonus_array[index_extra_active_skill_lv_50] += 1  # 50级主动+1， 1-50lv+1
        bonus_array[index_extra_percent_addtional_damage] += 12  # 攻击时，附加12%的伤害
        bonus_array[index_cool_correction] += 1.75  # 技能冷却每减1%，冷却矫正系数增加0.35
    elif creature in ['古国英豪莱恩', '太初之音薇泽达']:
        bonus_array[index_extra_percent_addtional_damage] += 10 * 10 / 30  # 宠物技能+10%攻击力，持续10s，冷却30s
        bonus_array[index_extra_percent_attack_speed] += 5 * 10 / 30  # 宠物技能+5%三速，只考虑普通技能，持续10s，冷却30s
        bonus_array[index_strength_and_intelligence] += 120  # 四维+120
        bonus_array[index_extra_percent_attack_speed] += 5  # 三速+5%
        bonus_array[index_extra_all_element_strength] += 24  # 所有属强+24
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 10  # 暴击率+10%
        bonus_array[index_extra_passive_transfer_skill] += 1  # 转职被动+1， 1-50lv+1
        bonus_array[index_extra_passive_first_awaken_skill] += 1  # 一觉被动+1， 1-50lv+1
        bonus_array[index_extra_active_skill_lv_1_45] += 1  # 1-45级主动+1， 1-50lv+1
        bonus_array[index_extra_active_skill_lv_50] += 1  # 50级主动+1， 1-50lv+1
        bonus_array[index_extra_percent_addtional_damage] += 15  # 攻击时，附加15%的伤害
        bonus_array[index_extra_percent_final_damage] += 5  # 最终伤害+5%
        bonus_array[index_cool_correction] += 1.75  # 技能冷却每减1%，冷却矫正系数增加0.35
    elif creature in ['雪兔蒂娅', '火狐艾芙']:
        bonus_array[index_extra_percent_addtional_damage] += 10 * 10 / 30  # 宠物技能+10%攻击力，持续10s，冷却30s
        bonus_array[index_extra_percent_attack_speed] += 5 * 10 / 30  # 宠物技能+5%三速，只考虑普通技能，持续10s，冷却30s
        bonus_array[index_strength_and_intelligence] += 100  # 四维+100
        bonus_array[index_extra_percent_attack_speed] += 5  # 三速+5%
        bonus_array[index_extra_all_element_strength] += 20  # 所有属强+20
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 10  # 暴击率+10%
        bonus_array[index_extra_passive_transfer_skill] += 1  # 转职被动+1， 1-50lv+1
        bonus_array[index_extra_passive_first_awaken_skill] += 1  # 一觉被动+1， 1-50lv+1
        bonus_array[index_extra_active_skill_lv_1_45] += 1  # 1-45级主动+1， 1-50lv+1
        bonus_array[index_extra_active_skill_lv_50] += 1  # 50级主动+1， 1-50lv+1
        bonus_array[index_extra_percent_addtional_damage] += 10  # 攻击时，附加10%的伤害
        bonus_array[index_cool_correction] += 1.75  # 技能冷却每减1%，冷却矫正系数增加0.35
    elif creature in ['冰雪魔法师蒂娅', '炽焰咒术师艾芙']:
        bonus_array[index_extra_percent_addtional_damage] += 10 * 10 / 30  # 宠物技能+10%攻击力，持续10s，冷却30s
        bonus_array[index_extra_percent_attack_speed] += 5 * 10 / 30  # 宠物技能+5%三速，只考虑普通技能，持续10s，冷却30s
        bonus_array[index_strength_and_intelligence] += 110  # 四维+110
        bonus_array[index_extra_percent_attack_speed] += 5  # 三速+5%
        bonus_array[index_extra_all_element_strength] += 22  # 所有属强+22
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 10  # 暴击率+10%
        bonus_array[index_extra_passive_transfer_skill] += 1  # 转职被动+1， 1-50lv+1
        bonus_array[index_extra_passive_first_awaken_skill] += 1  # 一觉被动+1， 1-50lv+1
        bonus_array[index_extra_active_skill_lv_1_45] += 1  # 1-45级主动+1， 1-50lv+1
        bonus_array[index_extra_active_skill_lv_50] += 1  # 50级主动+1， 1-50lv+1
        bonus_array[index_extra_percent_addtional_damage] += 12  # 攻击时，附加12%的伤害
        bonus_array[index_cool_correction] += 1.75  # 技能冷却每减1%，冷却矫正系数增加0.35
    elif creature in ['艾莉丝', '克里斯']:
        bonus_array[index_extra_percent_addtional_damage] += 10 * 10 / 20  # 宠物技能+10%攻击力，持续10s，冷却20s
        bonus_array[index_physical_magical_independent_attack_power] += 35  # 三攻+35
        bonus_array[index_extra_percent_attack_speed] += 4  # 三速+4%
        bonus_array[index_extra_all_element_strength] += 15  # 所有属强+15
        bonus_array[index_extra_percent_magic_physical_crit_rate] += 10  # 暴击率+10%
        bonus_array[index_extra_passive_transfer_skill] += 1  # 转职被动+1， 1-50lv+1
        bonus_array[index_extra_passive_first_awaken_skill] += 1  # 一觉被动+1， 1-50lv+1
        bonus_array[index_extra_active_skill_lv_1_45] += 1  # 1-45级主动+1， 1-50lv+1
        bonus_array[index_extra_active_skill_lv_50] += 1  # 50级主动+1， 1-50lv+1

    # todo：加上各种技能宝珠、光环、皮肤、徽章等国服特色的支持

    # from 韩械，原先他在Data.xlsx中所有武器的1-45和50级的被动技能中各增加了2级，其中1级为宠物的技能等级，另一极暂时不确定是哪里来的，可能是其他国服特色
    # 为了保持一致，根据他的建议，把data中所有武器的1-45和50级的主动技能减少两级，然后另外一个技能补正加在这个位置
    bonus_array[index_extra_active_skill_lv_1_45] += 1  # 1-45级主动+1， 1-50lv+1
    bonus_array[index_extra_active_skill_lv_50] += 1  # 50级主动+1， 1-50lv+1

    # 经韩械反馈，属强多出来了80点，应该是之前他在data.xlsx中补正的数值我这边重新计算了一遍- -，暂时先在这里减去吧
    bonus_array[index_extra_all_element_strength] -= 44

    return bonus_array


g_speed_first = True


def modify_slots_order(items, not_select_items, work_uniforms_items, transfer_slots_equips):
    if not g_speed_first:
        return

    # 所有已选装备
    _modify_slots_order(items)
    # 百变怪的各部位可选装备需要与上面的部位顺序一致
    _modify_slots_order(not_select_items)
    # 可升级得到的各部位工作服
    _modify_slots_order(work_uniforms_items)
    # 跨界的备选装备
    _modify_slots_order(transfer_slots_equips)


def _modify_slots_order(slots):
    # 默认槽位顺序为11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33
    # 这种情况下，神话分布在第一位、第六位、第十一位，由于不能同时搭配两个神话，会导致额外多计算很多搭配
    # hack: 优化：由于装备顺序不影响最终计算结果，所以把神话装备先放到前面，那么剪枝可以更高效
    #   所有已选装备、百变怪各部位可选装备、各部位工作服的顺序需要一致，比如第一个是鞋、头肩、腰带，则其余俩也要是这个顺序

    slots[0], slots[1], slots[2], slots[3], slots[4], slots[5], slots[6], slots[7], slots[8], slots[9], slots[10] = \
        slots[0], slots[5], slots[10], slots[1], slots[2], slots[3], slots[4], slots[6], slots[7], slots[8], slots[9]


## 计算函数##
def calc():
    global result_window
    try:
        result_window.destroy()
    except NameError as error:
        pass
    if select_speed.get() == '慢速':
        set_perfect = 1
    else:
        set_perfect = 0
    showsta(text="正在准备组合算法驱动...")
    start_time = time.time()
    load_excel = load_workbook("DATA.xlsx", read_only=True, data_only=True)

    db_one = load_excel["one"]
    opt_one = {}
    name_one = {}
    for row in db_one.rows:
        row_value = [cell.value for cell in row]

        name = row_value[0]
        row_value_cut = row_value[2:]

        opt_one[name] = row_value_cut
        name_one[name] = row_value

    db_set = load_excel["set"]
    opt_set = {}
    for row in db_set.rows:
        row_value = [cell.value for cell in row]

        if len(row_value) == 0:
            continue

        set_index = row_value[0]
        row_value_cut = row_value[2:]

        opt_set[set_index] = row_value_cut  ## DB 装入 ##

    db_buf = load_excel["buf"]
    opt_buf = {}
    name_buf = {}
    for row in db_buf.rows:
        row_value = [cell.value for cell in row]

        buf_index = row_value[0]
        row_value_cut = row_value[2:]

        opt_buf[buf_index] = row_value_cut  ## DB 装入 ##
        name_buf[buf_index] = row_value

    db_buflvl = load_excel["buflvl"]
    opt_buflvl = {}
    for row in db_buflvl.rows:
        row_value = [cell.value for cell in row]

        buf_name = row_value[0]
        row_value_cut = row_value[2:]

        opt_buflvl[buf_name] = row_value_cut

    load_presetc = load_workbook("preset.xlsx", data_only=True)
    db_preset = load_presetc["custom"]
    job_name = jobup_select.get()
    try:
        ele_skill = int(opt_job_ele[job_name][1])
    except KeyError as error:
        tkinter.messagebox.showerror('部分参数有误', "未选择职业或职业非法")
        return
    ele_in = (int(db_preset["B14"].value) + int(db_preset["B15"].value) + int(db_preset["B16"].value) +
              int(ele_skill) - int(db_preset["B18"].value) + int(db_preset["B19"].value) + 13)

    global count_valid, count_invalid, show_number, all_list_num, max_setopt, count_start_time, unique_index
    count_valid = 0
    count_invalid = 0
    show_number = 0

    if job_name[-4:] == "(奶系)":
        active_eff_one = 15
        active_eff_set = 18 - 3
    else:
        active_eff_one = 21
        active_eff_set = 27 - 3

    if time_select.get() == "60秒(觉醒占比↓)":
        lvl_shift = 6
    else:
        lvl_shift = 0

    job_lv1 = opt_job[job_name][11 + lvl_shift]
    job_lv2 = opt_job[job_name][12 + lvl_shift]
    job_lv3 = opt_job[job_name][13 + lvl_shift]
    job_lv4 = opt_job[job_name][14 + lvl_shift]
    job_lv5 = opt_job[job_name][15 + lvl_shift]
    job_lv6 = opt_job[job_name][16 + lvl_shift]
    job_pas0 = opt_job[job_name][0]
    job_pas1 = opt_job[job_name][1]
    job_pas2 = opt_job[job_name][2]
    job_pas3 = opt_job[job_name][3]

    if req_cool.get() == 'X(纯伤害)':
        cool_on = 0
    else:
        cool_on = 1

    valid_weapon = True
    weapon_indexs = []
    try:
        for weapon_name in wep_combopicker.get_selected_entrys():
            weapon_indexs.append(wep_name_to_index[weapon_name])
    except:
        valid_weapon = False
    if len(weapon_indexs) == 0:
        valid_weapon = False

    if not valid_weapon:
        tkinter.messagebox.showerror('部分参数有误', "未选择武器或武器非法")
        return

    # 获取当前装备、百变怪可选装备、工作服列表
    items, not_select_items, work_uniforms_items = get_equips()

    # 获取选定的账号的各部位所拥有的当前账户未拥有的装备列表
    transfer_slots_equips = get_transfer_slots_equips(items, load_presetc["one"])
    transfer_max_count = get_can_transfer_nums()

    # 根据需求决定是否需要开启将神话放到前面来加快剪枝的方案
    modify_slots_order(items, not_select_items, work_uniforms_items, transfer_slots_equips)

    # 已选装备的搭配数
    all_list_num = calc_ori_counts(items)
    # 百变怪增加的搭配数
    all_list_num += calc_bbg_add_counts(items, not_select_items)
    # 额外升级的工作服增加的搭配数
    all_list_num += calc_upgrade_work_uniforms_add_counts(items, not_select_items, work_uniforms_items)

    try:
        showsta(text='开始计算')
        count_start_time = time.time()  # 开始计时
    except MemoryError as error:
        tkinter.messagebox.showerror('内存误差', "已超过可用内存.")
        showsta(text='已中止')
        return

    global exit_calc
    # 开始计算
    exit_calc = 0

    has_baibainguai = baibianguai_select.get() == txt_has_baibianguai
    can_upgrade_work_unifrom_nums = get_can_upgrade_work_unifrom_nums()
    has_uniforms = pre_calc_has_uniforms(items, work_uniforms_items)
    # 超慢速时不进行任何剪枝操作，装备搭配对比的标准是最终计算出的伤害与奶量倍率
    dont_pruning = select_speed.get() == '超慢速'

    # 看了看，主要性能瓶颈在于直接使用了itertools.product遍历所有的笛卡尔积组合，导致无法提前剪枝，只能在每个组合计算前通过条件判断是否要跳过
    # 背景，假设当前处理到下标n（0-10）的装备，前面装备已选择的组合为selected_combination(of size n)，未处理装备为后面11-n-1个，其对应组合数为rcp=len(Cartesian Product(后面11-n-1个装备部位))
    def cartesianProduct(current_index, has_god, baibianguai, upgrade_work_uniforms, transfered_equips, selected_combination,
                         process_func):
        invalid_cnt = 1
        for i in range(current_index + 1, len(items)):
            invalid_cnt *= len(items[i])

        def try_equip(equip):
            global max_setopt

            # 增加处理后续未计算的百变怪
            bbg_invalid_cnt = 0
            if has_baibainguai and baibianguai is None:
                for i in range(current_index + 1, len(items)):
                    bbg_invalid_cnt += invalid_cnt / len(items[i]) * len(not_select_items[i])

            # 剪枝条件1：若当前组合序列已经有神话装备（god），且当前这个部位遍历到的仍是一个神话装备，则可以直接跳过rcp个组合，当前部位之前处理下一个备选装备
            if has_god and is_god(equip):
                inc_invalid_cnt_func(invalid_cnt + bbg_invalid_cnt)
                return

            selected_combination.append(equip)

            # re：剪枝条件2：预计算出后面装备部位能够获得的最大价值量，若当前已有价值量与之相加低于已处理的最高价值量，则剪枝
            if not dont_pruning:
                ub = upper_bound(selected_combination, has_god or is_god(equip), current_index + 1)
                if ub < max_setopt - set_perfect:
                    selected_combination.pop()
                    inc_invalid_cnt_func(invalid_cnt + bbg_invalid_cnt)
                    return

            if current_index < len(items) - 1:
                cartesianProduct(current_index + 1, has_god or is_god(equip), baibianguai, upgrade_work_uniforms, transfered_equips,
                                 selected_combination, process_func)
            else:  # 符合条件的装备搭配
                if dont_pruning:
                    # 不进行任何剪枝操作，装备搭配对比的标准是最终计算出的伤害与奶量倍率
                    process_func(selected_combination, baibianguai, upgrade_work_uniforms, transfered_equips)
                else:
                    # 仅当当前搭配的价值评估函数值不低于历史最高值时才视为有效搭配
                    god = 0
                    if has_god or is_god(equip):
                        god = 1
                    set_list = ["1" + str(selected_combination[x][2:4]) for x in range(0, 11)]
                    set_val = Counter(set_list)
                    del set_val['136', '137', '138']
                    # 1件价值量=0，两件=1，三件、四件=2，五件=3，神话额外增加1价值量
                    setopt_num = sum([floor(x * 0.7) for x in set_val.values()]) + god

                    if setopt_num >= max_setopt - set_perfect:
                        if max_setopt <= setopt_num - god * set_perfect:
                            max_setopt = setopt_num - god * set_perfect
                        process_func(selected_combination, baibianguai, upgrade_work_uniforms, transfered_equips)
                    else:
                        inc_invalid_cnt_func(1)

            selected_combination.pop()

        # 考虑当前部位的每一件可选装备
        for equip in items[current_index]:
            if exit_calc == 1:
                showsta(text='已终止')
                return
            try_equip(equip)

        # 当拥有百变怪，且目前的尝试序列尚未使用到百变怪的时候考虑使用百变怪充当当前部位
        if has_baibainguai and baibianguai is None:
            for equip in not_select_items[current_index]:
                if exit_calc == 1:
                    showsta(text='已终止')
                    return
                baibianguai = equip
                try_equip(equip)
                baibianguai = None

        # 若当前部位的工作服尚未拥有，且可升级工作服的次数尚未用完，则尝试本部位升级工作服
        if not has_uniforms[current_index] and len(upgrade_work_uniforms) < can_upgrade_work_unifrom_nums:
            work_uniform = work_uniforms_items[current_index]

            upgrade_work_uniforms.append(work_uniform)
            try_equip(work_uniform)
            upgrade_work_uniforms.pop()

        # 当当前部位有可以从选定账号跨界的装备，且已跨界数目未超过设定上限，则考虑跨界该部位的装备
        if len(transfer_slots_equips[current_index]) != 0 and len(transfered_equips) < transfer_max_count:
            for equip_to_transfer in transfer_slots_equips[current_index]:
                transfered_equips.append(equip_to_transfer)
                try_equip(equip_to_transfer)
                transfered_equips.pop()

        pass

    # items = [list11, list12, list13, list14, list15, list21, list22, list23, list31, list32, list33]
    # 预处理，计算每个部位是否拥有神话装备
    slot_has_god = []
    for slot_equips in items:
        hg = False
        for equip in slot_equips:
            if is_god(equip):
                hg = True
                break

        slot_has_god.append(hg)

    def has_god_since(idx):
        for i in range(idx, len(slot_has_god)):
            if slot_has_god[i]:
                return True

        return False

    def has_no_god_since(idx):
        return not has_god_since(idx)

    # 为当前已选择序列和后续剩余可选序列计算出一个尽可能精确的上限
    # note: 思路三：进一步降低上限，在当前已有序列的各套装个数的前提下，计算任意n个序列所能产生的价值量最大增益
    # note：思路四：进一步降低上限，在当前已有序列的各套装个数的前提下，计算后面n个序列的各套装配置下所能产生的价值量最大增益
    def upper_bound(selected_combination, selected_has_god, remaining_start_index):
        return upper_bound_2(selected_combination, selected_has_god, remaining_start_index)

    # 对照组：也就是后续
    def upper_bound_none(selected_combination, selected_has_god, remaining_start_index):
        return 1000000

    # note: 思路一：由于每个新增部位产生增益为1或0，因此计算当前序列的价值量，后续每个可选部位按照增益1来计算，可得到约束条件最小的最大上限
    def upper_bound_1(selected_combination, selected_has_god, remaining_start_index):
        # 计算至今为止已有的价值量
        current_value = calc_equip_value(selected_combination, selected_has_god)
        # 后续按最大价值量计算，即每个槽位按能产生1点增益计算
        remaining_max_value = len(items) - remaining_start_index
        hg = has_god_since(remaining_start_index)
        if hg:
            remaining_max_value += 1

        ub = current_value + remaining_max_value
        return ub

    # 新增k个装备所能产生的最大价值量（不计入神话）
    max_inc_values = [0 for i in range(11 + 1)]
    max_inc_values[1] = 1  # 2=>3
    max_inc_values[2] = 2  # 1,1 => 2,2
    max_inc_values[3] = 3  # 1,1,1 => 2,2,2
    max_inc_values[4] = 4  # 1,1,1,1 => 2,2,2,2
    max_inc_values[5] = 5  # 1,1,1,1 => 2,2,2,3
    max_inc_values[6] = 6  # 1,1,1,1 => 2,2,3,3
    max_inc_values[7] = 7  # 1,1,1,1 => 2,3,3,3
    max_inc_values[8] = 7  # upper limit = 533->7
    max_inc_values[9] = 7  # upper limit = 533->7
    max_inc_values[10] = 7  # upper limit = 533->7
    max_inc_values[11] = 7  # upper limit = 533->7

    # note: 思路二：计算新增k个序列所能产生的价值量最大增益
    def upper_bound_2(selected_combination, selected_has_god, remaining_start_index):
        # 计算至今为止已有的价值量
        current_value = calc_equip_value(selected_combination, selected_has_god)
        # 后续按最大价值量计算，即每个槽位按能产生1点增益计算
        remaining_max_value = max_inc_values[len(items) - remaining_start_index]
        hg = has_god_since(remaining_start_index)
        if hg:
            remaining_max_value += 1

        ub = current_value + remaining_max_value
        return ub

    def calc_equip_value(selected_combination, selected_has_god):
        god = 0
        if selected_has_god:
            god = 1
        set_list = ["1" + str(selected_combination[x][2:4]) for x in range(0, len(selected_combination))]
        set_val = Counter(set_list)
        del set_val['136', '137', '138']
        # 1件价值量=0，两件=1，三件、四件=2，五件=3，神话额外增加1价值量
        setopt_num = sum([floor(x * 0.7) for x in set_val.values()]) + god

        return setopt_num

    is_shuchu_job = job_name[4:7] not in ["奶爸", "奶妈", "奶萝"]
    if is_shuchu_job:

        # 代码名称
        # 0调节器 1추공  2증 3크 4추 5속추
        # 6모 7공 8스탯 9속강 10지속 11스증 12특수
        # 13공속 14크확 / 15 액티브 / 16~19 패시브 /20 신화여부/21세트코드/22쿨감보정/23 원쿨감
        # 단품:24~33 저장소 / 34 二觉캐액티브
        getone = opt_one.get
        minheap = MinHeap(5)
        unique_index = 0
        max_setopt = 0
        show_number = 1

        def process(calc_now, baibianguai, upgrade_work_uniforms, transfered_equips):
            set_list = ["1" + str(calc_now[x][2:4]) for x in range(0, 11)]
            set_on = [];
            setapp = set_on.append
            setcount = set_list.count
            set_oncount = set_on.count
            onecount = calc_now.count
            for i in range(101, 136):
                if setcount(str(i)) == 2:
                    setapp(str(i) + "1")
                if 4 >= setcount(str(i)) >= 3:
                    setapp(str(i) + "2")
                if setcount(str(i)) == 5:
                    setapp(str(i) + "3")
            for i in range(136, 139):
                if setcount(str(i)) == 2:
                    setapp(str(i) + "0")
                if 4 >= setcount(str(i)) >= 3:
                    setapp(str(i) + "1")
                if setcount(str(i)) == 5:
                    setapp(str(i) + "2")
            if onecount('32410650') == 1:
                if onecount('21400340') == 1:
                    setapp('1401')
                elif onecount('31400540') == 1:
                    setapp('1401')

            for wep_num in weapon_indexs:
                calc_wep = (wep_num,) + tuple(calc_now)
                damage = 0
                # 获取输出职业的国服特色数值作为初始值
                base_array = get_shuchu_bonus_attributes()
                # 加上职业的基础属强
                base_array[index_extra_all_element_strength] += ele_in

                skiper = 0
                for_calc = tuple(set_on) + calc_wep
                oneone = len(for_calc)
                oneonelist = []
                for i in range(oneone):
                    no_cut = getone(for_calc[i])  ## 11번 스증
                    if no_cut is None:
                        # hack：目前select是默认初始化时将tg{1101-3336}[0,1]范围的key对应的值都设为0，而百变怪会根据select的值为0来筛选出未选择的集合
                        #  因此在这里如果为None，就是这种情况，直接返回就可以了
                        global count_invalid
                        count_invalid = count_invalid + 1
                        return
                    cut = np.array(no_cut[0:20] + no_cut[22:23] + no_cut[34:35] + no_cut[38:44])
                    skiper = (skiper / 100 + 1) * (cut[index_extra_percent_skill_attack_power] / 100 + 1) * 100 - 100
                    oneonelist.append(cut)
                for i in range(oneone):
                    base_array = base_array + oneonelist[i]

                if set_oncount('1201') == 1 and onecount('32200') == 1:
                    base_array[index_extra_percent_crit_damage] -= 5
                if onecount('33230') == 1 or onecount('33231') == 1:
                    if onecount('31230') == 0:
                        base_array[index_extra_percent_addtional_damage] -= 10
                    if onecount('32230') == 0:
                        base_array[index_extra_all_element_strength] -= 40
                # 特殊处理天命无常套装
                if onecount('15340') == 1 or onecount('23340') == 1 or onecount('33340') == 1 or onecount(
                        '33341') == 1:
                    if set_oncount('1341') == 0 and set_oncount('1342') == 0:  # 1341=天命两件套 1342=天命三件套
                        if onecount('15340') == 1:
                            base_array[index_extra_all_element_strength] -= 20
                        elif onecount('23340') == 1:  # 天命无常-戒指-命运的捉弄
                            base_array[index_extra_percent_attack_damage] -= 10
                        elif onecount('33340') == 1:
                            base_array[index_extra_percent_final_damage] -= 5  #
                        else:  # 天命无常-神话耳环-命运反抗者
                            base_array[index_extra_all_element_strength] -= 4  # ele=4
                            base_array[index_extra_percent_attack_damage] -= 2  # damper=2
                            base_array[index_extra_percent_final_damage] -= 1  # allper=6
                            base_array[index_extra_percent_strength_and_intelligence] -= 1.93  # staper=15
                if onecount('11111') == 1:
                    if set_oncount('1112') == 1 or set_oncount('1113') == 1:
                        base_array[index_cool_correction] += 10
                if onecount('11301') == 1:
                    if onecount('22300') != 1:
                        base_array[index_extra_percent_addtional_damage] -= 10
                        base_array[index_extra_percent_physical_magical_independent_attack_power] += 10
                    if onecount('31300') != 1:
                        base_array[index_extra_percent_addtional_damage] -= 10
                        base_array[index_extra_percent_physical_magical_independent_attack_power] += 10

                base_array[index_extra_percent_skill_attack_power] = skiper  # 技能攻击力 +X%
                real_bon = (base_array[index_extra_percent_addtional_damage] +  # 攻击时，附加X%的伤害，也就是白字
                            base_array[index_extra_percent_elemental_damage] *  # 攻击时，附加X%的属性伤害
                            (base_array[index_extra_all_element_strength] * 0.0045 + 1.05))  # 所有属性强化+X
                actlvl = ((base_array[index_extra_active_second_awaken_skill] +  # 二觉主动技能
                           base_array[index_extra_active_skill_lv_1_45] * job_lv1 +  # 1_45主动技能
                           base_array[index_extra_active_skill_lv_50] * job_lv2 +  # 50主动技能
                           base_array[index_extra_active_skill_lv_60_80] * job_lv3 +  # 60_80主动技能
                           base_array[index_extra_active_skill_lv_85] * job_lv4 +  # 85主动技能
                           base_array[index_extra_active_skill_lv_95] * job_lv5 +  # 95主动技能
                           base_array[index_extra_active_skill_lv_100] * job_lv6  # 100主动技能
                           ) / 100 + 1)
                paslvl = (((100 + base_array[index_extra_passive_transfer_skill] * job_pas0) / 100) *  # 增加转职被动的等级
                          ((100 + base_array[index_extra_passive_first_awaken_skill] * job_pas1) / 100) *  # 增加一绝被动的等级
                          ((100 + base_array[index_extra_passive_second_awaken_skill] * job_pas2) / 100) *  # 增加二觉被动的等级
                          ((100 + base_array[index_extra_passive_third_awaken_skill] * job_pas3) / 100)  # 增加三觉被动的等级
                          )
                damage = ((base_array[index_extra_percent_attack_damage] / 100 + 1) *  # 攻击时额外增加X%的伤害增加量
                          (base_array[index_extra_percent_crit_damage] / 100 + 1) *  # 暴击时，额外增加X%的伤害增加量
                          (real_bon / 100 + 1) *  # 白字与属强的最终综合值
                          (base_array[index_extra_percent_final_damage] / 100 + 1) *  # 最终伤害+X%
                          (base_array[index_extra_percent_physical_magical_independent_attack_power] / 100 + 1) *  # 物理/魔法/独立攻击力 +X%
                          (base_array[index_extra_percent_strength_and_intelligence] / 100 + 1) *  # 力智+X%
                          (base_array[index_extra_all_element_strength] * 0.0045 + 1.05) *  # 所有属性强化+X
                          (base_array[index_extra_percent_continued_damage] / 100 + 1) *  # 发生持续伤害5秒，伤害量为对敌人造成伤害的X%
                          (skiper / 100 + 1) *  # 技能攻击力 +X%
                          (base_array[index_extra_percent_special_effect] / 100 + 1) *  # 特殊词条，作者为每个特殊词条打了相应的强度百分比分，如一叶障目对忍者一些技能的特殊改变被认为可以强化9%，守护的抉择（歧路鞋）的护石增强词条被认为可以增强21%
                          actlvl * paslvl *  # 主动技能与被动技能的影响
                          ((54500 + 3.31 * base_array[index_strength_and_intelligence]) / 54500) *  # 力智
                          ((4800 + base_array[index_physical_magical_independent_attack_power]) / 4800) *  # 物理/魔法/独立攻击力
                          (1 + cool_on * base_array[index_cool_correction] / 100) /  # 冷却矫正系数，每冷却1%，记0.35这个值
                          (1.05 + 0.0045 * int(ele_skill)))  # 不太确定，可能是属性抗性？속강

                base_array[index_extra_percent_addtional_damage] = real_bon
                global unique_index
                unique_index += 1
                not_owned_equips = [uwu for uwu in upgrade_work_uniforms]
                for equip in transfered_equips:
                    not_owned_equips.append(equip)
                minheap.add((damage, unique_index,
                             [calc_wep, base_array, baibianguai, tuple(not_owned_equips)]))

                global count_valid
                count_valid = count_valid + 1

        cartesianProduct(0, False, None, [], [], [], process)

        show_number = 0
        showsta(text='结果统计中')

        ranking = []
        for rank, data in enumerate(minheap.getTop()):
            damage = data[0]
            value = data[2]
            ranking.append((damage, value))
            showsta(text='结果统计中' + str(rank + 1) + " / 5")

        show_result(ranking, 'deal', ele_skill)

    else:
        # TODO: 增加支持对国服特色的支持
        base_b = 10 + int(db_preset['H2'].value) + int(db_preset['H4'].value) + int(db_preset['H5'].value) + 1
        base_c = 12 + int(db_preset['H3'].value) + 1
        base_pas0 = 0
        base_pas0_c = 3
        base_pas0_b = 0
        base_stat_s = 4166 + 74 - 126 + int(db_preset['H1'].value)
        base_stat_d = int(db_preset['H6'].value) - int(db_preset['H1'].value)
        base_stat_h = 4308 - 45 - 83 + int(db_preset['H1'].value)
        base_pas0_1 = 0
        lvlget = opt_buflvl.get
        minheap1 = MinHeap(5)
        minheap2 = MinHeap(5)
        minheap3 = MinHeap(5)
        unique_index = 0
        setget = opt_buf.get
        max_setopt = 0
        show_number = 1

        def process(calc_now, baibianguai, upgrade_work_uniforms, transfered_equips):
            set_list = ["1" + str(calc_now[x][2:4]) for x in range(0, 11)]
            set_val = Counter(set_list)
            del set_val['136', '137', '138']
            set_on = [];
            setapp = set_on.append
            setcount = set_list.count
            for i in range(101, 136):
                if setcount(str(i)) == 2:
                    setapp(str(i) + "1")
                if 4 >= setcount(str(i)) >= 3:
                    setapp(str(i) + "2")
                if setcount(str(i)) == 5:
                    setapp(str(i) + "3")
            for i in range(136, 139):
                if setcount(str(i)) == 2:
                    setapp(str(i) + "0")
                if 4 >= setcount(str(i)) >= 3:
                    setapp(str(i) + "1")
                if setcount(str(i)) == 5:
                    setapp(str(i) + "2")

            # 코드 이름
            # 0 체정 1 지능
            # 祝福 2 스탯% 3 물공% 4 마공% 5 독공%
            # 아포 6 고정 7 스탯%
            # 8 축렙 9 포렙
            # 10 아리아/보징증폭
            # 11 전직패 12 보징 13 각패1 14 각패2 15 二觉 16 각패3
            # 17 깡신념 18 깡신실 19 아리아쿨 20 하베쿨
            for wep_num in weapon_indexs:
                calc_wep = (wep_num,) + tuple(calc_now)
                base_array = np.array(
                    [base_stat_h, base_stat_s, 0, 0, 0, 0, 0, 0, base_b, base_c, 0, base_pas0, base_pas0_1, 0, 0, 0,
                     0, 0,
                     0, 0, 0])

                b_stat = 10.24  ##탈리스만
                b_phy = 0
                b_mag = 0
                b_ind = 0
                c_per = 0
                for_calc = tuple(set_on) + calc_wep
                oneone = len(for_calc)
                oneonelist = []
                for i in range(oneone):
                    # 获取该装备的buff属性
                    # re：阅读data.xlsx的buf sheet和buflvl确认奶系职业的各个计算维度的含义
                    cut = setget(for_calc[i])  ## 2 3 4 5 7
                    if cut is None:
                        # hack：目前select是默认初始化时将tg{1101-3336}[0,1]范围的key对应的值都设为0，而百变怪会根据select的值为0来筛选出未选择的集合
                        #  因此在这里如果为None，就是这种情况，直接返回就可以了
                        global count_invalid
                        count_invalid = count_invalid + 1
                        return
                    no_cut = np.array(cut)
                    base_array = base_array + no_cut
                    b_stat = (b_stat / 100 + 1) * (no_cut[2] / 100 + 1) * 100 - 100
                    b_phy = (b_phy / 100 + 1) * (no_cut[3] / 100 + 1) * 100 - 100
                    b_mag = (b_mag / 100 + 1) * (no_cut[4] / 100 + 1) * 100 - 100
                    b_ind = (b_ind / 100 + 1) * (no_cut[5] / 100 + 1) * 100 - 100
                    c_per = (c_per / 100 + 1) * (no_cut[7] / 100 + 1) * 100 - 100
                    oneonelist.append(no_cut)

                if job_name[4:7] == "奶爸":
                    b_base_att = lvlget('hol_b_atta')[int(base_array[8])]
                    stat_pas0lvl_b = lvlget('pas0')[int(base_array[11]) + base_pas0_b] + lvlget('hol_pas0_1')[
                        int(base_array[12])]
                    stat_pas0lvl_c = lvlget('pas0')[int(base_array[11]) + base_pas0_c] + lvlget('hol_pas0_1')[
                        int(base_array[12])]
                    stat_pas1lvl = lvlget('hol_pas1')[int(base_array[13])]
                    stat_pas2lvl = lvlget('hol_act2')[int(base_array[15])]
                    stat_pas3lvl = lvlget('pas3')[int(base_array[16])]
                    stat_b = base_array[0] + stat_pas0lvl_b + stat_pas1lvl + stat_pas2lvl + stat_pas3lvl + 19 * \
                             base_array[10] + base_stat_d
                    stat_c = base_array[0] + stat_pas0lvl_c + stat_pas1lvl + stat_pas2lvl + stat_pas3lvl + 19 * \
                             base_array[10]
                    b_stat_calc = int(
                        int(lvlget('hol_b_stat')[int(base_array[8])] * (b_stat / 100 + 1)) * (stat_b / 630 + 1))
                    b_phy_calc = int(int(b_base_att * (b_phy / 100 + 1)) * (stat_b / 630 + 1))
                    b_mag_calc = int(int(b_base_att * (b_mag / 100 + 1)) * (stat_b / 630 + 1))
                    b_ind_calc = int(int(b_base_att * (b_ind / 100 + 1)) * (stat_b / 630 + 1))
                    b_average = int((b_phy_calc + b_mag_calc + b_ind_calc) / 3)
                    c_calc = int(int((lvlget('c_stat')[int(base_array[9])] + base_array[6]) * (c_per / 100 + 1)) * (
                            stat_c / 750 + 1))
                    pas1_calc = int(lvlget('hol_pas1_out')[int(base_array[13])] + 213 + base_array[17])
                    pas1_out = str(
                        int(lvlget('hol_pas1_out')[int(base_array[13])] + 213 + base_array[17])) + "  (" + str(
                        int(20 + base_array[13])) + "级)"
                    save1 = str(b_stat_calc) + "/" + str(b_average) + "   [" + str(int(stat_b)) + "(" + str(
                        int(base_array[8])) + "级)]"

                else:
                    if job_name[4:7] == "奶妈":
                        b_value = 675
                        aria = 1.25 + 0.05 * base_array[10]
                    if job_name[4:7] == "奶萝":
                        b_value = 665
                        aria = (1.20 + 0.05 * base_array[10]) * 1.20

                    b_base_att = lvlget('se_b_atta')[int(base_array[8])]
                    stat_pas0lvl_b = lvlget('pas0')[int(base_array[11]) + int(base_pas0_b)]
                    stat_pas0lvl_c = lvlget('pas0')[int(base_array[11]) + int(base_pas0_c)]
                    stat_pas1lvl = lvlget('se_pas1')[int(base_array[13])] + base_array[18]
                    stat_pas2lvl = lvlget('se_pas2')[int(base_array[14])]
                    stat_pas3lvl = lvlget('pas3')[int(base_array[16])]
                    stat_b = base_array[
                                 1] + stat_pas0lvl_b + stat_pas1lvl + stat_pas2lvl + stat_pas3lvl + base_stat_d
                    stat_c = base_array[1] + stat_pas0lvl_c + stat_pas1lvl + stat_pas2lvl + stat_pas3lvl
                    b_stat_calc = int(int(lvlget('se_b_stat')[int(base_array[8])] * (b_stat / 100 + 1)) * (
                            stat_b / b_value + 1) * aria)
                    b_phy_calc = int(int(b_base_att * (b_phy / 100 + 1) * (stat_b / b_value + 1)) * aria)
                    b_mag_calc = int(int(b_base_att * (b_mag / 100 + 1) * (stat_b / b_value + 1)) * aria)
                    b_ind_calc = int(int(b_base_att * (b_ind / 100 + 1) * (stat_b / b_value + 1)) * aria)
                    b_average = int((b_phy_calc + b_mag_calc + b_ind_calc) / 3)
                    c_calc = int(int((lvlget('c_stat')[int(base_array[9])] + base_array[6]) * (c_per / 100 + 1)) * (
                            stat_c / 750 + 1))
                    pas1_calc = int(stat_pas1lvl + 442)
                    pas1_out = str(int(stat_pas1lvl + 442)) + "  (" + str(int(20 + base_array[13])) + "级)"
                    save1 = str(b_stat_calc) + "(" + str(int(b_stat_calc / aria)) + ")/ " + str(
                        b_average) + "(" + str(int(b_average / aria)) + ")\n                  [" + str(
                        int(stat_b)) + "(" + str(int(base_array[8])) + "级)]"

                save2 = str(c_calc) + "    [" + str(int(stat_c)) + "(" + str(int(base_array[9])) + "级)]"
                ##1축 2포 3합
                global unique_index
                unique_index += 1
                not_owned_equips = [uwu for uwu in upgrade_work_uniforms]
                for equip in transfered_equips:
                    not_owned_equips.append(equip)
                save_data = [calc_wep, [save1, save2, pas1_out], baibianguai, tuple(not_owned_equips)]
                minheap1.add((((15000 + b_stat_calc) / 250 + 1) * (2650 + b_average), unique_index, save_data))
                minheap2.add((((15000 + c_calc) / 250 + 1) * 2650, unique_index, save_data))
                minheap3.add(
                    (((15000 + pas1_calc + c_calc + b_stat_calc) / 250 + 1) * (2650 + b_average), unique_index, save_data))

                global count_valid
                count_valid = count_valid + 1

        cartesianProduct(0, False, None, [], [], [], process)
        show_number = 0
        showsta(text='结果统计中')

        ranking1 = [];
        ranking2 = [];
        ranking3 = []
        for rank, data in enumerate(minheap1.getTop()):
            damage = data[0]
            value = data[2]
            ranking1.append((damage, value))
        for rank, data in enumerate(minheap2.getTop()):
            damage = data[0]
            value = data[2]
            ranking2.append((damage, value))
        for rank, data in enumerate(minheap3.getTop()):
            damage = data[0]
            value = data[2]
            ranking3.append((damage, value))

        ranking = [ranking1, ranking2, ranking3]
        show_result(ranking, 'buf', ele_skill)

    load_presetc.close()
    load_excel.close()
    showsta(text='输出完成' + "时间 = " + format_time(time.time() - start_time))
    # 结束计算
    exit_calc = 1
    # print("时间 = " + str(time.time() - start_time) + "秒")


def calc_thread():
    threading.Thread(target=calc, daemon=True).start()


def stop_calc():
    global exit_calc
    exit_calc = 1


def get_equips():
    list11 = [];
    list12 = [];
    list13 = [];
    list14 = [];
    list15 = []
    list21 = [];
    list22 = [];
    list23 = [];
    list31 = [];
    list32 = [];
    list33 = []
    list_setnum = [];
    list_godnum = []
    # 未选中的装备
    listns11 = [];
    listns12 = [];
    listns13 = [];
    listns14 = [];
    listns15 = []
    listns21 = [];
    listns22 = [];
    listns23 = [];
    listns31 = [];
    listns32 = [];
    listns33 = []

    for i in range(1010, 1999):
        try:
            if eval('select_item["tg1{}"]'.format(i)) == 1:
                list11.append('1' + str(i))
                list_setnum.append(str(i)[1:])
            elif eval('select_item["tg1{}"]'.format(i)) == 0 and can_convert_from_baibianguai('1' + str(i)):
                listns11.append('1' + str(i))
        except KeyError as error:
            c = 1
    for i in range(2010, 2999):
        try:
            if eval('select_item["tg1{}"]'.format(i)) == 1:
                list12.append('1' + str(i))
                list_setnum.append(str(i)[1:])
            elif eval('select_item["tg1{}"]'.format(i)) == 0 and can_convert_from_baibianguai('1' + str(i)):
                listns12.append('1' + str(i))
        except KeyError as error:
            c = 1
    for i in range(3010, 3999):
        try:
            if eval('select_item["tg1{}"]'.format(i)) == 1:
                list13.append('1' + str(i))
                list_setnum.append(str(i)[1:])
            elif eval('select_item["tg1{}"]'.format(i)) == 0 and can_convert_from_baibianguai('1' + str(i)):
                listns13.append('1' + str(i))
        except KeyError as error:
            c = 1
    for i in range(4010, 4999):
        try:
            if eval('select_item["tg1{}"]'.format(i)) == 1:
                list14.append('1' + str(i))
                list_setnum.append(str(i)[1:])
            elif eval('select_item["tg1{}"]'.format(i)) == 0 and can_convert_from_baibianguai('1' + str(i)):
                listns14.append('1' + str(i))
        except KeyError as error:
            c = 1
    for i in range(5010, 5999):
        try:
            if eval('select_item["tg1{}"]'.format(i)) == 1:
                list15.append('1' + str(i))
                list_setnum.append(str(i)[1:])
            elif eval('select_item["tg1{}"]'.format(i)) == 0 and can_convert_from_baibianguai('1' + str(i)):
                listns15.append('1' + str(i))
        except KeyError as error:
            c = 1
    for i in range(1010, 1999):
        try:
            if eval('select_item["tg2{}"]'.format(i)) == 1:
                list21.append('2' + str(i))
                list_setnum.append(str(i)[1:])
            elif eval('select_item["tg2{}"]'.format(i)) == 0 and can_convert_from_baibianguai('2' + str(i)):
                listns21.append('2' + str(i))
        except KeyError as error:
            c = 1
    for i in range(2010, 2999):
        try:
            if eval('select_item["tg2{}"]'.format(i)) == 1:
                list22.append('2' + str(i))
                list_setnum.append(str(i)[1:])
            elif eval('select_item["tg2{}"]'.format(i)) == 0 and can_convert_from_baibianguai('2' + str(i)):
                listns22.append('2' + str(i))
        except KeyError as error:
            c = 1
    for i in range(3010, 3999):
        try:
            if eval('select_item["tg2{}"]'.format(i)) == 1:
                list23.append('2' + str(i))
                list_setnum.append(str(i)[1:])
            elif eval('select_item["tg2{}"]'.format(i)) == 0 and can_convert_from_baibianguai('2' + str(i)):
                listns23.append('2' + str(i))
        except KeyError as error:
            c = 1
    for i in range(1010, 1999):
        try:
            if eval('select_item["tg3{}"]'.format(i)) == 1:
                list31.append('3' + str(i))
                list_setnum.append(str(i)[1:])
            elif eval('select_item["tg3{}"]'.format(i)) == 0 and can_convert_from_baibianguai('3' + str(i)):
                listns31.append('3' + str(i))
        except KeyError as error:
            c = 1
    for i in range(2010, 2999):
        try:
            if eval('select_item["tg3{}"]'.format(i)) == 1:
                list32.append('3' + str(i))
                list_setnum.append(str(i)[1:])
            elif eval('select_item["tg3{}"]'.format(i)) == 0 and can_convert_from_baibianguai('3' + str(i)):
                listns32.append('3' + str(i))
        except KeyError as error:
            c = 1
    for i in range(3010, 3999):
        try:
            if eval('select_item["tg3{}"]'.format(i)) == 1:
                list33.append('3' + str(i))
                list_setnum.append(str(i)[1:])
            elif eval('select_item["tg3{}"]'.format(i)) == 0 and can_convert_from_baibianguai('3' + str(i)):
                listns33.append('3' + str(i))
        except KeyError as error:
            c = 1
    algo_list = ['11', '12', '13', '14', '15', '21', '22', '23', '31', '32', '33']
    if select_speed.get() == '快速':
        for i in list_setnum:
            if list_setnum.count(i) == 1:
                if i[-1] != '1':
                    for ca in algo_list:
                        try:
                            eval("list{}.remove('{}{}')".format(ca, ca, i))
                        except ValueError as error:
                            c = 1

    for know_one in know_list:
        if eval('select_item["tg{}"]'.format(know_one)) == 1:
            eval('list{}.append(str({}))'.format(know_one[0:2], know_one))

    # 为了计算结果更精确，永远将100传说防具、普雷首饰、普雷特殊加入备选方案
    list11.append('11360')
    list12.append('12360')
    list13.append('13360')
    list14.append('14360')
    list15.append('15360')

    list21.append('21370')
    list22.append('22370')
    list23.append('23370')

    list31.append('31380');
    list32.append('32380');
    list33.append('33380')

    # 所有已选装备
    items = [list11, list12, list13, list14, list15, list21, list22, list23, list31, list32, list33]
    # 百变怪的各部位可选装备需要与上面的部位顺序一致
    not_select_items = [listns11, listns12, listns13, listns14, listns15, listns21, listns22, listns23, listns31, listns32, listns33]
    # 可升级得到的各部位工作服
    work_uniforms_items = ["11150", "12150", "13150", "14150", "15150", "21190", "22190", "23190", "31230", "32230", "33230"]

    return items, not_select_items, work_uniforms_items


def inc_invalid_cnt_func(cnt):
    global count_invalid
    count_invalid += int(cnt)


# 装备编号的最后一位表示是否是神话装备，eg：33341
def is_god(equip):
    return int(equip[-1]) == 1


# 百变怪是否可以转换成该装备
def can_convert_from_baibianguai(equip):
    # 百变怪不能转换为神话装备
    if is_god(equip):
        return False
    # 百变怪不能转化为工作服
    if equip in work_uniforms:
        return False
    # 百变怪不能转化为智慧产物
    if equip in the_product_of_wisdoms:
        return False

    return True


def calc_ori_counts(all_slots_equips):
    cnt = 1
    for one_slot_equips in all_slots_equips:
        cnt *= len(one_slot_equips)
    return cnt


def calc_bbg_add_counts(slots_equips, slots_not_select_equips):
    if baibianguai_select.get() != txt_has_baibianguai:
        return 0

    ori_counts = calc_ori_counts(slots_equips)

    # 百变怪增加的搭配数
    bbg_add_num = 0
    for i in range(0, len(slots_not_select_equips)):
        bbg_add_num += ori_counts / len(slots_equips[i]) * len(slots_not_select_equips[i])

    return bbg_add_num


# 玩家各个部位是否已经升级了工作服
def pre_calc_has_uniforms(items, work_uniforms_items):
    return [work_uniforms_items[idx] in items[idx] for idx in range(len(items))]


def get_can_upgrade_work_unifrom_nums():
    # 用户配置的当前可升级的工作服数目
    can_upgrade_work_unifrom_nums = 0
    if can_upgrade_work_unifrom_nums_select.get() in can_upgrade_work_unifrom_nums_str_2_int:
        can_upgrade_work_unifrom_nums = can_upgrade_work_unifrom_nums_str_2_int[
            can_upgrade_work_unifrom_nums_select.get()]
    return can_upgrade_work_unifrom_nums


def get_can_transfer_nums():
    # 用户配置的当前可升级的工作服数目
    can_transfer_nums = 0
    if can_transfer_nums_select.get() in can_transfer_nums_str_2_int:
        can_transfer_nums = can_transfer_nums_str_2_int[
            can_transfer_nums_select.get()]
    return can_transfer_nums


# 获取选定的账号的各部位所拥有的当前账户未拥有的装备列表
def get_transfer_slots_equips(items, sheet):
    # 获取各存档的装备信息
    slot_name_list = ['11', '12', '13', '14', '15', '21', '22', '23', '31', '32', '33']
    slot_name_to_index = {}
    for idx, name in enumerate(slot_name_list):
        slot_name_to_index[name] = idx

    transfer_slots_equips = [[] for i in range(0, 11)]
    for account_name in transfer_equip_combopicker.get_selected_entrys():
        # 获取当前存档的index
        account_index = 0
        try:
            account_index = save_name_list.index(account_name)
        except Exception:
            continue

        # 读取各个装备的点亮情况
        for i in range(1, 264):
            has_equip = sheet.cell(i, 2 + account_index).value == 1
            equip_index = sheet.cell(i, 1).value
            if len(equip_index) == 6:
                # 六位为武器，过滤掉武器
                continue

            if has_equip:
                try:
                    slot_index = slot_name_to_index[equip_index[:2]]
                    # 如果该装备当前账号未拥有，且之前的账户中未添加过，则加入备选集
                    if equip_index not in items[slot_index] and equip_index not in transfer_slots_equips[slot_index] and not is_god(equip_index):
                        transfer_slots_equips[slot_index].append(equip_index)
                except KeyError as error:
                    pass

    return transfer_slots_equips


def calc_upgrade_work_uniforms_add_counts(slots_equips, slots_not_select_equips, slots_work_uniforms):
    # 找出所有尚未升级工作服的部位
    not_has_uniform_slots = []
    for slot, work_uniform in enumerate(slots_work_uniforms):
        if work_uniform not in slots_equips[slot]:
            not_has_uniform_slots.append(slot)

    total_add_counts = 0

    # 穷举尚未升级部位的大小小于等于最大可升级数目的所有组合
    max_upgrade_count = min(get_can_upgrade_work_unifrom_nums(), len(not_has_uniform_slots))
    # 遍历所有可能升级的件数
    for upgrade_count in range(1, max_upgrade_count + 1):
        # 遍历升级该件数的所有部位的组合
        for upgrade_slots in itertools.combinations(not_has_uniform_slots, upgrade_count):
            # 获取非升级部位的已有装备
            other_slots_equips = []
            for slot, slot_equips in enumerate(slots_equips):
                if slot not in upgrade_slots:
                    other_slots_equips.append(slot_equips)
            # 获取非升级部位的未选择装备
            other_slots_not_select_equips = []
            for slot, slot_not_select_equips in enumerate(slots_not_select_equips):
                if slot not in upgrade_slots:
                    other_slots_not_select_equips.append(slot_not_select_equips)

            # 计算该升级方案下其余部位的可能搭配数目
            total_add_counts += calc_bbg_add_counts(other_slots_equips, other_slots_not_select_equips)

    return total_add_counts


result_window_width = 585
result_window_readable_result_area_height = 18 * 3
result_window_height = 402 + result_window_readable_result_area_height

res_txt_readable_result_left_top_x = 0
res_txt_readable_result_left_top_y = result_window_height - result_window_readable_result_area_height
res_txt_readable_result_center_x = result_window_width // 2
res_txt_readable_result_center_y = result_window_height - result_window_readable_result_area_height // 2
res_txt_readable_result = None


# 用文字方式写成当前搭配，避免每次都得一个个肉眼对比图标来确认是啥装备
def change_readable_result_area(weapon, equips, is_create):
    global res_txt_readable_result, canvas_res

    readable_names = []
    readable_names.append(equip_index_to_realname[weapon])

    # 智慧产物以外的套装信息
    set_list = ["1" + str(equips[x][2:4]) for x in range(0, 11) if len(equips[x]) < 8]
    for set_index, count in collections.Counter(set_list).most_common():
        readable_names.append("{}-{}".format(equip_index_to_realname[set_index], count))

    # 智慧产物单独列出
    wisdom_indexs = [equips[x] for x in range(0, 11) if len(equips[x]) == 8]
    # 赤鬼的次元石改造五阶段词条：装备[青面修罗的面具]、[噙毒手套]中1种以上时，释放疯魔索伦之力。 - 攻击时，附加7%的伤害。
    if wisdom_indexs.count('32410650') == 1:
        if wisdom_indexs.count('21400340'):
            readable_names.append(equip_index_to_realname["1401"])
            wisdom_indexs.remove('32410650')
            wisdom_indexs.remove('21400340')
        elif wisdom_indexs.count('31400540') == 1:
            readable_names.append(equip_index_to_realname["1401"])
            wisdom_indexs.remove('32410650')
            wisdom_indexs.remove('31400540')
    for wisdom_index in wisdom_indexs:
        readable_names.append(equip_index_to_realname[wisdom_index])

    line_word_count = 0
    max_line_word_count = 40
    readable_result = ""
    for name in readable_names:
        if line_word_count + len(name) >= max_line_word_count:
            line_word_count = 0
            readable_result += "\n"
        elif line_word_count != 0:
            readable_result += ' | '

        readable_result += name
        line_word_count += len(name)

    content = pretty_words(readable_names, 40, ' | ')
    if is_create:
        res_txt_readable_result = canvas_res.create_text(res_txt_readable_result_center_x, res_txt_readable_result_center_y,
                                                     text=content,
                                                     font=guide_font, fill='white')
    else:
        canvas_res.itemconfig(res_txt_readable_result, text=content)


# 展示当前搭配的各装备名称
def show_name():
    global g_rank_equips, g_current_rank, g_current_job, g_current_buff_type

    equips = None
    if g_current_job == "deal":
        equips = g_rank_equips[g_current_rank]
    else:
        equips = g_rank_equips[g_current_buff_type][g_current_rank]

    readable_names = []
    for equip in equips:
        readable_names.append(equip_index_to_realname[equip])

    tkinter.messagebox.showinfo("装备详细信息", pretty_words(readable_names, 30, ' | '))


# 保证一行不会有太多词
def pretty_words(words, max_line_word_count, delimiter):
    pretty_result = ""
    line_word_count = 0
    for word in words:
        if line_word_count + len(word) >= max_line_word_count:
            line_word_count = 0
            pretty_result += "\n"
        elif line_word_count != 0:
            pretty_result += delimiter

        pretty_result += word
        line_word_count += len(word)

    return pretty_result


def show_result(rank_list, job_type, ele_skill):
    global g_rank_equips, g_current_rank, g_current_job, g_current_buff_type
    g_current_rank = 0
    g_current_job = job_type

    global result_window
    result_window = tkinter.Toplevel(self)
    # result_window.attributes("-topmost", True)
    result_window.geometry("{}x{}+800+400".format(result_window_width, result_window_height))
    result_window.title("计算结果")
    result_window.resizable(False, False)
    global canvas_res
    canvas_width = result_window_width + 2
    canvas_height = result_window_height + 2
    canvas_res = Canvas(result_window, width=canvas_width, height=canvas_height, bd=0)
    canvas_res.place(x=-2, y=-2)
    if job_type == 'deal':
        result_bg = tkinter.PhotoImage(file='ext_img/bg_result.png')
    else:
        result_bg = tkinter.PhotoImage(file='ext_img/bg_result2.png')
    canvas_res.create_image(canvas_width // 2, canvas_height // 2, image=result_bg)

    global image_list, image_list2
    global res_img11, res_img12, res_img13, res_img14, res_img15, res_img21, res_img22, res_img23, res_img31, res_img32, res_img33, res_txtbbg, res_imgbbg, wep_combopicker, jobup_select, res_txt_weapon

    wep_index = ""

    if job_type == 'deal':  ###########################

        global result_image_on, rank_dam, rank_stat, rank_stat2, req_cool, res_dam, res_stat, res_stat2
        rank_baibiaoguai = [0, 0, 0, 0, 0]
        rank_not_owned_equips = [0, 0, 0, 0, 0]
        rank_dam = [0, 0, 0, 0, 0]
        rank_setting = [0, 0, 0, 0, 0]
        rss = [0, 0, 0, 0, 0]
        result_image_on = [{}, {}, {}, {}, {}]
        try:
            # rank => [score, [calc_wep, base_array, baibianguai, not_owned_equips]]
            rank_baibiaoguai[0] = rank_list[0][1][2]
            rank_not_owned_equips[0] = rank_list[0][1][3]
            rank_dam[0] = str(int(100 * rank_list[0][0])) + "%"
            rank_setting[0] = rank_list[0][1][0]  ##0号是排名
            rss[0] = rank_list[0][1][1]
            for i in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                for j in rank_setting[0]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on[0][str(i)] = image_list[j]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips[0]:
                                result_image_on[0][str(i)] = image_list2[j]
            if rank_baibiaoguai[0] is not None:
                result_image_on[0]["bbg"] = image_list[rank_baibiaoguai[0]]

            rank_baibiaoguai[1] = rank_list[1][1][2]
            rank_not_owned_equips[1] = rank_list[1][1][3]
            rank_dam[1] = str(int(100 * rank_list[1][0])) + "%"
            rank_setting[1] = rank_list[1][1][0]
            rss[1] = rank_list[1][1][1]
            for i in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                for j in rank_setting[1]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on[1][str(i)] = image_list[j]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips[1]:
                                result_image_on[1][str(i)] = image_list2[j]
            if rank_baibiaoguai[1] is not None:
                result_image_on[1]["bbg"] = image_list[rank_baibiaoguai[1]]

            rank_baibiaoguai[2] = rank_list[2][1][2]
            rank_not_owned_equips[2] = rank_list[2][1][3]
            rank_dam[2] = str(int(100 * rank_list[2][0])) + "%"
            rank_setting[2] = rank_list[2][1][0]
            rss[2] = rank_list[2][1][1]
            for i in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                for j in rank_setting[2]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on[2][str(i)] = image_list[j]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips[2]:
                                result_image_on[2][str(i)] = image_list2[j]
            if rank_baibiaoguai[2] is not None:
                result_image_on[2]["bbg"] = image_list[rank_baibiaoguai[2]]

            rank_baibiaoguai[3] = rank_list[3][1][2]
            rank_not_owned_equips[3] = rank_list[3][1][3]
            rank_dam[3] = str(int(100 * rank_list[3][0])) + "%"
            rank_setting[3] = rank_list[3][1][0]
            rss[3] = rank_list[3][1][1]
            for i in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                for j in rank_setting[3]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on[3][str(i)] = image_list[j]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips[3]:
                                result_image_on[3][str(i)] = image_list2[j]
            if rank_baibiaoguai[3] is not None:
                result_image_on[3]["bbg"] = image_list[rank_baibiaoguai[3]]

            rank_baibiaoguai[4] = rank_list[4][1][2]
            rank_not_owned_equips[4] = rank_list[4][1][3]
            rank_dam[4] = str(int(100 * rank_list[4][0])) + "%"
            rank_setting[4] = rank_list[4][1][0]
            rss[4] = rank_list[4][1][1]
            for i in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                for j in rank_setting[4]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on[4][str(i)] = image_list[j]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips[4]:
                                result_image_on[4][str(i)] = image_list2[j]
            if rank_baibiaoguai[4] is not None:
                result_image_on[4]["bbg"] = image_list[rank_baibiaoguai[4]]
        except IndexError as error:
            c = 1

        # 0추스탯 1추공 2증 3크 4추
        # 6모 7공 8스탯 9속강 10지속 11스증 12특수
        # 13공속 14크확 / 15 특수액티브 / 16~19 패시브 /20 쿨감보정/21 二觉캐특수액티브 /22~27 액티브레벨링
        rank_stat = [0, 0, 0, 0, 0]
        rank_stat2 = [0, 0, 0, 0, 0]
        for i in range(0, 5):
            try:
                rank_stat[i] = ("增伤= " + str(int(rss[i][index_extra_percent_attack_damage])) +
                                "%\n爆伤= " + str(int(rss[i][index_extra_percent_crit_damage])) +
                                "%\n白字= " + str(int(rss[i][index_extra_percent_addtional_damage])) +
                                "%\n所攻= " + str(int(rss[i][index_extra_percent_final_damage])) +
                                "%\n三攻= " + str(int(rss[i][index_extra_percent_physical_magical_independent_attack_power])) +
                                "%\n力智= " + str(int(rss[i][index_extra_percent_strength_and_intelligence])) +
                                "%\n属强= " + str(int(rss[i][index_extra_all_element_strength])) +
                                "\n持续= " + str(int(rss[i][index_extra_percent_continued_damage])) +
                                "%\n技攻= " + str(int(rss[i][index_extra_percent_skill_attack_power])) +
                                "%\n特殊= " + str(int(rss[i][index_extra_percent_special_effect])) +
                                "%\n\n攻速= " + str(int(rss[i][index_extra_percent_attack_speed])) +
                                "%\n暴击率= " + str(int(rss[i][index_extra_percent_magic_physical_crit_rate])) + "%")
                rank_stat2[i] = ("   <主动>"
                                 "\n  1~45技能= " + str(round(rss[i][index_extra_active_skill_lv_1_45], 1)) + "级" +
                                 "\n    50技能= " + str(int(rss[i][index_extra_active_skill_lv_50])) + "级" +
                                 "\n 60~80技能= " + str(round(rss[i][index_extra_active_skill_lv_60_80], 1)) + "级" +
                                 "\n    85技能= " + str(int(rss[i][index_extra_active_skill_lv_85])) + "级" +
                                 "\n    95技能= " + str(int(rss[i][index_extra_active_skill_lv_95])) + "级" +
                                 "\n   100技能= " + str(int(rss[i][index_extra_active_skill_lv_100])) + "级" +
                                 "\n\n   <被动>" +
                                 "\n  转职被动= " + str(round(rss[i][index_extra_passive_transfer_skill], 1)) + "级" +
                                 "\n  一觉被动= " + str(int(rss[i][index_extra_passive_first_awaken_skill])) + "级" +
                                 "\n  二觉被动= " + str(int(rss[i][index_extra_passive_second_awaken_skill])) + "级" +
                                 "\n  三觉被动= " + str(int(rss[i][index_extra_passive_third_awaken_skill])) + "级")
            except TypeError as error:
                c = 1

        cool_check = req_cool.get()
        cool_txt = ""
        if cool_check == "O(打开)":
            cool_txt = "冷却补正"
        else:
            cool_txt = "纯伤害"
        canvas_res.create_text(122, 145, text=cool_txt, font=guide_font, fill='white')
        if int(ele_skill) != 0:
            canvas_res.create_text(122, 170, text="技能属强补正=" + str(int(ele_skill)) + " / 逆校正%=" + str(
                round(100 * (1.05 / (1.05 + int(ele_skill) * 0.0045) - 1), 1)) + "%", font=guide_font, fill='white')
        res_dam = canvas_res.create_text(122, 125, text=rank_dam[0], font=mid_font, fill='white')
        res_stat = canvas_res.create_text(65, 293, text=rank_stat[0], fill='white')
        res_stat2 = canvas_res.create_text(185, 293, text=rank_stat2[0], fill='white')

        res_img11 = canvas_res.create_image(57, 57, image=result_image_on[0]['11'])  # 上衣
        res_img12 = canvas_res.create_image(27, 87, image=result_image_on[0]['12'])  # 裤子
        res_img13 = canvas_res.create_image(27, 57, image=result_image_on[0]['13'])  # 头肩
        res_img14 = canvas_res.create_image(57, 87, image=result_image_on[0]['14'])  # 腰带
        res_img15 = canvas_res.create_image(27, 117, image=result_image_on[0]['15'])  # 鞋子
        res_img21 = canvas_res.create_image(189, 57, image=result_image_on[0]['21'])  # 手镯
        res_img22 = canvas_res.create_image(219, 57, image=result_image_on[0]['22'])  # 项链
        res_img23 = canvas_res.create_image(219, 87, image=result_image_on[0]['23'])  # 戒指
        res_img31 = canvas_res.create_image(189, 87, image=result_image_on[0]['31'])  # 辅助装备
        res_img32 = canvas_res.create_image(219, 117, image=result_image_on[0]['32'])  # 魔法石
        res_img33 = canvas_res.create_image(189, 117, image=result_image_on[0]['33'])  # 耳环
        if 'bbg' in result_image_on[0]:
            res_txtbbg = canvas_res.create_text(178, 147, text="百变怪=>", font=guide_font, fill='white')
            res_imgbbg = canvas_res.create_image(219, 147, image=result_image_on[0]['bbg'])  # 百变怪
        cn1 = 0
        for j in range(0, 5):
            try:
                for i in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                    canvas_res.create_image(268 + cn1 * 29, 67 + 78 * j, image=result_image_on[j][str(i)])
                    cn1 = cn1 + 1
                if 'bbg' in result_image_on[j]:
                    canvas_res.create_text(268 + 5 * 29 + 14, 38 + 78 * j, text="百变怪=>", font=guide_font, fill='white')
                    canvas_res.create_image(268 + 7 * 29, 37 + 78 * j, image=result_image_on[j]['bbg'])
                cn1 = 0
                canvas_res.create_text(346, 34 + 78 * j, text=rank_dam[j], font=mid_font, fill='white')
            except KeyError as error:
                c = 1

        weapon = rank_setting[0][0]
        equips = rank_setting[0][1:]
        change_readable_result_area(weapon, equips, True)

        wep_index = weapon

        g_rank_equips = {}
        for rank in range(0, 5):
            g_rank_equips[rank] = rank_setting[rank]

        length = len(rank_list)

    elif job_type == 'buf':  ##########################
        load_presetr = load_workbook("preset.xlsx", data_only=True)
        r_preset = load_presetr["custom"]
        global result_image_on1, result_image_on2, result_image_on3, rank_buf1, rank_buf2, rank_buf3, rank_type_buf, res_buf, res_img_list, res_buf_list, res_buf_ex1, res_buf_ex2, res_buf_ex3, rank_buf_ex1, rank_buf_ex2, rank_buf_ex3, res_buf_type_what
        rank_type_buf = 3
        rank_baibiaoguai1 = [0, 0, 0, 0, 0]
        rank_baibiaoguai2 = [0, 0, 0, 0, 0]
        rank_baibiaoguai3 = [0, 0, 0, 0, 0]
        rank_not_owned_equips1 = [0, 0, 0, 0, 0]
        rank_not_owned_equips2 = [0, 0, 0, 0, 0]
        rank_not_owned_equips3 = [0, 0, 0, 0, 0]
        rank_setting1 = [0, 0, 0, 0, 0]
        rank_setting2 = [0, 0, 0, 0, 0]
        rank_setting3 = [0, 0, 0, 0, 0]
        result_image_on1 = [{}, {}, {}, {}, {}]
        result_image_on2 = [{}, {}, {}, {}, {}]
        result_image_on3 = [{}, {}, {}, {}, {}]
        rank_buf1 = [0, 0, 0, 0, 0]
        rank_buf2 = [0, 0, 0, 0, 0]
        rank_buf3 = [0, 0, 0, 0, 0]
        rank_buf_ex1 = [0, 0, 0, 0, 0]
        rank_buf_ex2 = [0, 0, 0, 0, 0]
        rank_buf_ex3 = [0, 0, 0, 0, 0]
        ## rank_setting[rank]=rank_list[a][rank][b][c]
        ## a: 0=祝福,1=크오,2=합계
        ## b: 0=계수,1=스펙or증가량
        ## c: b에서 1 선택시, 0=스펙, 1=증가량
        try:
            # ranking = [ranking1, ranking2, ranking3]
            # ranking1 = rank => [score, [calc_wep, [save1, save2, pas1_out], baibianguai, not_owned_equips]]
            rank_baibiaoguai3[0] = rank_list[2][0][1][2]
            rank_baibiaoguai2[0] = rank_list[1][0][1][2]
            rank_baibiaoguai1[0] = rank_list[0][0][1][2]
            rank_not_owned_equips3[0] = rank_list[2][0][1][3]
            rank_not_owned_equips2[0] = rank_list[1][0][1][3]
            rank_not_owned_equips1[0] = rank_list[0][0][1][3]
            rank_setting3[0] = rank_list[2][0][1][0]  ##2번째 숫자가 랭킹임
            rank_setting2[0] = rank_list[1][0][1][0]
            rank_setting1[0] = rank_list[0][0][1][0]
            rank_buf3[0] = int(rank_list[2][0][0] / 10)
            rank_buf2[0] = int(rank_list[1][0][0] / 10)
            rank_buf1[0] = int(rank_list[0][0][0] / 10)
            rank_buf_ex3[0] = rank_list[2][0][1][1]
            rank_buf_ex2[0] = rank_list[1][0][1][1]
            rank_buf_ex1[0] = rank_list[0][0][1][1]
            for i in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                for j in rank_setting3[0]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on3[0][str(i)] = image_list[j]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips3[0]:
                                result_image_on3[0][str(i)] = image_list2[j]
                for j in rank_setting2[0]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on2[0][str(i)] = image_list[j]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips2[0]:
                                result_image_on2[0][str(i)] = image_list2[j]
                for j in rank_setting1[0]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on1[0][str(i)] = image_list[j]  ##
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips1[0]:
                                result_image_on1[0][str(i)] = image_list2[j]
            if rank_baibiaoguai3[0] is not None:
                result_image_on3[0]["bbg"] = image_list[rank_baibiaoguai3[0]]
            if rank_baibiaoguai2[0] is not None:
                result_image_on2[0]["bbg"] = image_list[rank_baibiaoguai2[0]]
            if rank_baibiaoguai1[0] is not None:
                result_image_on1[0]["bbg"] = image_list[rank_baibiaoguai1[0]]

            rank_baibiaoguai3[1] = rank_list[2][1][1][2]
            rank_baibiaoguai2[1] = rank_list[1][1][1][2]
            rank_baibiaoguai1[1] = rank_list[0][1][1][2]
            rank_not_owned_equips3[1] = rank_list[2][1][1][3]
            rank_not_owned_equips2[1] = rank_list[1][1][1][3]
            rank_not_owned_equips1[1] = rank_list[0][1][1][3]
            rank_setting3[1] = rank_list[2][1][1][0]
            rank_setting2[1] = rank_list[1][1][1][0]
            rank_setting1[1] = rank_list[0][1][1][0]
            rank_buf3[1] = int(rank_list[2][1][0] / 10)
            rank_buf2[1] = int(rank_list[1][1][0] / 10)
            rank_buf1[1] = int(rank_list[0][1][0] / 10)
            rank_buf_ex3[1] = rank_list[2][1][1][1]
            rank_buf_ex2[1] = rank_list[1][1][1][1]
            rank_buf_ex1[1] = rank_list[0][1][1][1]
            for i in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                for j in rank_setting3[1]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on3[1][str(i)] = image_list[j]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips3[1]:
                                result_image_on3[1][str(i)] = image_list2[j]
                for j in rank_setting2[1]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on2[1][str(i)] = image_list[j]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips2[1]:
                                result_image_on2[1][str(i)] = image_list2[j]
                for j in rank_setting1[1]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on1[1][str(i)] = image_list[j]  ##
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips1[1]:
                                result_image_on1[1][str(i)] = image_list2[j]
            if rank_baibiaoguai3[1] is not None:
                result_image_on3[1]["bbg"] = image_list[rank_baibiaoguai3[1]]
            if rank_baibiaoguai2[1] is not None:
                result_image_on2[1]["bbg"] = image_list[rank_baibiaoguai2[1]]
            if rank_baibiaoguai1[1] is not None:
                result_image_on1[1]["bbg"] = image_list[rank_baibiaoguai1[1]]

            rank_baibiaoguai3[2] = rank_list[2][2][1][2]
            rank_baibiaoguai2[2] = rank_list[1][2][1][2]
            rank_baibiaoguai1[2] = rank_list[0][2][1][2]
            rank_not_owned_equips3[2] = rank_list[2][2][1][3]
            rank_not_owned_equips2[2] = rank_list[1][2][1][3]
            rank_not_owned_equips1[2] = rank_list[0][2][1][3]
            rank_setting3[2] = rank_list[2][2][1][0]
            rank_setting2[2] = rank_list[1][2][1][0]
            rank_setting1[2] = rank_list[0][2][1][0]
            rank_buf3[2] = int(rank_list[2][2][0] / 10)
            rank_buf2[2] = int(rank_list[1][2][0] / 10)
            rank_buf1[2] = int(rank_list[0][2][0] / 10)
            rank_buf_ex3[2] = rank_list[2][2][1][1]
            rank_buf_ex2[2] = rank_list[1][2][1][1]
            rank_buf_ex1[2] = rank_list[0][2][1][1]
            for i in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                for j in rank_setting3[2]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on3[2][str(i)] = image_list[j]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips3[2]:
                                result_image_on3[2][str(i)] = image_list2[j]
                for j in rank_setting2[2]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on2[2][str(i)] = image_list[j]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips2[2]:
                                result_image_on2[2][str(i)] = image_list2[j]
                for j in rank_setting1[2]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on1[2][str(i)] = image_list[j]  ##
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips1[1]:
                                result_image_on1[1][str(i)] = image_list2[j]
            if rank_baibiaoguai3[2] is not None:
                result_image_on3[2]["bbg"] = image_list[rank_baibiaoguai3[2]]
            if rank_baibiaoguai2[2] is not None:
                result_image_on2[2]["bbg"] = image_list[rank_baibiaoguai2[2]]
            if rank_baibiaoguai1[2] is not None:
                result_image_on1[2]["bbg"] = image_list[rank_baibiaoguai1[2]]

            rank_baibiaoguai3[3] = rank_list[2][3][1][2]
            rank_baibiaoguai2[3] = rank_list[1][3][1][2]
            rank_baibiaoguai1[3] = rank_list[0][3][1][2]
            rank_not_owned_equips3[3] = rank_list[2][3][1][3]
            rank_not_owned_equips2[3] = rank_list[1][3][1][3]
            rank_not_owned_equips1[3] = rank_list[0][3][1][3]
            rank_setting3[3] = rank_list[2][3][1][0]
            rank_setting2[3] = rank_list[1][3][1][0]
            rank_setting1[3] = rank_list[0][3][1][0]
            rank_buf3[3] = int(rank_list[2][3][0] / 10)
            rank_buf2[3] = int(rank_list[1][3][0] / 10)
            rank_buf1[3] = int(rank_list[0][3][0] / 10)
            rank_buf_ex3[3] = rank_list[2][3][1][1]
            rank_buf_ex2[3] = rank_list[1][3][1][1]
            rank_buf_ex1[3] = rank_list[0][3][1][1]
            for i in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                for j in rank_setting3[3]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on3[3][str(i)] = image_list[j]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips3[3]:
                                result_image_on3[3][str(i)] = image_list2[j]
                for j in rank_setting2[3]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on2[3][str(i)] = image_list[j]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips2[3]:
                                result_image_on2[3][str(i)] = image_list2[j]
                for j in rank_setting1[3]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on1[3][str(i)] = image_list[j]  ##
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips1[3]:
                                result_image_on1[3][str(i)] = image_list2[j]
            if rank_baibiaoguai3[3] is not None:
                result_image_on3[3]["bbg"] = image_list[rank_baibiaoguai3[3]]
            if rank_baibiaoguai2[3] is not None:
                result_image_on2[3]["bbg"] = image_list[rank_baibiaoguai2[3]]
            if rank_baibiaoguai1[3] is not None:
                result_image_on1[3]["bbg"] = image_list[rank_baibiaoguai1[3]]

            rank_baibiaoguai3[4] = rank_list[2][4][1][2]
            rank_baibiaoguai2[4] = rank_list[1][4][1][2]
            rank_baibiaoguai1[4] = rank_list[0][4][1][2]
            rank_not_owned_equips3[4] = rank_list[2][4][1][3]
            rank_not_owned_equips2[4] = rank_list[1][4][1][3]
            rank_not_owned_equips1[4] = rank_list[0][4][1][3]
            rank_setting3[4] = rank_list[2][4][1][0]
            rank_setting2[4] = rank_list[1][4][1][0]
            rank_setting1[4] = rank_list[0][4][1][0]
            rank_buf3[4] = int(rank_list[2][4][0] / 10)
            rank_buf2[4] = int(rank_list[1][4][0] / 10)
            rank_buf1[4] = int(rank_list[0][4][0] / 10)
            rank_buf_ex3[4] = rank_list[2][4][1][1]
            rank_buf_ex2[4] = rank_list[1][4][1][1]
            rank_buf_ex1[4] = rank_list[0][4][1][1]
            for i in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                for j in rank_setting3[4]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on3[4][str(i)] = image_list[j]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips3[4]:
                                result_image_on3[4][str(i)] = image_list2[j]
                for j in rank_setting2[4]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on2[4][str(i)] = image_list[j]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips2[4]:
                                result_image_on2[4][str(i)] = image_list2[j]
                for j in rank_setting1[4]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on1[4][str(i)] = image_list[j]  ##
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if j in rank_not_owned_equips1[4]:
                                result_image_on1[4][str(i)] = image_list2[j]
            if rank_baibiaoguai3[4] is not None:
                result_image_on3[4]["bbg"] = image_list[rank_baibiaoguai3[4]]
            if rank_baibiaoguai2[4] is not None:
                result_image_on2[4]["bbg"] = image_list[rank_baibiaoguai2[4]]
            if rank_baibiaoguai1[4] is not None:
                result_image_on1[4]["bbg"] = image_list[rank_baibiaoguai1[4]]
        except IndexError as error:
            c = 1

        canvas_res.create_text(122, 193, text="自定义 祝福+" + str(
            int(r_preset['H2'].value) + int(r_preset['H4'].value) + int(
                r_preset['H5'].value)) + "级 / " + "自定义 一觉+" + str(int(r_preset['H3'].value)) + "级\n祝福数据+" + str(
            int(r_preset['H6'].value)) + " / 一觉数据+" + str(int(r_preset['H1'].value)), font=guide_font, fill='white')

        res_buf = canvas_res.create_text(122, 125, text=rank_buf3[0], font=mid_font, fill='white')
        res_buf_type_what = canvas_res.create_text(122, 145, text="综合标准", font=guide_font, fill='white')
        res_buf_ex1 = canvas_res.create_text(123, 247, text="祝福=" + rank_buf_ex3[0][0], font=guide_font, fill='white')
        res_buf_ex2 = canvas_res.create_text(123 - 17, 282, text="一觉=" + rank_buf_ex3[0][1], font=guide_font,
                                             fill='white')
        res_buf_ex3 = canvas_res.create_text(123 - 44, 312, text="一觉패=" + rank_buf_ex3[0][2], font=guide_font,
                                             fill='white')

        res_img11 = canvas_res.create_image(57, 57, image=result_image_on3[0]['11'])
        res_img12 = canvas_res.create_image(27, 87, image=result_image_on3[0]['12'])
        res_img13 = canvas_res.create_image(27, 57, image=result_image_on3[0]['13'])
        res_img14 = canvas_res.create_image(57, 87, image=result_image_on3[0]['14'])
        res_img15 = canvas_res.create_image(27, 117, image=result_image_on3[0]['15'])
        res_img21 = canvas_res.create_image(189, 57, image=result_image_on3[0]['21'])
        res_img22 = canvas_res.create_image(219, 57, image=result_image_on3[0]['22'])
        res_img23 = canvas_res.create_image(219, 87, image=result_image_on3[0]['23'])
        res_img31 = canvas_res.create_image(189, 87, image=result_image_on3[0]['31'])
        res_img32 = canvas_res.create_image(219, 117, image=result_image_on3[0]['32'])
        res_img33 = canvas_res.create_image(189, 117, image=result_image_on3[0]['33'])
        if 'bbg' in result_image_on3[0]:
            res_txtbbg = canvas_res.create_text(178, 147, text="百变怪=>", font=guide_font, fill='white')
            res_imgbbg = canvas_res.create_image(219, 147, image=result_image_on3[0]['bbg'])  # 百变怪
        cn1 = 0
        res_img_list = {}
        res_buf_list = {}
        for j in range(0, 5):
            try:
                for i in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                    temp_res = canvas_res.create_image(268 + cn1 * 29, 67 + 78 * j, image=result_image_on3[j][str(i)])
                    res_img_list[str(j) + str(i)] = temp_res
                    cn1 = cn1 + 1
                if 'bbg' in result_image_on3[j]:
                    canvas_res.create_text(268 + 5 * 29 + 14, 38 + 78 * j, text="百变怪=>", font=guide_font, fill='white')
                    canvas_res.create_image(268 + 7 * 29, 37 + 78 * j, image=result_image_on3[j]['bbg'])
                cn1 = 0
                temp_buf = canvas_res.create_text(346, 34 + 78 * j, text=rank_buf3[j], font=mid_font, fill='white')
                res_buf_list[j] = temp_buf
            except KeyError as error:
                c = 1
        length = len(rank_list[0])
        type1_img = tkinter.PhotoImage(file='ext_img/type_bless.png')
        type2_img = tkinter.PhotoImage(file='ext_img/type_crux.png')
        type3_img = tkinter.PhotoImage(file='ext_img/type_all.png')
        rank_type_but1 = tkinter.Button(result_window, command=lambda: change_rank_type(1), image=type1_img,
                                        bg=dark_main, borderwidth=0, activebackground=dark_main);
        rank_type_but1.place(x=8, y=337)
        rank_type_but2 = tkinter.Button(result_window, command=lambda: change_rank_type(2), image=type2_img,
                                        bg=dark_main, borderwidth=0, activebackground=dark_main);
        rank_type_but2.place(x=84, y=337)
        rank_type_but3 = tkinter.Button(result_window, command=lambda: change_rank_type(3), image=type3_img,
                                        bg=dark_main, borderwidth=0, activebackground=dark_main);
        rank_type_but3.place(x=160, y=337)
        rank_type_but1.image = type1_img
        rank_type_but2.image = type2_img
        rank_type_but3.image = type3_img

        weapon = rank_setting3[0][0]
        equips = rank_setting3[0][1:]
        change_readable_result_area(weapon, equips, True)

        wep_index = weapon

        g_rank_equips = {}
        g_current_buff_type = "综合"
        for buf_type, rank_setting in [("祝福", rank_setting1), ("一觉", rank_setting2), ("综合", rank_setting3)]:
            ranks = {}
            for rank in range(0, 5):
                ranks[rank] = rank_setting[rank]
            g_rank_equips[buf_type] = ranks

        load_presetr.close()

    wep_name = equip_index_to_realname[wep_index]
    job_name = jobup_select.get()
    res_txt_weapon = canvas_res.create_text(122, 20, text=wep_name, font=guide_font, fill='white')
    canvas_res.create_text(122, 50, text="<职业>", font=guide_font, fill='white')
    canvas_res.create_text(122, 87, text=job_name, font=guide_font, fill='white')

    show_name_img = tkinter.PhotoImage(file='ext_img/show_name.png')
    res_bt_show_name = tkinter.Button(result_window, command=lambda: show_name(), image=show_name_img,
                                      bg=dark_blue, borderwidth=0, activebackground=dark_blue);
    res_bt_show_name.place(x=8, y=135)

    show_detail_img = tkinter.PhotoImage(file='ext_img/show_detail.png')

    res_bt1 = tkinter.Button(result_window, command=lambda: change_rank(0, job_type), image=show_detail_img,
                             bg=dark_blue, borderwidth=0, activebackground=dark_blue);
    res_bt1.place(x=486, y=20 + 78 * 0)
    res_bt2 = tkinter.Button(result_window, command=lambda: change_rank(1, job_type), image=show_detail_img,
                             bg=dark_blue, borderwidth=0, activebackground=dark_blue)
    res_bt3 = tkinter.Button(result_window, command=lambda: change_rank(2, job_type), image=show_detail_img,
                             bg=dark_blue, borderwidth=0, activebackground=dark_blue)
    res_bt4 = tkinter.Button(result_window, command=lambda: change_rank(3, job_type), image=show_detail_img,
                             bg=dark_blue, borderwidth=0, activebackground=dark_blue)
    res_bt5 = tkinter.Button(result_window, command=lambda: change_rank(4, job_type), image=show_detail_img,
                             bg=dark_blue, borderwidth=0, activebackground=dark_blue)
    if length > 1:
        res_bt2.place(x=486, y=20 + 78 * 1)
    if length > 2:
        res_bt3.place(x=486, y=20 + 78 * 2)
    if length > 3:
        res_bt4.place(x=486, y=20 + 78 * 3)
    if length > 4:
        res_bt5.place(x=486, y=20 + 78 * 4)

    canvas_res.image = result_bg
    res_bt1.image = show_detail_img
    res_bt_show_name.image = show_name_img


def change_rank(now, job_type):
    global g_current_rank
    g_current_rank = now

    global image_list, canvas_res, res_img11, res_img12, res_img13, res_img14, res_img15, res_img21, res_img22, res_img23, res_img31, res_img32, res_img33, res_txtbbg, res_imgbbg
    if job_type == 'deal':
        global res_dam, res_stat, res_stat2, rank_stat, rank_stat2, result_image_on
        try:
            image_changed = result_image_on[now]
            canvas_res.itemconfig(res_img11, image=image_changed['11'])
            canvas_res.itemconfig(res_img12, image=image_changed['12'])
            canvas_res.itemconfig(res_img13, image=image_changed['13'])
            canvas_res.itemconfig(res_img14, image=image_changed['14'])
            canvas_res.itemconfig(res_img15, image=image_changed['15'])
            canvas_res.itemconfig(res_img21, image=image_changed['21'])
            canvas_res.itemconfig(res_img22, image=image_changed['22'])
            canvas_res.itemconfig(res_img23, image=image_changed['23'])
            canvas_res.itemconfig(res_img31, image=image_changed['31'])
            canvas_res.itemconfig(res_img32, image=image_changed['32'])
            canvas_res.itemconfig(res_img33, image=image_changed['33'])
            if 'res_txtbbg' in globals() and res_txtbbg is not None:
                canvas_res.delete(res_txtbbg)
            if 'res_imgbbg' in globals() and res_imgbbg is not None:
                canvas_res.delete(res_imgbbg)
            if 'bbg' in image_changed:
                res_txtbbg = canvas_res.create_text(178, 147, text="百变怪=>", fill='white')
                res_imgbbg = canvas_res.create_image(219, 147, image=image_changed['bbg'])  # 百变怪
            canvas_res.itemconfig(res_dam, text=rank_dam[now])
            canvas_res.itemconfig(res_stat, text=rank_stat[now])
            canvas_res.itemconfig(res_stat2, text=rank_stat2[now])

            current_weapon = g_rank_equips[now][0]
            current_equips = g_rank_equips[now][1:]
            canvas_res.itemconfig(res_txt_weapon, text=equip_index_to_realname[current_weapon])
            change_readable_result_area(current_weapon, current_equips, False)
        except KeyError as error:
            c = 1

    elif job_type == 'buf':
        global result_image_on1, result_image_on2, result_image_on3, rank_buf1, rank_buf2, rank_buf3, rank_type_buf, res_buf, res_buf_ex1, res_buf_ex2, res_buf_ex3, rank_buf_ex1, rank_buf_ex2, rank_buf_ex3
        try:
            if rank_type_buf == 1:
                image_changed = result_image_on1[now]
                rank_changed = rank_buf1[now]
                rank_buf_ex_changed = rank_buf_ex1
            elif rank_type_buf == 2:
                image_changed = result_image_on2[now]
                rank_changed = rank_buf2[now]
                rank_buf_ex_changed = rank_buf_ex2
            elif rank_type_buf == 3:
                image_changed = result_image_on3[now]
                rank_changed = rank_buf3[now]
                rank_buf_ex_changed = rank_buf_ex3
            canvas_res.itemconfig(res_buf, text=rank_changed)
            canvas_res.itemconfig(res_buf_ex1, text="祝福=" + rank_buf_ex_changed[now][0])
            canvas_res.itemconfig(res_buf_ex2, text="一觉=" + rank_buf_ex_changed[now][1])
            canvas_res.itemconfig(res_buf_ex3, text="一觉패=" + rank_buf_ex_changed[now][2])
            canvas_res.itemconfig(res_img11, image=image_changed['11'])
            canvas_res.itemconfig(res_img12, image=image_changed['12'])
            canvas_res.itemconfig(res_img13, image=image_changed['13'])
            canvas_res.itemconfig(res_img14, image=image_changed['14'])
            canvas_res.itemconfig(res_img15, image=image_changed['15'])
            canvas_res.itemconfig(res_img21, image=image_changed['21'])
            canvas_res.itemconfig(res_img22, image=image_changed['22'])
            canvas_res.itemconfig(res_img23, image=image_changed['23'])
            canvas_res.itemconfig(res_img31, image=image_changed['31'])
            canvas_res.itemconfig(res_img32, image=image_changed['32'])
            canvas_res.itemconfig(res_img33, image=image_changed['33'])
            if 'res_txtbbg' in globals() and res_txtbbg is not None:
                canvas_res.delete(res_txtbbg)
            if 'res_imgbbg' in globals() and res_imgbbg is not None:
                canvas_res.delete(res_imgbbg)
            if 'bbg' in image_changed:
                res_txtbbg = canvas_res.create_text(178, 147, text="百变怪=>", fill='white')
                res_imgbbg = canvas_res.create_image(219, 147, image=image_changed['bbg'])  # 百变怪

            current_weapon = g_rank_equips[g_current_buff_type][now][0]
            current_equips = g_rank_equips[g_current_buff_type][now][1:]
            canvas_res.itemconfig(res_txt_weapon, text=equip_index_to_realname[current_weapon])
            change_readable_result_area(current_weapon, current_equips, False)

        except KeyError as error:
            c = 1


def change_rank_type(in_type):
    global g_current_rank
    g_current_rank = 0
    global g_current_buff_type
    global image_list, canvas_res, res_img11, res_img12, res_img13, res_img14, res_img15, res_img21, res_img22, res_img23, res_img31, res_img32, res_img33, res_txtbbg, res_imgbbg
    global result_image_on1, result_image_on2, result_image_on3, rank_buf1, rank_buf2, rank_buf3, rank_type_buf, res_img_list, res_buf_list, res_buf_ex1, res_buf_ex2, res_buf_ex3, rank_buf_ex1, rank_buf_ex2, rank_buf_ex3, res_buf_type_what
    if in_type == 1:
        rank_type_buf = 1
        image_changed = result_image_on1[0]
        image_changed_all = result_image_on1
        rank_changed = rank_buf1
        rank_buf_ex_changed = rank_buf_ex1
        type_changed = "祝福标准"
        g_current_buff_type = "祝福"
    elif in_type == 2:
        rank_type_buf = 2
        image_changed = result_image_on2[0]
        image_changed_all = result_image_on2
        rank_changed = rank_buf2
        rank_buf_ex_changed = rank_buf_ex2
        type_changed = "一觉标准"
        g_current_buff_type = "一觉"
    elif in_type == 3:
        rank_type_buf = 3
        image_changed = result_image_on3[0]
        image_changed_all = result_image_on3
        rank_changed = rank_buf3
        rank_buf_ex_changed = rank_buf_ex3
        type_changed = "综合标准"
        g_current_buff_type = "综合"
    canvas_res.itemconfig(res_buf_type_what, text=type_changed)
    canvas_res.itemconfig(res_buf_ex1, text="祝福=" + rank_buf_ex_changed[0][0])
    canvas_res.itemconfig(res_buf_ex2, text="一觉=" + rank_buf_ex_changed[0][1])
    canvas_res.itemconfig(res_buf_ex3, text="一觉패=" + rank_buf_ex_changed[0][2])
    canvas_res.itemconfig(res_buf, text=rank_changed[0])
    canvas_res.itemconfig(res_img11, image=image_changed['11'])
    canvas_res.itemconfig(res_img12, image=image_changed['12'])
    canvas_res.itemconfig(res_img13, image=image_changed['13'])
    canvas_res.itemconfig(res_img14, image=image_changed['14'])
    canvas_res.itemconfig(res_img15, image=image_changed['15'])
    canvas_res.itemconfig(res_img21, image=image_changed['21'])
    canvas_res.itemconfig(res_img22, image=image_changed['22'])
    canvas_res.itemconfig(res_img23, image=image_changed['23'])
    canvas_res.itemconfig(res_img31, image=image_changed['31'])
    canvas_res.itemconfig(res_img32, image=image_changed['32'])
    canvas_res.itemconfig(res_img33, image=image_changed['33'])
    if 'res_txtbbg' in globals() and res_txtbbg is not None:
        canvas_res.delete(res_txtbbg)
    if 'res_imgbbg' in globals() and res_imgbbg is not None:
        canvas_res.delete(res_imgbbg)
    if 'bbg' in image_changed:
        res_txtbbg = canvas_res.create_text(178, 147, text="百变怪=>", fill='white')
        res_imgbbg = canvas_res.create_image(219, 147, image=image_changed['bbg'])  # 百变怪
    cn2 = 0
    for j in range(0, 5):
        try:
            for i in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                canvas_res.itemconfig(res_img_list[str(j) + str(i)], image=image_changed_all[j][str(i)])
                cn2 = cn2 + 2
            if 'bbg' in image_changed_all[j]:
                canvas_res.create_text(268 + 5 * 29 + 14, 38 + 78 * j, text="百变怪=>", font=guide_font, fill='white')
                canvas_res.create_image(268 + 7 * 29, 37 + 78 * j, image=image_changed_all[j]['bbg'])
            cn2 = 0
            canvas_res.itemconfig(res_buf_list[j], text=rank_changed[j], font=mid_font, fill='white')
        except KeyError as error:
            c = 1

    current_weapon = g_rank_equips[g_current_buff_type][0][0]
    current_equips = g_rank_equips[g_current_buff_type][0][1:]
    canvas_res.itemconfig(res_txt_weapon, text=equip_index_to_realname[current_weapon])
    change_readable_result_area(current_weapon, current_equips, False)


def costum():
    global custom_window
    custom_window = tkinter.Toplevel(self)
    custom_window.attributes("-topmost", True)
    custom_window.geometry("620x400+750+20")

    load_preset = load_workbook("preset.xlsx", data_only=True)
    db_preset = load_preset["custom"]

    tkinter.Label(custom_window, text="<输出环境>", font=mid_font).place(x=100, y=10)
    tkinter.Label(custom_window, text="属性攻击=", font=guide_font).place(x=10, y=50)
    ele_list = ['火', '冰', '光', '暗']
    ele_type = tkinter.ttk.Combobox(custom_window, width=5, values=ele_list);
    ele_type.place(x=80, y=52)  ##
    ele_type.set(db_preset['B1'].value)
    tkinter.Label(custom_window, text="冷却补正比例=          %", font=guide_font).place(x=160, y=50)  ##Y11/Z11
    cool_con = tkinter.Entry(custom_window, width=5);
    cool_con.place(x=230, y=52)
    cool_con.insert(END, db_preset['B2'].value)

    tkinter.Label(custom_window, text="<特殊装备补正>", font=mid_font).place(x=100, y=85)
    tkinter.Label(custom_window, text=" 输入窗口的数值会以对应百分比加成最终数值", fg="Red").place(x=30, y=120)
    tkinter.Label(custom_window, text="歧路腰带=          %", font=guide_font).place(x=160, y=155)  ##O164
    cus1 = tkinter.Entry(custom_window, width=5);
    cus1.place(x=230, y=157)
    cus1.insert(END, db_preset['B3'].value)
    tkinter.Label(custom_window, text="歧路鞋子=          %", font=guide_font).place(x=160, y=185)  ##O180
    cus2 = tkinter.Entry(custom_window, width=5);
    cus2.place(x=230, y=187)
    cus2.insert(END, db_preset['B4'].value)
    tkinter.Label(custom_window, text="经验等级=          ", font=guide_font).place(x=160, y=215)  ##G276
    lvl_list = ['传说↓', '英雄↑']
    cus3 = tkinter.ttk.Combobox(custom_window, width=5, values=lvl_list);
    cus3.place(x=230, y=217)
    cus3.set(db_preset['B12'].value)
    tkinter.Label(custom_window, text="恍惚增幅等级=          강", font=guide_font).place(x=160, y=245)
    lvl_list = [10, 11, 12, 13]
    cus4 = tkinter.ttk.Combobox(custom_window, width=2, values=lvl_list);
    cus4.place(x=230, y=247)
    cus4.set(db_preset['B13'].value)

    tkinter.Label(custom_window, text="不息上衣=          %", font=guide_font).place(x=10, y=155)  ##O100
    cus6 = tkinter.Entry(custom_window, width=5);
    cus6.place(x=80, y=157)
    cus6.insert(END, db_preset['B5'].value)
    tkinter.Label(custom_window, text="不息裤子=          %", font=guide_font).place(x=10, y=185)  ##O127
    cus7 = tkinter.Entry(custom_window, width=5);
    cus7.place(x=80, y=187)
    cus7.insert(END, db_preset['B6'].value)
    tkinter.Label(custom_window, text="不息护肩=          %", font=guide_font).place(x=10, y=215)  ##O147
    cus8 = tkinter.Entry(custom_window, width=5);
    cus8.place(x=80, y=217)
    cus8.insert(END, db_preset['B7'].value)
    tkinter.Label(custom_window, text="不息腰带=          %", font=guide_font).place(x=10, y=245)  ##O163
    cus9 = tkinter.Entry(custom_window, width=5);
    cus9.place(x=80, y=247)
    cus9.insert(END, db_preset['B8'].value)
    tkinter.Label(custom_window, text="不息鞋子=          %", font=guide_font).place(x=10, y=275)  ##O179
    cus10 = tkinter.Entry(custom_window, width=5);
    cus10.place(x=80, y=277)
    cus10.insert(END, db_preset['B9'].value)
    tkinter.Label(custom_window, text="不息2件套=           %", font=guide_font).place(x=10, y=305)  ##O295
    cus11 = tkinter.Entry(custom_window, width=5);
    cus11.place(x=80, y=307)
    cus11.insert(END, db_preset['B10'].value)
    tkinter.Label(custom_window, text="不息3件套=           %", font=guide_font).place(x=10, y=335)  ##O296,O297
    cus12 = tkinter.Entry(custom_window, width=5);
    cus12.place(x=80, y=337)
    cus12.insert(END, db_preset['B11'].value)

    tkinter.Label(custom_window, text="<奶量增幅相关>", font=mid_font, fg='blue').place(x=410, y=5)
    tkinter.Label(custom_window, text="补正辅助角色的表现", fg="Red").place(x=350, y=33)
    tkinter.Label(custom_window, text="面板体精智+          ", font=guide_font).place(x=320, y=75)  ##
    c_stat = tkinter.Entry(custom_window, width=7);
    c_stat.place(x=390, y=82)
    c_stat.insert(END, db_preset['H1'].value)
    tkinter.Label(custom_window, text="面板体精智+          ", font=guide_font).place(x=470, y=75)  ##
    b_stat = tkinter.Entry(custom_window, width=7);
    b_stat.place(x=540, y=82)
    b_stat.insert(END, db_preset['H6'].value)
    three = [0, 1, 2, 3];
    two = [0, 1, 2]
    tkinter.Label(custom_window, text="祝福称号=", font=guide_font).place(x=320, y=110)
    b_style_lvl = tkinter.ttk.Combobox(custom_window, width=5, values=three);
    b_style_lvl.place(x=390, y=112)  ##
    b_style_lvl.set(db_preset['H2'].value)
    tkinter.Label(custom_window, text="一觉称号=", font=guide_font).place(x=470, y=110)
    c_style_lvl = tkinter.ttk.Combobox(custom_window, width=5, values=two);
    c_style_lvl.place(x=540, y=112)  ##
    c_style_lvl.set(db_preset['H3'].value)
    tkinter.Label(custom_window, text="祝福等级=", font=guide_font).place(x=320, y=140)
    b_plt = tkinter.ttk.Combobox(custom_window, width=5, values=two);
    b_plt.place(x=390, y=142)  ##
    b_plt.set(db_preset['H4'].value)
    tkinter.Label(custom_window, text="祝福等级=", font=guide_font).place(x=470, y=140)
    b_cri = tkinter.ttk.Combobox(custom_window, width=5, values=[0, 1]);
    b_cri.place(x=540, y=142)  ##
    b_cri.set(db_preset['H5'].value)

    tkinter.Label(custom_window, text="<属强提升>", font=mid_font).place(x=410, y=175)
    tkinter.Label(custom_window, text="基础属强=", font=guide_font).place(x=470, y=210)
    ele1 = tkinter.Entry(custom_window, width=7);
    ele1.place(x=540, y=212)  ##
    ele1.insert(END, db_preset['B14'].value)
    tkinter.Label(custom_window, text="其他属强=", font=guide_font).place(x=470, y=240)
    ele2 = tkinter.Entry(custom_window, width=7);
    ele2.place(x=540, y=242)  ##
    ele2.insert(END, db_preset['B15'].value)
    tkinter.Label(custom_window, text="勋章属强=", font=guide_font).place(x=470, y=270)
    ele3 = tkinter.Entry(custom_window, width=7);
    ele3.place(x=540, y=272)  ##
    ele3.insert(END, db_preset['B16'].value)
    tkinter.Label(custom_window, text="技能属强= ", font=guide_font).place(x=320, y=210)
    ele4 = tkinter.Entry(custom_window, width=7);
    ele4.place(x=390, y=212)  ##
    ele4.insert(END, db_preset['B17'].value)
    tkinter.Label(custom_window, text="怪物属抗=", font=guide_font).place(x=320, y=240)
    ele5 = tkinter.Entry(custom_window, width=7);
    ele5.place(x=390, y=242)  ##
    ele5.insert(END, db_preset['B18'].value)
    tkinter.Label(custom_window, text="辅助减抗=", font=guide_font).place(x=320, y=270)
    ele6 = tkinter.Entry(custom_window, width=7);
    ele6.place(x=390, y=272)  ##
    ele6.insert(END, db_preset['B19'].value)

    load_preset.close()
    save_command = lambda: save_custom(ele_type.get(), cool_con.get(), cus1.get(), cus2.get(), cus3.get(), cus4.get(),
                                       cus6.get(), cus7.get(), cus8.get(), cus9.get(), cus10.get(), cus11.get(),
                                       cus12.get(),
                                       c_stat.get(), b_stat.get(), b_style_lvl.get(), c_style_lvl.get(), b_plt.get(),
                                       b_cri.get(),
                                       ele1.get(), ele2.get(), ele3.get(), ele4.get(), ele5.get(), ele6.get())
    tkinter.Button(custom_window, text="保存", font=mid_font, command=save_command, bg="lightyellow").place(x=190, y=295)


# re:如果调整了装备顺序，一定要记得一一去矫正着这个函数中的各个装备对应的行数
def save_custom(ele_type, cool_con, cus1, cus2, cus3, cus4, cus6, cus7, cus8, cus9, cus10, cus11, cus12, c_stat, b_stat,
                b_style_lvl, c_style_lvl, b_plt, b_cri, ele1, ele2, ele3, ele4, ele5, ele6):
    try:
        load_excel3 = load_workbook("DATA.xlsx")
        load_preset1 = load_workbook("preset.xlsx")
        db_custom1 = load_preset1["custom"]
        db_save_one = load_excel3["one"]
        db_save_set = load_excel3["set"]

        #########################################################
        #                     输出环境                           #
        #########################################################

        # 属性攻击
        db_custom1['B1'] = ele_type
        # 大自然防具会根据属性不同部位有不同的属强加成
        pos_elemental_strength_work_uniform_pants = 'L130'  # 工作服裤子
        pos_elemental_strength_work_uniform_shoulders = 'L150'  # 工作服头肩
        pos_elemental_strength_work_uniform_belt = 'L166'  # 工作服腰带
        pos_elemental_strength_work_uniform_shoes = 'L182'  # 工作服鞋子

        db_save_one[pos_elemental_strength_work_uniform_pants] = 0
        db_save_one[pos_elemental_strength_work_uniform_shoulders] = 0
        db_save_one[pos_elemental_strength_work_uniform_belt] = 0
        db_save_one[pos_elemental_strength_work_uniform_shoes] = 0

        if ele_type == '火':
            # 工作服头肩在火属性攻击时会增加火属性属强24点
            db_save_one[pos_elemental_strength_work_uniform_shoulders] = 24
        elif ele_type == '冰':
            # 工作服腰带在冰属性攻击时会增加冰属性属强24点
            db_save_one[pos_elemental_strength_work_uniform_belt] = 24
        elif ele_type == '光':
            # 工作服鞋子在光属性攻击时会增加光属性属强24点
            db_save_one[pos_elemental_strength_work_uniform_shoes] = 24
        elif ele_type == '暗':
            # 工作服裤子在暗属性攻击时会增加暗属性属强24点
            db_save_one[pos_elemental_strength_work_uniform_pants] = 24

        # 冷却补正比例
        db_custom1['B2'] = float(cool_con)
        # 行1：中文列名（我新加的）
        # 行2-257：装备列表
        # 行258-259：原作者加的英文缩写列名与序号
        # 行260-351：套装列表
        # 行352-353：原作者加的英文缩写列名与序号
        # 行354-361：智慧产物列表
        # 行362：套装编号与套装名称列
        # 行363-400：套装编号与套装名称
        for i in range(1, 400 + 1):
            try:
                db_save_one.cell(i, 25).value = db_save_one.cell(i, 26).value * float(cool_con) / 100
            except:
                pass

        for i in range(1, 93):
            try:
                db_save_set.cell(i, 25).value = db_save_one.cell(i, 26).value * float(cool_con) / 100
            except:
                pass

        #########################################################
        #                     特殊装备补正                       #
        #########################################################

        # 歧路腰带=X%
        db_custom1['B3'] = float(cus1)
        db_save_one['O165'] = float(cus1)
        # 歧路鞋子=X%
        db_custom1['B4'] = float(cus2)
        db_save_one['O181'] = float(cus2)
        # 不息上衣=X%
        db_custom1['B5'] = float(cus6)
        db_save_one['O101'] = float(cus6)
        db_save_one['O102'] = float(cus6)
        # 不息裤子=X%
        db_custom1['B6'] = float(cus7)
        db_save_one['O128'] = float(cus7)
        # 不息护肩=X%
        db_custom1['B7'] = float(cus8)
        db_save_one['O148'] = float(cus8)
        # 不息腰带=X%
        db_custom1['B8'] = float(cus9)
        db_save_one['O164'] = float(cus9)
        # 不息鞋子=X%
        db_custom1['B9'] = float(cus10)
        db_save_one['O180'] = float(cus10)
        # 不息2件套=X%
        db_custom1['B10'] = float(cus11)
        db_save_one['O296'] = float(cus11)
        # 不息3件套=X%
        db_custom1['B11'] = float(cus12)
        db_save_one['O297'] = float(cus12)
        db_save_one['O298'] = float(cus12)
        # 经验等级=英雄↑ 或 传说↓
        db_custom1['B12'] = cus3
        if cus3 == '传说↓':
            # 传说↓
            db_save_one['J87'] = 34  # 龙血玄黄-上衣
            db_save_one['F121'] = 34  # 龙血玄黄-裤子
            db_save_one['N141'] = 34  # 龙血玄黄-头肩
            db_save_one['L157'] = 68  # 龙血玄黄-腰带
            db_save_one['K173'] = 34  # 龙血玄黄-鞋子
            db_save_one['G277'] = 40  # 龙血玄黄5
        else:
            # 英雄↑
            db_save_one['J87'] = 35  # 龙血玄黄-上衣
            db_save_one['F121'] = 35  # 龙血玄黄-裤子
            db_save_one['N141'] = 35  # 龙血玄黄-头肩
            db_save_one['L157'] = 72  # 龙血玄黄-腰带
            db_save_one['K173'] = 35  # 龙血玄黄-鞋子
            db_save_one['G277'] = 41  # 龙血玄黄5
        # 恍惚增幅等级
        db_custom1['B13'] = cus4
        db_save_one['N190'] = int(cus4) + 4  # 破晓-手镯
        db_save_one['N191'] = int(cus4) + 4  # 破晓-神话手镯
        db_save_one['K206'] = int(cus4) + 4  # 破晓-项链
        db_save_one['E215'] = int(cus4) + 4  # 破晓-戒指

        #########################################################
        #                     奶量增幅相关                       #
        #########################################################

        # 面板体精智（左边的）
        db_custom1['H1'] = c_stat
        # 面板体精智（右边的）
        db_custom1['H6'] = b_stat
        # 祝福称号等级
        db_custom1['H2'] = b_style_lvl
        # 一觉称号等级
        db_custom1['H3'] = c_style_lvl
        # 祝福等级（左）
        db_custom1['H4'] = b_plt
        # 祝福等级（右）
        db_custom1['H5'] = b_cri

        #########################################################
        #                        属强相关                        #
        #########################################################

        # 基础属强
        db_custom1['B14'] = ele1
        # 其他属强
        db_custom1['B15'] = ele2
        # 勋章属强
        db_custom1['B16'] = ele3
        # 技能属强
        db_custom1['B17'] = ele4
        # 怪物属抗
        db_custom1['B18'] = ele5
        # 辅助减抗（如奶妈-60抗性）
        db_custom1['B19'] = ele6

        load_preset1.save("preset.xlsx")
        load_preset1.close()
        load_excel3.save("DATA.xlsx")
        load_excel3.close()
        custom_window.destroy()
        tkinter.messagebox.showinfo("通知", "保存完成")
    except PermissionError as error:
        tkinter.messagebox.showerror("错误", "请重试")


def load_checklist():
    ask_msg1 = tkinter.messagebox.askquestion('确认', "找回保存明细？")
    for snum in range(0, 10):
        if save_select.get() == save_name_list[snum]:
            ssnum1 = snum
    if ask_msg1 == 'yes':
        load_checklist_noconfirm(ssnum1)


def load_checklist_noconfirm(ssnum1):
    load_preset3 = load_workbook("preset.xlsx")
    db_load_check = load_preset3["one"]
    load_cell = db_load_check.cell
    # 读取各个装备的点亮情况
    k = 1
    for i in range(1, 264):
        if load_cell(i, 2 + ssnum1).value == 1:
            try:
                select_item['tg{}'.format(load_cell(i, 1).value)] = 1
            except KeyError as error:
                passss = 1
        elif load_cell(i, 2 + ssnum1).value == 0:
            try:
                select_item['tg{}'.format(load_cell(i, 1).value)] = 0
            except KeyError as error:
                passss = 1

    # 增加读取武器、职业等选项
    col_custom_save_value = g_col_custom_save_value_begin + ssnum1
    wep_combopicker.set((load_cell(g_row_custom_save_weapon, col_custom_save_value).value or "").split(','))
    jobup_select.set(load_cell(g_row_custom_save_job, col_custom_save_value).value)
    time_select.set(load_cell(g_row_custom_save_fight_time, col_custom_save_value).value)
    style = load_cell(g_row_custom_save_title, col_custom_save_value).value
    # 由于调整了国服特色的实现，若找不到之前版本存档的称号，则换为第一个称号
    if style not in styles:
        style = styles[0]
    style_select.set(style)
    creature = load_cell(g_row_custom_save_pet, col_custom_save_value).value
    # 由于调整了国服特色的实现，若找不到之前版本存档的称号，则换为第一个称号
    if creature not in creatures:
        creature = creatures[0]
    creature_select.set(creature)
    req_cool.set(load_cell(g_row_custom_save_cd, col_custom_save_value).value)
    select_speed.set(load_cell(g_row_custom_save_speed, col_custom_save_value).value)
    baibianguai_select.set(
        load_cell(g_row_custom_save_has_baibianguai, col_custom_save_value).value or txt_no_baibianguai)
    can_upgrade_work_unifrom_nums_select.set(
        load_cell(g_row_custom_save_can_upgrade_work_uniforms_nums, col_custom_save_value).value or
        txt_can_upgrade_work_unifrom_nums[0])
    transfer_equip_combopicker.set((load_cell(g_row_custom_save_transfer_from, col_custom_save_value).value or "").split(','))
    can_transfer_nums_select.set(load_cell(g_row_custom_save_max_transfer_count, col_custom_save_value).value or txt_can_transfer_nums[0])

    load_preset3.close()
    check_equipment()
    for i in range(101, 136):
        check_set(i)


# save_idx为存档的下标，从0到9
def save_my_custom(sc, row, col_custom_save_value, name, value):
    sc(row, g_col_custom_save_key).value = name
    sc(row, col_custom_save_value).value = value


def save_checklist():
    ask_msg2 = tkinter.messagebox.askquestion('确认', "确认保存吗？")
    for snum in range(0, 10):
        if save_select.get() == save_name_list[snum]:
            ssnum2 = snum
    try:
        if ask_msg2 == 'yes':
            load_preset4 = load_workbook("preset.xlsx")
            db_save_check = load_preset4["one"]
            save_cell = db_save_check.cell

            # 保存装备按钮的点亮情况
            opt_save = {}  # 装备按钮的index => 对应的行号（1-263）
            for i in range(1, 264):
                opt_save[save_cell(i, 1).value] = i

            for code in opt_save.keys():
                try:
                    if eval("select_item['tg{}']".format(code)) == 1:
                        save_cell(opt_save[code], 2 + ssnum2).value = 1
                except KeyError as error:
                    passss1 = 1

                try:
                    if eval("select_item['tg{}']".format(code)) == 0:
                        save_cell(opt_save[code], 2 + ssnum2).value = 0
                except KeyError as error:
                    passss1 = 1

                passss = 1

            # 增加保存武器、职业等选项
            col_custom_save_value = g_col_custom_save_value_begin + ssnum2
            save_my_custom(save_cell, g_row_custom_save_save_name, col_custom_save_value, "存档名", save_select.get())
            save_my_custom(save_cell, g_row_custom_save_weapon, col_custom_save_value, "武器", wep_combopicker.current_value)
            save_my_custom(save_cell, g_row_custom_save_job, col_custom_save_value, "职业选择", jobup_select.get())
            save_my_custom(save_cell, g_row_custom_save_fight_time, col_custom_save_value, "输出时间", time_select.get())
            save_my_custom(save_cell, g_row_custom_save_title, col_custom_save_value, "称号选择", style_select.get())
            save_my_custom(save_cell, g_row_custom_save_pet, col_custom_save_value, "宠物选择", creature_select.get())
            save_my_custom(save_cell, g_row_custom_save_cd, col_custom_save_value, "冷却补正", req_cool.get())
            save_my_custom(save_cell, g_row_custom_save_speed, col_custom_save_value, "选择速度", select_speed.get())
            save_my_custom(save_cell, g_row_custom_save_has_baibianguai, col_custom_save_value, "是否拥有百变怪", baibianguai_select.get())
            save_my_custom(save_cell, g_row_custom_save_can_upgrade_work_uniforms_nums, col_custom_save_value, "材料够升级的工作服数目", can_upgrade_work_unifrom_nums_select.get())
            save_my_custom(save_cell, g_row_custom_save_transfer_from, col_custom_save_value, "跨界来源账户列表", transfer_equip_combopicker.current_value)
            save_my_custom(save_cell, g_row_custom_save_max_transfer_count, col_custom_save_value, "最大跨界数目", can_transfer_nums_select.get())

            load_preset4.save("preset.xlsx")
            load_preset4.close()
            tkinter.messagebox.showinfo("通知", "保存完成")
    except PermissionError as error:
        tkinter.messagebox.showerror("错误", "请关闭preset.xlsx之后重试")


def change_list_name():
    global change_window
    change_window = tkinter.Toplevel(self)
    change_window.geometry("190x320+750+200")
    tkinter.Label(change_window, text="1套").place(x=20, y=10)
    tkinter.Label(change_window, text="2套").place(x=20, y=35)
    tkinter.Label(change_window, text="3套").place(x=20, y=60)
    tkinter.Label(change_window, text="4套").place(x=20, y=85)
    tkinter.Label(change_window, text="5套").place(x=20, y=110)
    tkinter.Label(change_window, text="6套").place(x=20, y=135)
    tkinter.Label(change_window, text="7套").place(x=20, y=160)
    tkinter.Label(change_window, text="8套").place(x=20, y=185)
    tkinter.Label(change_window, text="9套").place(x=20, y=210)
    tkinter.Label(change_window, text="10套").place(x=20, y=235)
    entry1 = tkinter.Entry(change_window, width=10);
    entry1.place(x=95, y=12);
    entry1.insert(END, save_name_list[0])
    entry2 = tkinter.Entry(change_window, width=10);
    entry2.place(x=95, y=37);
    entry2.insert(END, save_name_list[1])
    entry3 = tkinter.Entry(change_window, width=10);
    entry3.place(x=95, y=62);
    entry3.insert(END, save_name_list[2])
    entry4 = tkinter.Entry(change_window, width=10);
    entry4.place(x=95, y=87);
    entry4.insert(END, save_name_list[3])
    entry5 = tkinter.Entry(change_window, width=10);
    entry5.place(x=95, y=112);
    entry5.insert(END, save_name_list[4])
    entry6 = tkinter.Entry(change_window, width=10);
    entry6.place(x=95, y=137);
    entry6.insert(END, save_name_list[5])
    entry7 = tkinter.Entry(change_window, width=10);
    entry7.place(x=95, y=162);
    entry7.insert(END, save_name_list[6])
    entry8 = tkinter.Entry(change_window, width=10);
    entry8.place(x=95, y=187);
    entry8.insert(END, save_name_list[7])
    entry9 = tkinter.Entry(change_window, width=10);
    entry9.place(x=95, y=212);
    entry9.insert(END, save_name_list[8])
    entry10 = tkinter.Entry(change_window, width=10);
    entry10.place(x=95, y=237);
    entry10.insert(END, save_name_list[9])

    tkinter.Button(change_window, text="保存", font=mid_font,
                   command=lambda: change_savelist(entry1.get(), entry2.get(), entry3.get(), entry4.get(), entry5.get(),
                                                   entry6.get(), entry7.get(), entry8.get(), entry9.get(),
                                                   entry10.get())).place(x=60, y=270)


def change_savelist(in1, in2, in3, in4, in5, in6, in7, in8, in9, in10):
    in_list = [in1, in2, in3, in4, in5, in6, in7, in8, in9, in10]
    try:
        load_preset5 = load_workbook("preset.xlsx", data_only=True)
        db_custom2 = load_preset5["custom"]

        for i in range(1, 11):
            db_custom2.cell(i, 5).value = in_list[i - 1]
        global save_name_list

        # 更改存档名后，仍保持当前位置这个存档
        current_index = 0
        try:
            current_index = save_name_list.index(save_select.get())
        except ValueError:
            current_index = 0

        save_name_list = in_list
        load_preset5.save("preset.xlsx")
        load_preset5.close()
        save_select.set(save_name_list[current_index])
        save_select['values'] = save_name_list
        change_window.destroy()
        tkinter.messagebox.showinfo("通知", "保存完成")

        # 更新存档名称后，同步更新picker的列表，同时清空已选列表
        transfer_equip_combopicker.set("")
    except PermissionError as error:
        tkinter.messagebox.showerror("错误", "请关闭preset.xlsx之后重试")


def update_count():
    global count_valid, count_invalid, show_number
    global showcon, all_list_num, count_start_time
    global exit_calc
    hours = 0
    minutes = 0
    seconds = 0
    using_time_str = "0s"
    remaining_time_str = "0s"
    while True:
        try:
            show_str = str(count_valid) + "有效搭配/" + str(count_invalid) + "无效"
            if exit_calc == 0:
                using_time = time.time() - count_start_time
                using_time_str = format_time(using_time)
                processed = count_valid + count_invalid
                if all_list_num >= processed:
                    remaining_time = (all_list_num - count_valid - count_invalid) / (
                            count_valid + count_invalid + 1) * using_time
                    remaining_time_str = format_time(remaining_time)
                else:
                    remaining_time_str = "0s(未经确计数)"
            showcon(text=(str(count_valid) + "有效搭配/" + str(count_invalid) + "无效" +
                          "\n用时=" + using_time_str + "" +
                          "\n剩余=" + remaining_time_str)
                    )
            time.sleep(0.1)
        except Exception as e:
            print("update_count except: {}".format(e))


def display_realtime_counting_info():
    while True:
        try:
            items, not_select_items, work_uniforms_items = get_equips()

            # 已选装备的搭配数
            all_list_num = calc_ori_counts(items)
            # 百变怪增加的搭配数
            all_list_num += calc_bbg_add_counts(items, not_select_items)
            # 额外升级的工作服增加的搭配数
            all_list_num += calc_upgrade_work_uniforms_add_counts(items, not_select_items, work_uniforms_items)

            current_equips = 0
            for slot_equips in items:
                current_equips += len(slot_equips)
            total_equips = 177
            percent = current_equips / total_equips * 100

            show_txt = "{}/{}({:.2f}%) N={}".format(current_equips, total_equips, percent, int(all_list_num))
            showcon2(text=show_txt)
            time.sleep(1)
        except Exception as e:
            print("display_realtime_counting_info except: {}".format(e))


def load_checklist_on_start():
    # 启动时自动读取第一个配置
    load_checklist_noconfirm(0)


def update_thread():
    threading.Thread(target=update_count, daemon=True).start()
    threading.Thread(target=display_realtime_counting_info, daemon=True).start()
    threading.Thread(target=load_checklist_on_start, daemon=True).start()


def reset():
    know_list2 = ['13390150', '22390240', '23390450', '33390750', '21400340', '31400540', '32410650']
    for i in range(1101, 3336):
        try:
            select_item['tg{}0'.format(i)] = 0
        except KeyError as error:
            passss = 1
        try:
            select_item['tg{}1'.format(i)] = 0
        except KeyError as error:
            passss = 1
    for i in know_list2:
        select_item['tg{}'.format(i)] = 0
    check_equipment()
    for i in range(101, 136):
        check_set(i)

    # 处理百变怪与工作服升级数目
    baibianguai_select.set(txt_no_baibianguai)
    can_upgrade_work_unifrom_nums_select.set(txt_can_upgrade_work_unifrom_nums[0])

    wep_combopicker.set("")
    transfer_equip_combopicker.set("")
    can_transfer_nums_select.set(txt_can_transfer_nums[0])


###########################################################
#                         逻辑初始化                       #
###########################################################

exit_calc = 1
save_name_list = []
save_select = 0
count_valid = 0
unique_index = 0
count_invalid = 0
show_number = 0
all_list_num = 0
g_current_rank = 0
g_current_job = ""
g_current_buff_type = "祝福"  # 祝福 一觉 综合
g_rank_equips = {}
count_start_time = time.time()  # 开始计算的时间点

# 由于这里不需要对data.xlsx写入，设置read_only为True可以大幅度加快读取速度，在我的电脑上改动前读取耗时0.67s，占启动时间32%，改动之后用时0.1s，占启动时间4%
load_excel1 = load_workbook("DATA.xlsx", read_only=True, data_only=True)
db_one = load_excel1["one"]
name_one = {}
equip_index_to_realname = {}
for row in db_one.rows:
    row_value = [cell.value for cell in row]

    index = row_value[0]
    realname = row_value[1]

    name_one[index] = row_value
    equip_index_to_realname[index] = realname

db_job = load_excel1["lvl"]
opt_job = {}
opt_job_ele = {}
jobs = []

for row in db_job.rows:
    row_value = [cell.value for cell in row]

    job = row_value[0]
    if job in ["empty", "직업명"]:
        continue

    opt_job[job] = row_value[3:]
    opt_job_ele[job] = row_value[:3]
    jobs.append(job)

load_excel1.close()

# 该变量用来控制是否要检查preset.xlsx正确初始化，若有些必要的cell没有正确运行，则会赋值，为了更快启动，魔改版本不再检查
need_check_preset_file = False
load_preset0 = load_workbook("preset.xlsx", read_only=not need_check_preset_file, data_only=True)
db_custom = load_preset0["custom"]
save_name_list = []
for i in range(1, 11):
    save_name_list.append(db_custom.cell(i, 5).value)

if need_check_preset_file:
    ########## 버전 최초 구동 프리셋 업데이트 ###########
    try:
        db_save = load_preset0["one"]
        print("DATABASE 버전= " + str(db_custom['K1'].value))
        print("클라이언트 버전= " + now_version)
        if str(db_custom['K1'].value) != now_version:
            # print("DB 업데이트")
            db_custom['K1'] = now_version
        if db_custom['H1'].value == None:
            db_custom['G1'] = "up_stat"
            db_custom['H1'] = 0
        if db_custom['H2'].value == None:
            db_custom['G2'] = "bless_style"
            db_custom['H2'] = 3
        if db_custom['H3'].value == None:
            db_custom['G3'] = "crux_style"
            db_custom['H3'] = 2
        if db_custom['H4'].value == None:
            db_custom['G4'] = "bless_plt"
            db_custom['H4'] = 2
        if db_custom['H5'].value == None:
            db_custom['G5'] = "bless_cri"
            db_custom['H5'] = 1
        if db_custom['H6'].value == None:
            db_custom['G6'] = "up_stat_b"
            db_custom['H6'] = 0

        if db_custom['B14'].value == None:
            db_custom['A14'] = "ele_inchant"
            db_custom['B14'] = 116
        if db_custom['B15'].value == None:
            db_custom['A15'] = "ele_ora"
            db_custom['B15'] = 20
        if db_custom['B16'].value == None:
            db_custom['A16'] = "ele_gem"
            db_custom['B16'] = 7
        if db_custom['B17'].value == None:
            db_custom['A17'] = "ele_skill"
        db_custom['B17'] = 0  ## 자속강 비활성화
        if db_custom['B18'].value == None:
            db_custom['A18'] = "ele_mob_resist"
            db_custom['B18'] = 50
        if db_custom['B19'].value == None:
            db_custom['A19'] = "ele_buf_anti"
            db_custom['B19'] = 60

        if db_save['A257'].value == None:
            db_save['A257'] = '13390150';
            db_save['B257'] = '+5 퍼펙트컨트롤'
            db_save['C257'] = 0;
            db_save['D257'] = 0;
            db_save['E257'] = 0;
            db_save['F257'] = 0;
            db_save['G257'] = 0
            db_save['H257'] = 0;
            db_save['I257'] = 0;
            db_save['J257'] = 0;
            db_save['K257'] = 0;
            db_save['L257'] = 0
        if db_save['A258'].value == None:
            db_save['A258'] = '22390240';
            db_save['B258'] = '+4 선지자의 목걸이'
            db_save['C258'] = 0;
            db_save['D258'] = 0;
            db_save['E258'] = 0;
            db_save['F258'] = 0;
            db_save['G258'] = 0
            db_save['H258'] = 0;
            db_save['I258'] = 0;
            db_save['J258'] = 0;
            db_save['K258'] = 0;
            db_save['L258'] = 0
        if db_save['A259'].value == None:
            db_save['A259'] = '21400340';
            db_save['B259'] = '+4 독을 머금은 가시장갑'
            db_save['C259'] = 0;
            db_save['D259'] = 0;
            db_save['E259'] = 0;
            db_save['F259'] = 0;
            db_save['G259'] = 0
            db_save['H259'] = 0;
            db_save['I259'] = 0;
            db_save['J259'] = 0;
            db_save['K259'] = 0;
            db_save['L259'] = 0
        if db_save['A260'].value == None:
            db_save['A260'] = '23390450';
            db_save['B260'] = '+5 할기의 링'
            db_save['C260'] = 0;
            db_save['D260'] = 0;
            db_save['E260'] = 0;
            db_save['F260'] = 0;
            db_save['G260'] = 0
            db_save['H260'] = 0;
            db_save['I260'] = 0;
            db_save['J260'] = 0;
            db_save['K260'] = 0;
            db_save['L260'] = 0
        if db_save['A261'].value == None:
            db_save['A261'] = '31400540';
            db_save['B261'] = '+4 청면수라의 가면'
            db_save['C261'] = 0;
            db_save['D261'] = 0;
            db_save['E261'] = 0;
            db_save['F261'] = 0;
            db_save['G261'] = 0
            db_save['H261'] = 0;
            db_save['I261'] = 0;
            db_save['J261'] = 0;
            db_save['K261'] = 0;
            db_save['L261'] = 0
        if db_save['A262'].value == None:
            db_save['A262'] = '32410650';
            db_save['B262'] = '+5 적귀의 차원석'
            db_save['C262'] = 0;
            db_save['D262'] = 0;
            db_save['E262'] = 0;
            db_save['F262'] = 0;
            db_save['G262'] = 0
            db_save['H262'] = 0;
            db_save['I262'] = 0;
            db_save['J262'] = 0;
            db_save['K262'] = 0;
            db_save['L262'] = 0
        if db_save['A263'].value == None:
            db_save['A263'] = '33390750';
            db_save['B263'] = '+5 패스트퓨처 이어링'
            db_save['C263'] = 0;
            db_save['D263'] = 0;
            db_save['E263'] = 0;
            db_save['F263'] = 0;
            db_save['G263'] = 0
            db_save['H263'] = 0;
            db_save['I263'] = 0;
            db_save['J263'] = 0;
            db_save['K263'] = 0;
            db_save['L263'] = 0

        load_preset0.save("preset.xlsx")

    except PermissionError as error:
        tkinter.messagebox.showerror("错误", "更新失败. 请重新运行.")

load_preset0.close()


###########################################################
#                韩服特有的一些功能，与我们无关              #
###########################################################

def timeline_select():
    global timeline_window
    timeline_window = tkinter.Toplevel(self)
    timeline_window.attributes("-topmost", True)
    timeline_window.geometry("310x150+750+20")
    tkinter.Label(timeline_window, text="角色名=\n(准确的)", font=guide_font).place(x=10, y=9)
    cha_name = tkinter.Entry(timeline_window, width=13)
    cha_name.place(x=80, y=12)
    tkinter.Label(timeline_window, text="서버명=", font=guide_font).place(x=10, y=59)
    sever_list = ['카인', '디레지에', '바칼', '힐더', '안톤', '카시야스', '프레이', '시로코']
    serv_name = tkinter.ttk.Combobox(timeline_window, values=sever_list, width=11)
    serv_name.place(x=80, y=62)
    serv_name.set('카인')
    load_timeline = tkinter.Button(timeline_window, command=lambda: show_timeline(cha_name.get(), serv_name.get()),
                                   text="불러오기", font=mid_font)
    load_timeline.place(x=200, y=25)
    tkinter.Label(timeline_window, text="타임라인에 있는 에픽만 불러옵니다(일부X)", fg="Red").place(x=10, y=100)
    tkinter.Label(timeline_window, text="如果服务器不稳定请多试几次", fg="Red").place(x=10, y=120)


def show_timeline(name, server):
    # 国服没有这个东西，干掉它，提升运行效率
    pass


def cha_select(jobs):
    # 国服没有这个东西，干掉它，提升运行效率
    pass


def calc_my_cha(cha_name, serv_name, job_name):
    # 国服没有这个东西，干掉它，提升运行效率
    pass


def cha_image(cha_code, serv_code):
    # 国服没有这个东西，干掉它，提升运行效率
    pass


def show_my_cha(equipment, final_list, type_code, job_name, wep_name, ele_skill, cha_name, info_stat, *equipment2):
    # 国服没有这个东西，干掉它，提升运行效率
    pass


def toggle_cha_swi():
    # 国服没有这个东西，干掉它，提升运行效率
    pass


###########################################################
#                        ui相关变量                        #
###########################################################

select_item = {}

dark_main = _from_rgb((32, 34, 37))
dark_sub = _from_rgb((46, 49, 52))
dark_blue = _from_rgb((29, 30, 36))

# 目前可升级的工作服数目
txt_can_upgrade_work_unifrom_nums = [
    '材料够升级零件', '材料够升级一件', '材料够升级两件', '材料够升级三件', '材料够升级四件', '材料够升级五件',
    '材料够升级六件', '材料够升级七件', '材料够升级八件', '材料够升级九件', '材料够升级十件', '材料够升级十一件',
]
# 预先将升级工作服数目的字符串与对应数目映射
can_upgrade_work_unifrom_nums_str_2_int = {}
for idx, txt in enumerate(txt_can_upgrade_work_unifrom_nums):
    can_upgrade_work_unifrom_nums_str_2_int[txt] = idx

# 目前最多可跨界的装备数目
txt_can_transfer_nums = [
    '0', '1', '2', '3', '4', '5',
    '6', '7', '8', '9', '10', '11',
]
# 预先将目前最多可跨界的装备数目与对应数目映射
can_transfer_nums_str_2_int = {}
for idx, txt in enumerate(txt_can_transfer_nums):
    can_transfer_nums_str_2_int[txt] = idx


###########################################################
#                        ui相关函数                        #
###########################################################

def guide_speed():
    tkinter.messagebox.showinfo("准确度选择", (
        "快速=不太精确-删除单一散件\n"
        "中速=稍精确-包括散件, 神话优先\n"
        "慢速=比较精确-所有限制解除(非常慢)(保留价值预估函数过滤)\n"
        "超慢速=非常精确-所有限制解除(天荒地老海枯石烂的慢)"))


def click_equipment(code):
    if eval("select_item['tg{}']".format(code)) == 0:
        eval('select_{}'.format(code))['image'] = image_list[str(code)]
        select_item['tg' + str('{}'.format(code))] = 1
    elif eval("select_item['tg{}']".format(code)) == 1:
        eval('select_{}'.format(code))['image'] = image_list2[str(code)]
        select_item['tg' + str('{}'.format(code))] = 0
    if len(str(code)) == 5:
        check_set(int('1' + str(code)[2:4]))


def check_equipment():
    know_list2 = ['13390150', '22390240', '23390450', '33390750', '21400340', '31400540', '32410650']
    for i in range(11010, 33352):
        try:
            if eval("select_item['tg{}']".format(i)) == 0:
                eval('select_{}'.format(i))['image'] = image_list2[str(i)]
            elif eval("select_item['tg{}']".format(i)) == 1:
                eval('select_{}'.format(i))['image'] = image_list[str(i)]
        except KeyError as error:
            c = 1
    for i in know_list2:
        try:
            if eval("select_item['tg{}']".format(i)) == 0:
                eval('select_{}'.format(i))['image'] = image_list2[str(i)]
            elif eval("select_item['tg{}']".format(i)) == 1:
                eval('select_{}'.format(i))['image'] = image_list[str(i)]
        except KeyError as error:
            c = 1


def click_set(code):
    code_add = code - 100
    code_str = str(code)[1:3]
    set_checked = 0
    if code >= 116:  ##악세/특장/스까면
        if 116 <= code <= 119:
            for i in range(21, 24):  ## 악세부위에서
                try:
                    if select_item['tg' + str(i) + code_str + '0'] == 1:  ##채택된 숫자를 찾는다
                        set_checked = set_checked + 1  ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c = 1
        elif 123 >= code >= 120:
            for i in range(31, 34):  ## 특장부위에서
                try:
                    if select_item['tg' + str(i) + code_str + '0'] == 1:  ##채택된 숫자를 찾는다
                        set_checked = set_checked + 1  ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c = 1
        elif 131 >= code >= 128:
            for i in [11, 22, 31]:  ## 상목보부위에서
                try:
                    if select_item['tg' + str(i) + code_str + '0'] == 1:  ##채택된 숫자를 찾는다
                        set_checked = set_checked + 1  ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c = 1
        elif 127 >= code >= 124:
            for i in [12, 21, 32]:  ## 하팔법부위에서
                try:
                    if select_item['tg' + str(i) + code_str + '0'] == 1:  ##채택된 숫자를 찾는다
                        set_checked = set_checked + 1  ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c = 1
        elif 135 >= code >= 132:
            for i in [15, 23, 33]:  ## 신반귀부위에서
                try:
                    if select_item['tg' + str(i) + code_str + '0'] == 1:  ##채택된 숫자를 찾는다
                        set_checked = set_checked + 1  ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c = 1
        if set_checked == 3:  ## 채택 숫자가 3이면
            for i in range(11, 36):  ##모든 부위에서
                try:
                    eval('select_' + str(i) + code_str + '0')['image'] = image_list2[
                        str(i) + code_str + '0']  ##이미지도 오프로 바꿈
                    select_item['tg' + str(i) + code_str + '0'] = 0  ##모든 체크를 0으로 만듬
                except KeyError as error:
                    c = 1
            eval('set' + str(code))['image'] = image_list_set2[str(code)]  ##세트이미지도 오프로 바꿈
        else:  ## 채택 숫자가 3미만이면
            for i in range(11, 36):  ##모든 부위에서
                try:
                    eval('select_' + str(i) + code_str + '0')['image'] = image_list[
                        str(i) + code_str + '0']  ##이미지도 온으로 바꿈
                    select_item['tg' + str(i) + code_str + '0'] = 1  ##모든 체크를 1으로 만듬
                except KeyError as error:
                    c = 1
            eval('set' + str(code))['image'] = image_list_set[str(code)]  ##세트이미지도 온으로 바꿈


    else:
        for i in range(11, 16):  ## 방어구 부위에서
            try:
                if select_item['tg' + str(i) + code_str + '0'] == 1:  ##채택된 숫자를 찾는다
                    set_checked = set_checked + 1  ##그럼 변수에 +1을 더함
            except KeyError as error:
                c = 1

        if set_checked == 5:  ## 채택 숫자가 5이면
            for i in range(11, 16):  ## 방어구 부위에서
                try:
                    eval('select_' + str(i) + code_str + '0')['image'] = image_list2[
                        str(i) + code_str + '0']  ##이미지도 오프로 바꿈
                    select_item['tg' + str(i) + code_str + '0'] = 0  ##모든 체크를 0으로 만듬
                except KeyError as error:
                    c = 1
            eval('set' + str(code))['image'] = image_list_set2[str(code)]  ##세트이미지도 오프로 바꿈

        else:  ## 채택 숫자가 5미만이면
            for i in range(11, 16):  ## 방어구 부위에서
                try:
                    eval('select_' + str(i) + code_str + '0')['image'] = image_list[
                        str(i) + code_str + '0']  ##이미지도 온으로 바꿈
                    select_item['tg' + str(i) + code_str + '0'] = 1  ##모든 체크를 1으로 만듬
                except KeyError as error:
                    c = 1
            eval('set' + str(code))['image'] = image_list_set[str(code)]  ##세트이미지도 온으로 바꿈


def check_set(code):
    code_str = str(code)[1:3]
    set_checked = 0
    if code < 116:
        for i in [11, 12, 13, 14, 15]:
            if select_item['tg' + str(i) + code_str + '0'] == 1:
                set_checked = set_checked + 1
    elif code < 120:
        for i in [21, 22, 23]:
            if select_item['tg' + str(i) + code_str + '0'] == 1:
                set_checked = set_checked + 1
    elif code < 124:
        for i in [31, 32, 33]:
            if select_item['tg' + str(i) + code_str + '0'] == 1:
                set_checked = set_checked + 1
    elif code < 128:
        for i in [12, 21, 32]:
            if select_item['tg' + str(i) + code_str + '0'] == 1:
                set_checked = set_checked + 1
    elif code < 132:
        for i in [11, 22, 31]:
            if select_item['tg' + str(i) + code_str + '0'] == 1:
                set_checked = set_checked + 1
    elif code < 136:
        for i in [15, 23, 33]:
            if select_item['tg' + str(i) + code_str + '0'] == 1:
                set_checked = set_checked + 1

    if code < 116:
        if set_checked == 5:
            eval('set' + str(code))['image'] = image_list_set[str(code)]
        else:
            eval('set' + str(code))['image'] = image_list_set2[str(code)]
    else:
        if set_checked == 3:
            eval('set' + str(code))['image'] = image_list_set[str(code)]
        else:
            eval('set' + str(code))['image'] = image_list_set2[str(code)]


def donate():
    webbrowser.open('https://twip.kr/dawnclass16')


def dunfaoff():
    webbrowser.open('https://space.bilibili.com/4952736')


def blog():
    webbrowser.open('https://blog.naver.com/dawnclass16/221837654941')


def hamjung():
    tkinter.messagebox.showinfo("제작자 크레딧",
                                "총제작자=Dawnclass(새벽반)\n이미지/그래픽=경철부동산\n직업/버퍼DB=대략볼록할철\n서버제공=던파오프\n기타조언=히든 도비 4,5,6호\n\n오류 제보는 블로그 덧글이나 던조 쪽지로")


def get_other_account_names():
    current_name = save_select.get()
    return [name for name in save_name_list if name != current_name]


###########################################################
#                        tkinter初始化                    #
###########################################################

self = tkinter.Tk()
self.title("一键史诗搭配计算器-支持百变怪/升级工作服/跨界/多武器 ver" + now_version)
self.geometry("710x720+0+0")
self.resizable(False, False)
self.configure(bg=dark_main)
self.iconbitmap(r'ext_img/icon.ico')

###########################################################
#                      拼接ui的琐碎代码                    #
###########################################################


guide_font = tkinter.font.Font(family="맑은 고딕", size=10, weight='bold')
mid_font = tkinter.font.Font(family="맑은 고딕", size=14, weight='bold')
big_font = tkinter.font.Font(family="맑은 고딕", size=18, weight='bold')

## 내부 구조 ##
know_list = ['13390150', '22390240', '23390450', '33390750', '21400340', '31400540', '32410650']
image_list = {}
image_list2 = {}
image_list_set = {}
image_list_set2 = {}

# 读取装备图片
# 通过遍历文件夹来实现加载所需的图片，而不是穷举所有可能，最后导致启动时要卡顿两秒，根据测试，目前读取图片共使用0:00:01.780298秒, 总共尝试加载6749个， 有效的加载为351个
image_directory = "image"
for filename in os.listdir(image_directory):
    # 示例文件：22390240f.png
    index = filename[:-5]  # 装备的key(除去后五位后剩余的字符串)：22390240
    newImage = PhotoImage(file="image/{}".format(filename))  #
    if filename[-5] == "n":  # 根据倒数第五位决定使用哪个list
        # 神话装备会有三个文件，以11011为例，分别为11011f.png/11011n.gif/11011n.png，其中后面两个为点亮时的样式，
        # 为了跟原版一致，当是神话装备时，加载点亮样式时，优先使用gif版本的
        if  is_god(index) and index in image_list and filename.endswith(".png"):
            continue
        image_list[index] = newImage
    else:
        image_list2[index] = newImage

# 读取套装图片
for i in range(1, 36):
    image_list_set[str(100 + i)] = eval('PhotoImage(file="set_name/{}.png")'.format(i + 100))
    image_list_set2[str(100 + i)] = eval('PhotoImage(file="set_name/{}f.png")'.format(i + 100))

bg_img = PhotoImage(file="ext_img/bg_img.png")
bg_wall = tkinter.Label(self, image=bg_img)
bg_wall.place(x=0, y=0)

select_speed = tkinter.ttk.Combobox(self, values=['快速', '中速', '慢速', '超慢速'], width=15)
select_speed.place(x=145, y=11)
select_speed.set('中速')
select_speed_img = PhotoImage(file="ext_img/select_speed.png")
tkinter.Button(self, command=guide_speed, image=select_speed_img, borderwidth=0, activebackground=dark_main,
               bg=dark_main).place(x=29, y=7)
reset_img = PhotoImage(file="ext_img/reset.png")
tkinter.Button(self, command=reset, image=reset_img, borderwidth=0, activebackground=dark_main, bg=dark_main).place(
    x=302, y=476)

wep_list = []
wep_name_to_index = {}
for i in range(0, 75):
    wep_index = name_one[str(i + 111001)][0]
    wep_name = name_one[str(i + 111001)][1]

    wep_list.append(wep_name)
    wep_name_to_index[wep_name] = wep_index

wep_image = PhotoImage(file="ext_img/wep.png")
wep_g = tkinter.Label(self, image=wep_image, borderwidth=0, activebackground=dark_main, bg=dark_main)
wep_g.place(x=29, y=55)
wep_combopicker = Combopicker(self, values=wep_list, entrywidth=30)
wep_combopicker.place(x=110, y=60)

time_select = tkinter.ttk.Combobox(self, width=13, values=['20秒(觉醒占比↑)', '60秒(觉醒占比↓)'])
time_select.set('20秒(觉醒占比↑)')
time_select.place(x=390 - 17, y=220 + 52)
jobup_select = tkinter.ttk.Combobox(self, width=13, values=jobs)
jobup_select.set('职业选择')
jobup_select.place(x=390 - 17, y=190 + 52)
style_list = styles
style_select = tkinter.ttk.Combobox(self, width=13, values=style_list)
style_select.set(styles[0])
style_select.place(x=390 - 17, y=250 + 52)
creature_list = creatures
creature_select = tkinter.ttk.Combobox(self, width=13, values=creature_list)
creature_select.set(creatures[0])
creature_select.place(x=390 - 17, y=280 + 52)
req_cool = tkinter.ttk.Combobox(self, width=13, values=['X(纯伤害)', 'O(打开)'])
req_cool.set('X(纯伤害)')
req_cool.place(x=390 - 17, y=310 + 52)

calc_img = PhotoImage(file="ext_img/calc.png")
select_all = tkinter.Button(self, image=calc_img, borderwidth=0, activebackground=dark_main, command=calc_thread,
                            bg=dark_main)
select_all.place(x=390 - 35, y=7)
stop_img = PhotoImage(file="ext_img/stop.png")
tkinter.Button(self, image=stop_img, borderwidth=0, activebackground=dark_main, command=stop_calc, bg=dark_main).place(
    x=390 - 35, y=62)

timeline_img = PhotoImage(file="ext_img/timeline.png")
select_custom = tkinter.Button(self, image=timeline_img, borderwidth=0, activebackground=dark_main,
                               command=timeline_select, bg=dark_sub)
select_custom.place(x=345 + 165, y=340 - 100)
custom_img = PhotoImage(file="ext_img/custom.png")
select_custom2 = tkinter.Button(self, image=custom_img, borderwidth=0, activebackground=dark_main, command=costum,
                                bg=dark_sub)
select_custom2.place(x=435 + 165, y=340 - 100)

save_select = tkinter.ttk.Combobox(self, width=8, values=save_name_list)
save_select.place(x=345 + 165, y=410 - 100);
save_select.set(save_name_list[0])
save_img = PhotoImage(file="ext_img/SAVE.png")
save = tkinter.Button(self, image=save_img, borderwidth=0, activebackground=dark_main, command=save_checklist,
                      bg=dark_sub)
save.place(x=345 + 165, y=440 - 100)
load_img = PhotoImage(file="ext_img/LOAD.png")
load = tkinter.Button(self, image=load_img, borderwidth=0, activebackground=dark_main, command=load_checklist,
                      bg=dark_sub)
load.place(x=435 + 165, y=440 - 100)
change_name_img = PhotoImage(file="ext_img/name_change.png")
change_list_but = tkinter.Button(self, image=change_name_img, borderwidth=0, activebackground=dark_main,
                                 command=change_list_name, bg=dark_sub)
change_list_but.place(x=435 + 165, y=405 - 100)

# 百变怪选项
txt_no_baibianguai = 'No(没有百变怪)'
txt_has_baibianguai = 'Yes(拥有百变怪)'
baibianguai_txt = tkinter.Label(self, text="  百变怪  ", font=guide_font, fg="white", bg=dark_sub)
baibianguai_txt.place(x=300, y=395)
baibianguai_select = tkinter.ttk.Combobox(self, width=13, values=[txt_no_baibianguai, txt_has_baibianguai])
baibianguai_select.set(txt_no_baibianguai)
baibianguai_select.place(x=390 - 17, y=395)

can_upgrade_work_unifrom_nums_txt = tkinter.Label(self, text="  工作服  ", font=guide_font, fg="white", bg=dark_sub)
can_upgrade_work_unifrom_nums_txt.place(x=300, y=421)
can_upgrade_work_unifrom_nums_select = tkinter.ttk.Combobox(self, width=13,
                                                            values=txt_can_upgrade_work_unifrom_nums)
can_upgrade_work_unifrom_nums_select.set(txt_can_upgrade_work_unifrom_nums[0])
can_upgrade_work_unifrom_nums_select.place(x=390 - 17, y=421)

transfer_equip_txt = tkinter.Label(self, text="  跨界  ", font=guide_font, fg="white", bg=dark_sub)
transfer_equip_txt.place(x=300, y=447)
transfer_equip_combopicker = Combopicker(self, values=get_other_account_names(), entrywidth=11)
transfer_equip_combopicker.place(x=390 - 17, y=447)

can_transfer_nums_select = tkinter.ttk.Combobox(self, width=2, values=txt_can_transfer_nums)
can_transfer_nums_select.set(txt_can_transfer_nums[0])
can_transfer_nums_select.place(x=457, y=447)

show_count = tkinter.Label(self, font=guide_font, fg="white", bg=dark_sub)
show_count.place(x=490, y=40)
showcon = show_count.configure
show_state = tkinter.Label(self, text="计算栏", font=guide_font, fg="white", bg=dark_sub)
show_state.place(x=490, y=20)
showsta = show_state.configure

display_realtime_counting_info_label = tkinter.Label(self, font=guide_font, fg="white", bg=dark_sub)
display_realtime_counting_info_label.place(x=430, y=480)
showcon2 = display_realtime_counting_info_label.configure

set101 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['101'],
                        command=lambda: click_set(101));
set101.place(x=29, y=100)
set102 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['102'],
                        command=lambda: click_set(102));
set102.place(x=29, y=130)
set103 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['103'],
                        command=lambda: click_set(103));
set103.place(x=29, y=160)
set104 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['104'],
                        command=lambda: click_set(104));
set104.place(x=29, y=190)
set105 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['105'],
                        command=lambda: click_set(105));
set105.place(x=29, y=220)
set106 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['106'],
                        command=lambda: click_set(106));
set106.place(x=29, y=250)
set107 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['107'],
                        command=lambda: click_set(107));
set107.place(x=29, y=280)
set108 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['108'],
                        command=lambda: click_set(108));
set108.place(x=29, y=310)
set109 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['109'],
                        command=lambda: click_set(109));
set109.place(x=29, y=340)
set110 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['110'],
                        command=lambda: click_set(110));
set110.place(x=29, y=370)
set111 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['111'],
                        command=lambda: click_set(111));
set111.place(x=29, y=400)
set112 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['112'],
                        command=lambda: click_set(112));
set112.place(x=29, y=430)
set113 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['113'],
                        command=lambda: click_set(113));
set113.place(x=29, y=460)
set114 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['114'],
                        command=lambda: click_set(114));
set114.place(x=29, y=490)
set115 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['115'],
                        command=lambda: click_set(115));
set115.place(x=29, y=520)  ##
set116 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['116'],
                        command=lambda: click_set(116));
set116.place(x=320 - 33, y=100)
set117 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['117'],
                        command=lambda: click_set(117));
set117.place(x=320 - 33, y=130)
set118 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['118'],
                        command=lambda: click_set(118));
set118.place(x=320 - 33, y=160)
set119 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['119'],
                        command=lambda: click_set(119));
set119.place(x=320 - 33, y=190)  ##
set120 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['120'],
                        command=lambda: click_set(120));
set120.place(x=500 - 17, y=100)
set121 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['121'],
                        command=lambda: click_set(121));
set121.place(x=500 - 17, y=130)
set122 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['122'],
                        command=lambda: click_set(122));
set122.place(x=500 - 17, y=160)
set123 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['123'],
                        command=lambda: click_set(123));
set123.place(x=500 - 17, y=190)  ##
set128 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['128'],
                        command=lambda: click_set(128));
set128.place(x=29, y=570)
set129 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['129'],
                        command=lambda: click_set(129));
set129.place(x=29, y=600)
set130 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['130'],
                        command=lambda: click_set(130));
set130.place(x=29, y=630)
set131 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['131'],
                        command=lambda: click_set(131));
set131.place(x=29, y=660)  ##
set124 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['124'],
                        command=lambda: click_set(124));
set124.place(x=225, y=570)
set125 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['125'],
                        command=lambda: click_set(125));
set125.place(x=225, y=600)
set126 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['126'],
                        command=lambda: click_set(126));
set126.place(x=225, y=630)
set127 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['127'],
                        command=lambda: click_set(127));
set127.place(x=225, y=660)  ##
set132 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['132'],
                        command=lambda: click_set(132));
set132.place(x=421, y=570)
set133 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['133'],
                        command=lambda: click_set(133));
set133.place(x=421, y=600)
set134 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['134'],
                        command=lambda: click_set(134));
set134.place(x=421, y=630)
set135 = tkinter.Button(self, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2['135'],
                        command=lambda: click_set(135));
set135.place(x=421, y=660)  ##

##지혜의 산물
know_image = PhotoImage(file="set_name/know_name.png")
tkinter.Label(self, bg=dark_main, image=know_image).place(x=302, y=520)
select_item['tg13390150'] = 0;
select_13390150 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                                 image=image_list2['13390150'], command=lambda: click_equipment(13390150))
select_13390150.place(x=403, y=520)
select_item['tg22390240'] = 0;
select_22390240 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                                 image=image_list2['22390240'], command=lambda: click_equipment(22390240))
select_22390240.place(x=433, y=520)
select_item['tg23390450'] = 0;
select_23390450 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                                 image=image_list2['23390450'], command=lambda: click_equipment(23390450))
select_23390450.place(x=463, y=520)
select_item['tg33390750'] = 0;
select_33390750 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                                 image=image_list2['33390750'], command=lambda: click_equipment(33390750))
select_33390750.place(x=493, y=520)

select_item['tg21400340'] = 0;
select_21400340 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                                 image=image_list2['21400340'], command=lambda: click_equipment(21400340))
select_21400340.place(x=524, y=520)
select_item['tg31400540'] = 0;
select_31400540 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                                 image=image_list2['31400540'], command=lambda: click_equipment(31400540))
select_31400540.place(x=554, y=520)
select_item['tg32410650'] = 0;
select_32410650 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                                 image=image_list2['32410650'], command=lambda: click_equipment(32410650))
select_32410650.place(x=584, y=520)

##상의
select_item['tg11010'] = 0;
select_11010 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11010'], command=lambda: click_equipment(11010))
select_11010.place(x=100, y=100)
select_item['tg11011'] = 0;
select_11011 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11011'], command=lambda: click_equipment(11011))
select_11011.place(x=130, y=100)
select_item['tg11020'] = 0;
select_11020 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11020'], command=lambda: click_equipment(11020))
select_11020.place(x=100, y=130)
select_item['tg11021'] = 0;
select_11021 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11021'], command=lambda: click_equipment(11021))
select_11021.place(x=130, y=130)
select_item['tg11030'] = 0;
select_11030 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11030'], command=lambda: click_equipment(11030))
select_11030.place(x=100, y=160)
select_item['tg11031'] = 0;
select_11031 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11031'], command=lambda: click_equipment(11031))
select_11031.place(x=130, y=160)
select_item['tg11040'] = 0;
select_11040 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11040'], command=lambda: click_equipment(11040))
select_11040.place(x=100, y=190)
select_item['tg11041'] = 0;
select_11041 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11041'], command=lambda: click_equipment(11041))
select_11041.place(x=130, y=190)
select_item['tg11050'] = 0;
select_11050 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11050'], command=lambda: click_equipment(11050))
select_11050.place(x=100, y=220)
select_item['tg11051'] = 0;
select_11051 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11051'], command=lambda: click_equipment(11051))
select_11051.place(x=130, y=220)
select_item['tg11060'] = 0;
select_11060 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11060'], command=lambda: click_equipment(11060))
select_11060.place(x=100, y=250)
select_item['tg11061'] = 0;
select_11061 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11061'], command=lambda: click_equipment(11061))
select_11061.place(x=130, y=250)
select_item['tg11070'] = 0;
select_11070 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11070'], command=lambda: click_equipment(11070))
select_11070.place(x=100, y=280)
select_item['tg11071'] = 0;
select_11071 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11071'], command=lambda: click_equipment(11071))
select_11071.place(x=130, y=280)
select_item['tg11080'] = 0;
select_11080 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11080'], command=lambda: click_equipment(11080))
select_11080.place(x=100, y=310)
select_item['tg11081'] = 0;
select_11081 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11081'], command=lambda: click_equipment(11081))
select_11081.place(x=130, y=310)
select_item['tg11090'] = 0;
select_11090 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11090'], command=lambda: click_equipment(11090))
select_11090.place(x=100, y=340)
select_item['tg11091'] = 0;
select_11091 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11091'], command=lambda: click_equipment(11091))
select_11091.place(x=130, y=340)
select_item['tg11100'] = 0;
select_11100 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11100'], command=lambda: click_equipment(11100))
select_11100.place(x=100, y=370)
select_item['tg11101'] = 0;
select_11101 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11101'], command=lambda: click_equipment(11101))
select_11101.place(x=130, y=370)
select_item['tg11110'] = 0;
select_11110 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11110'], command=lambda: click_equipment(11110))
select_11110.place(x=100, y=400)
select_item['tg11111'] = 0;
select_11111 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11111'], command=lambda: click_equipment(11111))
select_11111.place(x=130, y=400)
select_item['tg11120'] = 0;
select_11120 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11120'], command=lambda: click_equipment(11120))
select_11120.place(x=100, y=430)
select_item['tg11121'] = 0;
select_11121 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11121'], command=lambda: click_equipment(11121))
select_11121.place(x=130, y=430)
select_item['tg11130'] = 0;
select_11130 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11130'], command=lambda: click_equipment(11130))
select_11130.place(x=100, y=460)
select_item['tg11131'] = 0;
select_11131 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11131'], command=lambda: click_equipment(11131))
select_11131.place(x=130, y=460)
select_item['tg11140'] = 0;
select_11140 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11140'], command=lambda: click_equipment(11140))
select_11140.place(x=100, y=490)
select_item['tg11141'] = 0;
select_11141 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11141'], command=lambda: click_equipment(11141))
select_11141.place(x=130, y=490)
select_item['tg11150'] = 0;
select_11150 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11150'], command=lambda: click_equipment(11150))
select_11150.place(x=100, y=520)
select_item['tg11151'] = 0;
select_11151 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11151'], command=lambda: click_equipment(11151))
select_11151.place(x=130, y=520)

select_item['tg11280'] = 0;
select_11280 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11280'], command=lambda: click_equipment(11280))
select_11280.place(x=100, y=570)
select_item['tg11281'] = 0;
select_11281 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11281'], command=lambda: click_equipment(11281))
select_11281.place(x=130, y=570)
select_item['tg11290'] = 0;
select_11290 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11290'], command=lambda: click_equipment(11290))
select_11290.place(x=100, y=600)
select_item['tg11291'] = 0;
select_11291 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11291'], command=lambda: click_equipment(11291))
select_11291.place(x=130, y=600)
select_item['tg11300'] = 0;
select_11300 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11300'], command=lambda: click_equipment(11300))
select_11300.place(x=100, y=630)
select_item['tg11301'] = 0;
select_11301 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11301'], command=lambda: click_equipment(11301))
select_11301.place(x=130, y=630)
select_item['tg11310'] = 0;
select_11310 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11310'], command=lambda: click_equipment(11310))
select_11310.place(x=100, y=660)
select_item['tg11311'] = 0;
select_11311 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['11311'], command=lambda: click_equipment(11311))
select_11311.place(x=130, y=660)
##하의
select_item['tg12010'] = 0;
select_12010 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12010'], command=lambda: click_equipment(12010))
select_12010.place(x=161, y=100)
select_item['tg12020'] = 0;
select_12020 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12020'], command=lambda: click_equipment(12020))
select_12020.place(x=161, y=130)
select_item['tg12030'] = 0;
select_12030 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12030'], command=lambda: click_equipment(12030))
select_12030.place(x=161, y=160)
select_item['tg12040'] = 0;
select_12040 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12040'], command=lambda: click_equipment(12040))
select_12040.place(x=161, y=190)
select_item['tg12050'] = 0;
select_12050 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12050'], command=lambda: click_equipment(12050))
select_12050.place(x=161, y=220)
select_item['tg12060'] = 0;
select_12060 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12060'], command=lambda: click_equipment(12060))
select_12060.place(x=161, y=250)
select_item['tg12070'] = 0;
select_12070 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12070'], command=lambda: click_equipment(12070))
select_12070.place(x=161, y=280)
select_item['tg12080'] = 0;
select_12080 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12080'], command=lambda: click_equipment(12080))
select_12080.place(x=161, y=310)
select_item['tg12090'] = 0;
select_12090 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12090'], command=lambda: click_equipment(12090))
select_12090.place(x=161, y=340)
select_item['tg12100'] = 0;
select_12100 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12100'], command=lambda: click_equipment(12100))
select_12100.place(x=161, y=370)
select_item['tg12110'] = 0;
select_12110 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12110'], command=lambda: click_equipment(12110))
select_12110.place(x=161, y=400)
select_item['tg12120'] = 0;
select_12120 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12120'], command=lambda: click_equipment(12120))
select_12120.place(x=161, y=430)
select_item['tg12130'] = 0;
select_12130 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12130'], command=lambda: click_equipment(12130))
select_12130.place(x=161, y=460)
select_item['tg12140'] = 0;
select_12140 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12140'], command=lambda: click_equipment(12140))
select_12140.place(x=161, y=490)
select_item['tg12150'] = 0;
select_12150 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12150'], command=lambda: click_equipment(12150))
select_12150.place(x=161, y=520)
select_item['tg12240'] = 0;
select_12240 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12240'], command=lambda: click_equipment(12240))
select_12240.place(x=296, y=570)
select_item['tg12250'] = 0;
select_12250 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12250'], command=lambda: click_equipment(12250))
select_12250.place(x=296, y=600)
select_item['tg12260'] = 0;
select_12260 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12260'], command=lambda: click_equipment(12260))
select_12260.place(x=296, y=630)
select_item['tg12270'] = 0;
select_12270 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['12270'], command=lambda: click_equipment(12270))
select_12270.place(x=296, y=660)
##어깨
select_item['tg13010'] = 0;
select_13010 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['13010'], command=lambda: click_equipment(13010))
select_13010.place(x=192, y=100)
select_item['tg13020'] = 0;
select_13020 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['13020'], command=lambda: click_equipment(13020))
select_13020.place(x=192, y=130)
select_item['tg13030'] = 0;
select_13030 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['13030'], command=lambda: click_equipment(13030))
select_13030.place(x=192, y=160)
select_item['tg13040'] = 0;
select_13040 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['13040'], command=lambda: click_equipment(13040))
select_13040.place(x=192, y=190)
select_item['tg13050'] = 0;
select_13050 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['13050'], command=lambda: click_equipment(13050))
select_13050.place(x=192, y=220)
select_item['tg13060'] = 0;
select_13060 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['13060'], command=lambda: click_equipment(13060))
select_13060.place(x=192, y=250)
select_item['tg13070'] = 0;
select_13070 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['13070'], command=lambda: click_equipment(13070))
select_13070.place(x=192, y=280)
select_item['tg13080'] = 0;
select_13080 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['13080'], command=lambda: click_equipment(13080))
select_13080.place(x=192, y=310)
select_item['tg13090'] = 0;
select_13090 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['13090'], command=lambda: click_equipment(13090))
select_13090.place(x=192, y=340)
select_item['tg13100'] = 0;
select_13100 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['13100'], command=lambda: click_equipment(13100))
select_13100.place(x=192, y=370)
select_item['tg13110'] = 0;
select_13110 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['13110'], command=lambda: click_equipment(13110))
select_13110.place(x=192, y=400)
select_item['tg13120'] = 0;
select_13120 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['13120'], command=lambda: click_equipment(13120))
select_13120.place(x=192, y=430)
select_item['tg13130'] = 0;
select_13130 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['13130'], command=lambda: click_equipment(13130))
select_13130.place(x=192, y=460)
select_item['tg13140'] = 0;
select_13140 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['13140'], command=lambda: click_equipment(13140))
select_13140.place(x=192, y=490)
select_item['tg13150'] = 0;
select_13150 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['13150'], command=lambda: click_equipment(13150))
select_13150.place(x=192, y=520)
##벨트
select_item['tg14010'] = 0;
select_14010 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['14010'], command=lambda: click_equipment(14010))
select_14010.place(x=223, y=100)
select_item['tg14020'] = 0;
select_14020 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['14020'], command=lambda: click_equipment(14020))
select_14020.place(x=223, y=130)
select_item['tg14030'] = 0;
select_14030 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['14030'], command=lambda: click_equipment(14030))
select_14030.place(x=223, y=160)
select_item['tg14040'] = 0;
select_14040 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['14040'], command=lambda: click_equipment(14040))
select_14040.place(x=223, y=190)
select_item['tg14050'] = 0;
select_14050 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['14050'], command=lambda: click_equipment(14050))
select_14050.place(x=223, y=220)
select_item['tg14060'] = 0;
select_14060 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['14060'], command=lambda: click_equipment(14060))
select_14060.place(x=223, y=250)
select_item['tg14070'] = 0;
select_14070 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['14070'], command=lambda: click_equipment(14070))
select_14070.place(x=223, y=280)
select_item['tg14080'] = 0;
select_14080 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['14080'], command=lambda: click_equipment(14080))
select_14080.place(x=223, y=310)
select_item['tg14090'] = 0;
select_14090 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['14090'], command=lambda: click_equipment(14090))
select_14090.place(x=223, y=340)
select_item['tg14100'] = 0;
select_14100 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['14100'], command=lambda: click_equipment(14100))
select_14100.place(x=223, y=370)
select_item['tg14110'] = 0;
select_14110 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['14110'], command=lambda: click_equipment(14110))
select_14110.place(x=223, y=400)
select_item['tg14120'] = 0;
select_14120 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['14120'], command=lambda: click_equipment(14120))
select_14120.place(x=223, y=430)
select_item['tg14130'] = 0;
select_14130 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['14130'], command=lambda: click_equipment(14130))
select_14130.place(x=223, y=460)
select_item['tg14140'] = 0;
select_14140 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['14140'], command=lambda: click_equipment(14140))
select_14140.place(x=223, y=490)
select_item['tg14150'] = 0;
select_14150 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['14150'], command=lambda: click_equipment(14150))
select_14150.place(x=223, y=520)
##신발
select_item['tg15010'] = 0;
select_15010 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15010'], command=lambda: click_equipment(15010))
select_15010.place(x=254, y=100)
select_item['tg15020'] = 0;
select_15020 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15020'], command=lambda: click_equipment(15020))
select_15020.place(x=254, y=130)
select_item['tg15030'] = 0;
select_15030 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15030'], command=lambda: click_equipment(15030))
select_15030.place(x=254, y=160)
select_item['tg15040'] = 0;
select_15040 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15040'], command=lambda: click_equipment(15040))
select_15040.place(x=254, y=190)
select_item['tg15050'] = 0;
select_15050 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15050'], command=lambda: click_equipment(15050))
select_15050.place(x=254, y=220)
select_item['tg15060'] = 0;
select_15060 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15060'], command=lambda: click_equipment(15060))
select_15060.place(x=254, y=250)
select_item['tg15070'] = 0;
select_15070 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15070'], command=lambda: click_equipment(15070))
select_15070.place(x=254, y=280)
select_item['tg15080'] = 0;
select_15080 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15080'], command=lambda: click_equipment(15080))
select_15080.place(x=254, y=310)
select_item['tg15090'] = 0;
select_15090 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15090'], command=lambda: click_equipment(15090))
select_15090.place(x=254, y=340)
select_item['tg15100'] = 0;
select_15100 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15100'], command=lambda: click_equipment(15100))
select_15100.place(x=254, y=370)
select_item['tg15110'] = 0;
select_15110 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15110'], command=lambda: click_equipment(15110))
select_15110.place(x=254, y=400)
select_item['tg15120'] = 0;
select_15120 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15120'], command=lambda: click_equipment(15120))
select_15120.place(x=254, y=430)
select_item['tg15130'] = 0;
select_15130 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15130'], command=lambda: click_equipment(15130))
select_15130.place(x=254, y=460)
select_item['tg15140'] = 0;
select_15140 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15140'], command=lambda: click_equipment(15140))
select_15140.place(x=254, y=490)
select_item['tg15150'] = 0;
select_15150 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15150'], command=lambda: click_equipment(15150))
select_15150.place(x=254, y=520)
select_item['tg15320'] = 0;
select_15320 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15320'], command=lambda: click_equipment(15320))
select_15320.place(x=492, y=570)
select_item['tg15330'] = 0;
select_15330 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15330'], command=lambda: click_equipment(15330))
select_15330.place(x=492, y=600)
select_item['tg15340'] = 0;
select_15340 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15340'], command=lambda: click_equipment(15340))
select_15340.place(x=492, y=630)
select_item['tg15350'] = 0;
select_15350 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['15350'], command=lambda: click_equipment(15350))
select_15350.place(x=492, y=660)
##팔찌
select_item['tg21160'] = 0;
select_21160 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21160'], command=lambda: click_equipment(21160))
select_21160.place(x=370 - 12, y=100)
select_item['tg21161'] = 0;
select_21161 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21161'], command=lambda: click_equipment(21161))
select_21161.place(x=370 - 12 + 30, y=100)
select_item['tg21170'] = 0;
select_21170 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21170'], command=lambda: click_equipment(21170))
select_21170.place(x=370 - 12, y=130)
select_item['tg21171'] = 0;
select_21171 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21171'], command=lambda: click_equipment(21171))
select_21171.place(x=370 - 12 + 30, y=130)
select_item['tg21180'] = 0;
select_21180 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21180'], command=lambda: click_equipment(21180))
select_21180.place(x=370 - 12, y=160)
select_item['tg21181'] = 0;
select_21181 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21181'], command=lambda: click_equipment(21181))
select_21181.place(x=370 - 12 + 30, y=160)
select_item['tg21190'] = 0;
select_21190 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21190'], command=lambda: click_equipment(21190))
select_21190.place(x=370 - 12, y=190)
select_item['tg21191'] = 0;
select_21191 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21191'], command=lambda: click_equipment(21191))
select_21191.place(x=370 - 12 + 30, y=190)
select_item['tg21240'] = 0;
select_21240 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21240'], command=lambda: click_equipment(21240))
select_21240.place(x=327, y=570)
select_item['tg21241'] = 0;
select_21241 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21241'], command=lambda: click_equipment(21241))
select_21241.place(x=357, y=570)
select_item['tg21250'] = 0;
select_21250 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21250'], command=lambda: click_equipment(21250))
select_21250.place(x=327, y=600)
select_item['tg21251'] = 0;
select_21251 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21251'], command=lambda: click_equipment(21251))
select_21251.place(x=357, y=600)
select_item['tg21260'] = 0;
select_21260 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21260'], command=lambda: click_equipment(21260))
select_21260.place(x=327, y=630)
select_item['tg21261'] = 0;
select_21261 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21261'], command=lambda: click_equipment(21261))
select_21261.place(x=357, y=630)
select_item['tg21270'] = 0;
select_21270 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21270'], command=lambda: click_equipment(21270))
select_21270.place(x=327, y=660)
select_item['tg21271'] = 0;
select_21271 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['21271'], command=lambda: click_equipment(21271))
select_21271.place(x=357, y=660)
##목걸이
select_item['tg22160'] = 0;
select_22160 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['22160'], command=lambda: click_equipment(22160))
select_22160.place(x=419, y=100)
select_item['tg22170'] = 0;
select_22170 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['22170'], command=lambda: click_equipment(22170))
select_22170.place(x=419, y=130)
select_item['tg22180'] = 0;
select_22180 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['22180'], command=lambda: click_equipment(22180))
select_22180.place(x=419, y=160)
select_item['tg22190'] = 0;
select_22190 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['22190'], command=lambda: click_equipment(22190))
select_22190.place(x=419, y=190)
select_item['tg22280'] = 0;
select_22280 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['22280'], command=lambda: click_equipment(22280))
select_22280.place(x=161, y=570)
select_item['tg22290'] = 0;
select_22290 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['22290'], command=lambda: click_equipment(22290))
select_22290.place(x=161, y=600)
select_item['tg22300'] = 0;
select_22300 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['22300'], command=lambda: click_equipment(22300))
select_22300.place(x=161, y=630)
select_item['tg22310'] = 0;
select_22310 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['22310'], command=lambda: click_equipment(22310))
select_22310.place(x=161, y=660)
##반지
select_item['tg23160'] = 0;
select_23160 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['23160'], command=lambda: click_equipment(23160))
select_23160.place(x=450, y=100)
select_item['tg23170'] = 0;
select_23170 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['23170'], command=lambda: click_equipment(23170))
select_23170.place(x=450, y=130)
select_item['tg23180'] = 0;
select_23180 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['23180'], command=lambda: click_equipment(23180))
select_23180.place(x=450, y=160)
select_item['tg23190'] = 0;
select_23190 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['23190'], command=lambda: click_equipment(23190))
select_23190.place(x=450, y=190)
select_item['tg23320'] = 0;
select_23320 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['23320'], command=lambda: click_equipment(23320))
select_23320.place(x=523, y=570)
select_item['tg23330'] = 0;
select_23330 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['23330'], command=lambda: click_equipment(23330))
select_23330.place(x=523, y=600)
select_item['tg23340'] = 0;
select_23340 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['23340'], command=lambda: click_equipment(23340))
select_23340.place(x=523, y=630)
select_item['tg23350'] = 0;
select_23350 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['23350'], command=lambda: click_equipment(23350))
select_23350.place(x=523, y=660)
##보조장비
select_item['tg31200'] = 0;
select_31200 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['31200'], command=lambda: click_equipment(31200))
select_31200.place(x=554, y=100)
select_item['tg31210'] = 0;
select_31210 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['31210'], command=lambda: click_equipment(31210))
select_31210.place(x=554, y=130)
select_item['tg31220'] = 0;
select_31220 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['31220'], command=lambda: click_equipment(31220))
select_31220.place(x=554, y=160)
select_item['tg31230'] = 0;
select_31230 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['31230'], command=lambda: click_equipment(31230))
select_31230.place(x=554, y=190)
select_item['tg31280'] = 0;
select_31280 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['31280'], command=lambda: click_equipment(31280))
select_31280.place(x=192, y=570)
select_item['tg31290'] = 0;
select_31290 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['31290'], command=lambda: click_equipment(31290))
select_31290.place(x=192, y=600)
select_item['tg31300'] = 0;
select_31300 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['31300'], command=lambda: click_equipment(31300))
select_31300.place(x=192, y=630)
select_item['tg31310'] = 0;
select_31310 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['31310'], command=lambda: click_equipment(31310))
select_31310.place(x=192, y=660)
##마법석
select_item['tg32200'] = 0;
select_32200 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['32200'], command=lambda: click_equipment(32200))
select_32200.place(x=585, y=100)
select_item['tg32210'] = 0;
select_32210 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['32210'], command=lambda: click_equipment(32210))
select_32210.place(x=585, y=130)
select_item['tg32220'] = 0;
select_32220 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['32220'], command=lambda: click_equipment(32220))
select_32220.place(x=585, y=160)
select_item['tg32230'] = 0;
select_32230 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['32230'], command=lambda: click_equipment(32230))
select_32230.place(x=585, y=190)
select_item['tg32240'] = 0;
select_32240 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['32240'], command=lambda: click_equipment(32240))
select_32240.place(x=388, y=570)
select_item['tg32250'] = 0;
select_32250 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['32250'], command=lambda: click_equipment(32250))
select_32250.place(x=388, y=600)
select_item['tg32260'] = 0;
select_32260 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['32260'], command=lambda: click_equipment(32260))
select_32260.place(x=388, y=630)
select_item['tg32270'] = 0;
select_32270 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['32270'], command=lambda: click_equipment(32270))
select_32270.place(x=388, y=660)
##귀걸이
select_item['tg33200'] = 0;
select_33200 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33200'], command=lambda: click_equipment(33200))
select_33200.place(x=616, y=100)
select_item['tg33201'] = 0;
select_33201 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33201'], command=lambda: click_equipment(33201))
select_33201.place(x=646, y=100)
select_item['tg33210'] = 0;
select_33210 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33210'], command=lambda: click_equipment(33210))
select_33210.place(x=616, y=130)
select_item['tg33211'] = 0;
select_33211 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33211'], command=lambda: click_equipment(33211))
select_33211.place(x=646, y=130)
select_item['tg33220'] = 0;
select_33220 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33220'], command=lambda: click_equipment(33220))
select_33220.place(x=616, y=160)
select_item['tg33221'] = 0;
select_33221 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33221'], command=lambda: click_equipment(33221))
select_33221.place(x=646, y=160)
select_item['tg33230'] = 0;
select_33230 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33230'], command=lambda: click_equipment(33230))
select_33230.place(x=616, y=190)
select_item['tg33231'] = 0;
select_33231 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33231'], command=lambda: click_equipment(33231))
select_33231.place(x=646, y=190)
select_item['tg33320'] = 0;
select_33320 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33320'], command=lambda: click_equipment(33320))
select_33320.place(x=554, y=570)
select_item['tg33321'] = 0;
select_33321 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33321'], command=lambda: click_equipment(33321))
select_33321.place(x=584, y=570)
select_item['tg33330'] = 0;
select_33330 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33330'], command=lambda: click_equipment(33330))
select_33330.place(x=554, y=600)
select_item['tg33331'] = 0;
select_33331 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33331'], command=lambda: click_equipment(33331))
select_33331.place(x=584, y=600)
select_item['tg33340'] = 0;
select_33340 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33340'], command=lambda: click_equipment(33340))
select_33340.place(x=554, y=630)
select_item['tg33341'] = 0;
select_33341 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33341'], command=lambda: click_equipment(33341))
select_33341.place(x=584, y=630)
select_item['tg33350'] = 0;
select_33350 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33350'], command=lambda: click_equipment(33350))
select_33350.place(x=554, y=660)
select_item['tg33351'] = 0;
select_33351 = tkinter.Button(self, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main,
                              image=image_list2['33351'], command=lambda: click_equipment(33351))
select_33351.place(x=584, y=660)

donate_image = PhotoImage(file='ext_img/donate.png')
donate_bt = tkinter.Button(self, image=donate_image, command=donate, borderwidth=0, bg=dark_main,
                           activebackground=dark_main)
donate_bt.place(x=625, y=550 - 28)

dunfaoff_image = PhotoImage(file='ext_img/dunfaoff.png')
dunfaoff_url = tkinter.Button(self, image=dunfaoff_image, command=dunfaoff, borderwidth=0, bg=dark_main,
                              activebackground=dark_main)
dunfaoff_url.place(x=535, y=410)

blog_image = PhotoImage(file='ext_img/blog.png')
blog_url = tkinter.Button(self, image=blog_image, command=blog, borderwidth=0, bg=dark_main,
                          activebackground=dark_main)
blog_url.place(x=615, y=410)

maker_image = PhotoImage(file='ext_img/maker.png')
maker = tkinter.Button(self, image=maker_image, command=hamjung, borderwidth=0, bg=dark_main,
                       activebackground=dark_main)
version = tkinter.Label(self, text='V ' + str(now_version) + '\n' + ver_time, font=guide_font, fg="white",
                        bg=dark_main)

maker.place(x=625, y=590)
version.place(x=630, y=650)

###########################################################
#                 启动工作线程并进入ui主循环                #
###########################################################

if __name__ == "__main__":
    update_thread()

self.mainloop()
self.quit()
